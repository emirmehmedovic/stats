#!/usr/bin/env python3
"""
Test Excel creation to verify openpyxl works correctly
"""
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import re

def sanitize_text(value):
    """Remove control characters that can corrupt Excel XML."""
    if value is None:
        return None
    if not isinstance(value, str):
        return value
    try:
        from openpyxl.utils.cell import ILLEGAL_CHARACTERS_RE
        value = ILLEGAL_CHARACTERS_RE.sub('', value)
    except Exception:
        value = re.sub(r'[\x00-\x08\x0B\x0C\x0E-\x1F]', '', value)
    return value

def create_merged_cell(ws, row, start_col, end_col, value, font=None, fill=None, alignment=None, border=None):
    """Create merged cell with proper formatting."""
    if isinstance(start_col, str):
        start_col = openpyxl.utils.column_index_from_string(start_col)
    if isinstance(end_col, str):
        end_col = openpyxl.utils.column_index_from_string(end_col)

    for col in range(start_col, end_col + 1):
        cell = ws.cell(row=row, column=col)
        if col == start_col and value is not None:
            cell.value = sanitize_text(value) if isinstance(value, str) else value
        if font:
            cell.font = font
        if fill:
            cell.fill = fill
        if alignment:
            cell.alignment = alignment
        if border:
            cell.border = border

    if start_col != end_col:
        ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=end_col)

# Test 1: Create workbook from scratch
print("Test 1: Creating Excel from scratch...")
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Test"

# Add title with merged cells
title_font = Font(name='Arial', size=14, bold=True)
center_align = Alignment(horizontal='center', vertical='center')
header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

create_merged_cell(ws, 1, 'A', 'E', "Test Title", font=title_font, alignment=center_align)

# Add headers
headers = ['Col A', 'Col B', 'Col C', 'Col D', 'Col E']
for col_idx, header in enumerate(headers, start=1):
    cell = ws.cell(row=2, column=col_idx)
    cell.value = sanitize_text(header)
    cell.font = Font(bold=True)
    cell.fill = header_fill
    cell.alignment = center_align

# Add some data
for row_idx in range(3, 10):
    for col_idx in range(1, 6):
        ws.cell(row=row_idx, column=col_idx).value = sanitize_text(f"Data {row_idx}-{col_idx}")

# Save
wb.save("izvještaji/generated/test_from_scratch.xlsx")
print("✅ Saved: izvještaji/generated/test_from_scratch.xlsx")

# Test 2: Load template and modify
print("\nTest 2: Loading template and modifying...")
try:
    wb2 = openpyxl.load_workbook("izvještaji/09. BHDCA Septembar 2025.xlsx")
    ws2 = wb2['Sheet2']

    # Just set a simple value without touching merged cells
    ws2['A1'] = sanitize_text("Test modification")

    wb2.save("izvještaji/generated/test_template_modified.xlsx")
    print("✅ Saved: izvještaji/generated/test_template_modified.xlsx")
except Exception as e:
    print(f"❌ Error: {e}")

print("\nDone! Please try opening both files in Excel.")
