#!/usr/bin/env python3
"""
Minimal BHANSA report test - with hardcoded safe data
This will help identify if the problem is in data or in code
"""
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from pathlib import Path

OUTPUT_DIR = Path("izvještaji/generated")
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

# Test 1: Absolute minimal Excel file
print("Test 1: Creating absolute minimal Excel...")
wb1 = openpyxl.Workbook()
ws1 = wb1.active
ws1.title = "Test"
ws1['A1'] = "Test"
ws1['B1'] = 123
wb1.iso_dates = True
wb1.save(OUTPUT_DIR / "test_minimal.xlsx")
wb1.close()
print("✅ Saved: test_minimal.xlsx")

# Test 2: With merged cells (like BHANSA header)
print("\nTest 2: Creating with merged cells...")
wb2 = openpyxl.Workbook()
ws2 = wb2.active
ws2.title = "Test"

title_font = Font(name='Arial', size=14, bold=True)
center_align = Alignment(horizontal='center', vertical='center')

# Merge cells A1:N1
for col in range(1, 15):
    cell = ws2.cell(row=1, column=col)
    if col == 1:
        cell.value = "AERODROM TUZLA RIJD Januar 2026"
    cell.font = title_font
    cell.alignment = center_align

ws2.merge_cells('A1:N1')

# Add some headers
headers = ['Nr.', 'DATE', 'COMPANY', 'A/C', 'REG.', 'MTOW/T', 'FLT.NMB',
           'FROM', 'TO', 'ADDRESS', 'DEP PAX', 'ARR PAX', 'Trip fare', 'Day']

header_font = Font(bold=True)
header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

for col_idx, header in enumerate(headers, start=1):
    cell = ws2.cell(row=2, column=col_idx)
    cell.value = header
    cell.font = header_font
    cell.fill = header_fill
    cell.border = border
    cell.alignment = center_align

# Add one row of safe data
ws2.cell(row=3, column=1).value = 1
ws2.cell(row=3, column=2).value = "01/01/2026"
ws2.cell(row=3, column=3).value = "Wizz Air"
ws2.cell(row=3, column=4).value = "A320"
ws2.cell(row=3, column=5).value = "9H-WDK"
ws2.cell(row=3, column=6).value = 75.5
ws2.cell(row=3, column=7).value = "W64297"
ws2.cell(row=3, column=8).value = "TZL"
ws2.cell(row=3, column=9).value = "DTM"
ws2.cell(row=3, column=10).value = "Wizz Air Hungary Ltd"
ws2.cell(row=3, column=11).value = "215"
ws2.cell(row=3, column=12).value = ""
ws2.cell(row=3, column=13).value = "R"
ws2.cell(row=3, column=14).value = "Monday"

wb2.iso_dates = True
wb2.save(OUTPUT_DIR / "test_with_merged.xlsx")
wb2.close()
print("✅ Saved: test_with_merged.xlsx")

# Test 3: Exact same structure as BHANSA but minimal data
print("\nTest 3: Creating BHANSA-like structure...")
wb3 = openpyxl.Workbook()
ws3 = wb3.active
ws3.title = "Sheet1"

# Stilovi
title_font = Font(name='Arial', size=14, bold=True)
header_font = Font(name='Arial', size=11, bold=True)
header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
regular_font = Font(name='Arial', size=10)
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
center_align = Alignment(horizontal='center', vertical='center')
left_align = Alignment(horizontal='left', vertical='center')

# Header - Row 1 (merged)
for col in range(1, 15):
    cell = ws3.cell(row=1, column=col)
    if col == 1:
        cell.value = "AERODROM TUZLA                    RIJD JANUAR 2026"
    cell.font = title_font
    cell.alignment = center_align

ws3.merge_cells(start_row=1, start_column=1, end_row=1, end_column=14)

# Column Headers - Row 2
headers = ['Nr.', 'DATE', 'COMPANY', 'A/C', 'REG.', 'MTOW/T', 'FLT.NMB',
           'FROM', 'TO', 'ADDRESS', 'DEP PAX', 'ARR PAX', 'Trip fare', 'Day']

for col_idx, header in enumerate(headers, start=1):
    cell = ws3.cell(row=2, column=col_idx)
    cell.value = header
    cell.font = header_font
    cell.fill = header_fill
    cell.border = border
    cell.alignment = center_align

# Add 5 rows of safe data
for row_idx in range(3, 8):
    ws3.cell(row=row_idx, column=1).value = row_idx - 2
    ws3.cell(row=row_idx, column=2).value = "01/01/2026"
    ws3.cell(row=row_idx, column=3).value = "Wizz Air Hungary Ltd"
    ws3.cell(row=row_idx, column=4).value = "A320"
    ws3.cell(row=row_idx, column=5).value = "9H-WDK"
    ws3.cell(row=row_idx, column=6).value = 75.5
    ws3.cell(row=row_idx, column=7).value = f"W6429{row_idx}"
    ws3.cell(row=row_idx, column=8).value = "TZL"
    ws3.cell(row=row_idx, column=9).value = "DTM"
    ws3.cell(row=row_idx, column=10).value = "Wizz Air Hungary Ltd, Budapest, Hungary"
    ws3.cell(row=row_idx, column=11).value = "215"
    ws3.cell(row=row_idx, column=12).value = ""
    ws3.cell(row=row_idx, column=13).value = "R"
    ws3.cell(row=row_idx, column=14).value = "Monday"

    # Apply styles
    for col in range(1, 15):
        cell = ws3.cell(row=row_idx, column=col)
        cell.font = regular_font
        cell.border = border
        if col in [1, 2, 4, 5, 6, 7, 8, 9, 13, 14]:
            cell.alignment = center_align
        else:
            cell.alignment = left_align

# Set column widths
ws3.column_dimensions['A'].width = 5
ws3.column_dimensions['B'].width = 12
ws3.column_dimensions['C'].width = 25
ws3.column_dimensions['D'].width = 8
ws3.column_dimensions['E'].width = 10
ws3.column_dimensions['F'].width = 10
ws3.column_dimensions['G'].width = 12
ws3.column_dimensions['H'].width = 8
ws3.column_dimensions['I'].width = 8
ws3.column_dimensions['J'].width = 50
ws3.column_dimensions['K'].width = 12
ws3.column_dimensions['L'].width = 12
ws3.column_dimensions['M'].width = 10
ws3.column_dimensions['N'].width = 12

wb3.iso_dates = True
wb3.save(OUTPUT_DIR / "test_bhansa_structure.xlsx")
wb3.close()
print("✅ Saved: test_bhansa_structure.xlsx")

print("\n" + "="*60)
print("MOLIM VAS:")
print("1. Otvorite sve 3 fajla u Excel-u")
print("2. Javite da li bilo koji od njih daje corruption warning")
print("3. Ako SVIH 3 rade OK, problem je u PODACIMA iz baze")
print("4. Ako neki ne radi, problem je u strukturi Excel fajla")
print("="*60)
