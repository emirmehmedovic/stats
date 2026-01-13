#!/usr/bin/env python3
"""
Test: Add workbook properties to prevent Excel warnings
"""
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from pathlib import Path
from datetime import datetime

OUTPUT_DIR = Path("izvještaji/generated")
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

print("Creating Excel with full workbook properties...")
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Test Report"

# Set workbook properties
wb.properties.creator = "Tuzla International Airport"
wb.properties.title = "Monthly Report"
wb.properties.subject = "Airport Statistics"
wb.properties.description = "Generated report"
wb.properties.keywords = "airport, statistics, report"
wb.properties.category = "Reports"
wb.properties.created = datetime.now()
wb.properties.modified = datetime.now()
wb.properties.lastModifiedBy = "Statistics System"

# Title row (merged)
title_font = Font(name='Arial', size=14, bold=True)
center_align = Alignment(horizontal='center', vertical='center')

for col in range(1, 6):
    cell = ws.cell(row=1, column=col)
    if col == 1:
        cell.value = "TEST REPORT WITH PROPERTIES"
    cell.font = title_font
    cell.alignment = center_align
ws.merge_cells('A1:E1')

# Headers
header_font = Font(name='Arial', size=10, bold=True)
header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

headers = ['ID', 'Name', 'Value', 'Date', 'Status']
for col_idx, header in enumerate(headers, start=1):
    cell = ws.cell(row=2, column=col_idx)
    cell.value = header
    cell.font = header_font
    cell.fill = header_fill
    cell.border = border
    cell.alignment = center_align

# Data rows
for row_idx in range(3, 10):
    ws.cell(row=row_idx, column=1).value = row_idx - 2
    ws.cell(row=row_idx, column=2).value = f"Item {row_idx - 2}"
    ws.cell(row=row_idx, column=3).value = 100 + row_idx
    ws.cell(row=row_idx, column=4).value = "2026-01-01"
    ws.cell(row=row_idx, column=5).value = "OK"

# Save with all best practices
wb.iso_dates = True
wb.security.lockStructure = False
wb.security.lockWindows = False

output_path = OUTPUT_DIR / "test_with_properties.xlsx"
wb.save(output_path)
wb.close()

print(f"✅ Created: {output_path}")
print("\nOtvorite ovaj fajl i provjerite da li ima corruption warning!")
