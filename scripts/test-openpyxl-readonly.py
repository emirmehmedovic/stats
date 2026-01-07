#!/usr/bin/env python3

import openpyxl
import os

STATS_DIR = "/Users/emir_mw/stats/STATS/2025/Dnevni izvještaji"
month_path = os.path.join(STATS_DIR, "03. MART")

files = [
    f for f in os.listdir(month_path)
    if 'Dnevni izvještaj o saobraćaju' in f and f.endswith('.xlsx')
]

file_path = os.path.join(month_path, files[0])

print("Testing openpyxl read_only mode:\n")

wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)

# Get header from first sheet (like the extraction script does)
first_sheet = wb[wb.sheetnames[0]]
header_row = next(first_sheet.iter_rows(values_only=True))
print(f"Got header from sheet '{wb.sheetnames[0]}': {len([h for h in header_row if h])} columns\n")

# Now iterate through all sheets
print("Iterating through all sheets:")
for sheet_name in wb.sheetnames[:5]:  # First 5 sheets
    ws = wb[sheet_name]
    row_count = 0

    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if i == 1:
            continue
        if row[0]:
            row_count += 1

    print(f"  Sheet '{sheet_name}': {row_count} data rows")

wb.close()
