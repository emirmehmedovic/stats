#!/usr/bin/env python3

import openpyxl
import os
import sys
import re
from datetime import datetime
from prisma import Prisma
import asyncio

STATS_DIR = "/Users/emir_mw/stats/STATS/2025/Dnevni izvještaji"

# Map Excel operation types to database codes
OPERATION_TYPE_MAP = {
    'SCHEDULED': 'SCHEDULED',
    'DIVERTED': 'DIVERTED',
    'MEDEVAC': 'MEDEVAC',
    'HELICOPTER': 'GENERAL_AVIATION',
    'GA': 'GENERAL_AVIATION',
    'NONSCHEDULED': 'CHARTER',
    'NON-SCHEDULED': 'CHARTER',
    'UNSCHEDULED': 'CHARTER',
    'CHARTER': 'CHARTER',
    'CARGO': 'CARGO',
    'MILITARY': 'MILITARY',
}

def parse_passengers(value):
    """Parse passenger string like '165+6 INF' or '110' or '-'"""
    if not value or value == '-' or value == 'N/A':
        return None

    str_val = str(value).strip()

    # Pattern: "110+6 INF" or "110+6"
    match = re.match(r'^(\d+)\s*\+\s*(\d+)', str_val)
    if match:
        adults = int(match.group(1))
        infants = int(match.group(2))
        return {'adults': adults, 'infants': infants, 'total': adults + infants}

    # Pattern: just a number "110"
    match = re.match(r'^(\d+)$', str_val)
    if match:
        adults = int(match.group(1))
        return {'adults': adults, 'infants': 0, 'total': adults}

    return None

def parse_route(route):
    """Parse route like 'TZL-AYT' and return [departure, arrival]"""
    if not route or route == '-' or route == 'N/A':
        return None

    parts = [p.strip() for p in route.split('-') if p.strip()]

    if len(parts) >= 2:
        return {
            'departure': parts[0],
            'arrival': parts[-1],
        }

    return None

def map_operation_type(excel_type):
    """Map operation type from Excel to database code"""
    if not excel_type:
        return 'SCHEDULED'

    normalized = str(excel_type).strip().upper()
    return OPERATION_TYPE_MAP.get(normalized, 'SCHEDULED')

async def get_or_create_airport(prisma, iata_code):
    """Get or create airport by IATA code"""
    airport = await prisma.airport.find_unique(where={'iataCode': iata_code})

    if airport:
        return airport.id

    # Create new airport with minimal data
    airport = await prisma.airport.create(data={
        'iataCode': iata_code,
        'name': iata_code,  # Placeholder
        'country': 'Unknown',  # Placeholder
    })

    return airport.id

async def import_month(prisma, month_folder, dry_run=False):
    """Import one month of flight data"""
    month_path = os.path.join(STATS_DIR, month_folder)

    # Find Excel file
    files = [
        f for f in os.listdir(month_path)
        if 'Dnevni izvještaj o saobraćaju' in f and f.endswith('.xlsx')
    ]

    if not files:
        print(f"   ⚠️  No Excel file found in {month_folder}")
        return {'success': False, 'count': 0, 'errors': 0}

    file_path = os.path.join(month_path, files[0])

    try:
        print(f"\n📊 {month_folder}:")
        print(f"   File: {files[0]}")

        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        flight_count = 0
        error_count = 0

        # Iterate through all sheets (days)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            for i, row in enumerate(ws.iter_rows(values_only=True), 1):
                # Skip header
                if i == 1:
                    continue

                # Skip empty rows
                if not row[0]:
                    continue

                try:
                    # Extract data from row
                    # [0] datum, [1] kompanija, [2] ruta, [3] tip a/c, [4] reg, [5] tip OPER
                    date_value = row[0]
                    airline_name = row[1]
                    route_str = row[2]
                    aircraft_model = row[3]
                    registration = row[4]
                    operation_type_str = row[5]
                    passengers_str = row[6] if len(row) > 6 else None

                    # Parse date
                    if isinstance(date_value, datetime):
                        flight_date = date_value
                    else:
                        try:
                            flight_date = datetime.strptime(str(date_value), '%Y-%m-%d')
                        except:
                            print(f"   ⚠️  Invalid date: {date_value}")
                            error_count += 1
                            continue

                    # Find airline
                    airline = await prisma.airline.find_first(
                        where={'name': str(airline_name).strip().upper()}
                    )

                    if not airline:
                        print(f"   ⚠️  Airline not found: {airline_name}")
                        error_count += 1
                        continue

                    # Parse route
                    route = parse_route(str(route_str))
                    if not route:
                        print(f"   ⚠️  Invalid route: {route_str}")
                        error_count += 1
                        continue

                    # Get or create airports
                    departure_airport_id = await get_or_create_airport(prisma, route['departure'])
                    arrival_airport_id = await get_or_create_airport(prisma, route['arrival'])

                    # Find aircraft type
                    aircraft_type = await prisma.aircrafttype.find_first(
                        where={'model': str(aircraft_model).strip()}
                    )

                    if not aircraft_type:
                        print(f"   ⚠️  Aircraft type not found: {aircraft_model}")
                        error_count += 1
                        continue

                    # Get operation type
                    operation_type_code = map_operation_type(operation_type_str)
                    operation_type = await prisma.operationtype.find_unique(
                        where={'code': operation_type_code}
                    )

                    if not operation_type:
                        print(f"   ⚠️  Operation type not found: {operation_type_code}")
                        error_count += 1
                        continue

                    # Parse passengers
                    passengers = parse_passengers(passengers_str)

                    if not dry_run:
                        # Create flight record
                        await prisma.flight.create(data={
                            'date': flight_date,
                            'airlineId': airline.id,
                            'route': route_str,
                            'registration': str(registration or ''),
                            'aircraftTypeId': aircraft_type.id,
                            'operationTypeId': operation_type.id,
                            'departureAirportId': departure_airport_id,
                            'arrivalAirportId': arrival_airport_id,

                            # Passenger data
                            'arrivalPassengers': passengers['adults'] if passengers else None,
                            'arrivalInfants': passengers['infants'] if passengers else None,

                            # Metadata
                            'dataSource': 'EXCEL_IMPORT_2025',
                            'importedFile': files[0],
                            'isVerified': False,
                        })

                    flight_count += 1

                except Exception as e:
                    print(f"   ❌ Error importing row {i}: {str(e)}")
                    error_count += 1

        status = '(DRY RUN)' if dry_run else ''
        print(f"   ✅ Imported {flight_count} flights {status}")
        if error_count > 0:
            print(f"   ⚠️  {error_count} errors")

        wb.close()
        return {'success': True, 'count': flight_count, 'errors': error_count}

    except Exception as e:
        print(f"   ❌ Error: {str(e)}")
        return {'success': False, 'count': 0, 'errors': 1}

async def main():
    """Main function"""
    args = sys.argv[1:]
    dry_run = '--dry-run' in args
    month_arg = None

    for arg in args:
        if not arg.startswith('--'):
            month_arg = arg
            break

    print('🚀 Starting 2025 Flight Import...')
    print(f"   Mode: {'DRY RUN' if dry_run else 'LIVE IMPORT'}\n")

    prisma = Prisma()
    await prisma.connect()

    try:
        if month_arg:
            # Import specific month
            print(f"📅 Importing single month: {month_arg}")
            await import_month(prisma, month_arg, dry_run)
        else:
            # Import all months
            month_folders = sorted([
                f for f in os.listdir(STATS_DIR)
                if os.path.isdir(os.path.join(STATS_DIR, f))
            ])

            print(f"📅 Found {len(month_folders)} months\n")

            total_flights = 0
            total_errors = 0

            for month_folder in month_folders:
                result = await import_month(prisma, month_folder, dry_run)
                total_flights += result['count']
                total_errors += result.get('errors', 0)

            print('\n' + '='*80)
            print('✅ IMPORT COMPLETED')
            print('='*80)
            print(f'\n📋 Summary:')
            print(f'   - Total flights: {total_flights}')
            print(f'   - Total errors: {total_errors}')

    except Exception as e:
        print(f'❌ Fatal error: {str(e)}')
        raise e
    finally:
        await prisma.disconnect()

if __name__ == '__main__':
    asyncio.run(main())
