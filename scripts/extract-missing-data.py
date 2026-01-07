#!/usr/bin/env python3

import openpyxl
import os
import sys
import json
from collections import defaultdict

def extract_data_from_year(year):
    """Extract all unique airlines and aircraft types from a year"""
    stats_dir = f"/Users/emir_mw/stats/STATS/{year}/Dnevni izvještaji"

    # Try both directory structures
    if not os.path.exists(stats_dir):
        stats_dir = f"/Users/emir_mw/stats/STATS/{year}/Mjesečni izvještaji"

    airlines = set()
    aircraft_types = set()

    if not os.path.exists(stats_dir):
        print(f"⚠️  Directory not found: {stats_dir}")
        return airlines, aircraft_types

    month_folders = sorted([
        f for f in os.listdir(stats_dir)
        if os.path.isdir(os.path.join(stats_dir, f))
    ])

    print(f"📊 Processing {year}...")

    for month_folder in month_folders:
        month_path = os.path.join(stats_dir, month_folder)

        # Find Excel file
        files = [
            f for f in os.listdir(month_path)
            if f.endswith('.xlsx') and not f.startswith('~')
        ]

        if not files:
            continue

        file_path = os.path.join(month_path, files[0])

        try:
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)

            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                rows = list(ws.iter_rows(values_only=True))

                if len(rows) < 2:
                    continue

                # Find airline and aircraft columns (usually at index 1 and 3/4)
                for row in rows[1:]:  # Skip header
                    if not row or len(row) < 5:
                        continue

                    # Airline is usually at index 1
                    airline = row[1]
                    if airline and str(airline).strip() and str(airline).strip() not in ['-', 'N/A', 'None']:
                        airline_str = str(airline).strip().upper()

                        # Skip summary/header labels
                        invalid_airlines = ['UKUPNO', 'REDOVNI PROMET', 'VANREDNI PROMET', 'OSTALA SLIJETANJA']
                        if airline_str not in invalid_airlines:
                            airlines.add(airline_str)

                    # Aircraft is usually at index 3 or 4
                    for idx in [3, 4]:
                        if len(row) > idx:
                            aircraft = row[idx]
                            if aircraft and str(aircraft).strip() and str(aircraft).strip() not in ['-', 'N/A', 'None', '0']:
                                aircraft_str = str(aircraft).strip().upper()

                                # Skip if it's just digits
                                if aircraft_str.isdigit():
                                    continue

                                # Skip if too long (likely registration or something else)
                                if len(aircraft_str) > 15:
                                    continue

                                # Skip if it looks like a registration (has - in middle and is long)
                                # Registrations: 9H-WAV, TC-NBC, YU-BTC, T7-ACA
                                if '-' in aircraft_str and len(aircraft_str) > 5:
                                    continue

                                # Skip common registration prefixes (country codes)
                                # 9A/9H (Croatia/Malta), TC (Turkey), YU (Serbia), HA/HB (Hungary/Swiss)
                                # E7 (Bosnia), 5A (Libya), LY (Lithuania), S5 (Slovenia), T7 (San Marino)
                                # OE (Austria), D- (Germany), F- (France)
                                registration_prefixes = [
                                    '9A', '9H', 'TC', 'YU', 'HA', 'HB', 'E7', '5A',
                                    'LY', 'S5', 'T7', 'OE', 'D-', 'F-', 'DB', 'DC', 'DF', 'OM'
                                ]
                                if any(aircraft_str.startswith(prefix) for prefix in registration_prefixes):
                                    continue

                                # Skip if it contains 'N' followed by digits (US registrations like N6509Z)
                                if aircraft_str.startswith('N') and any(c.isdigit() for c in aircraft_str[1:3]):
                                    continue

                                # Skip obvious non-aircraft entries
                                skip_terms = ['FAA', 'SCHEDULED', 'HAWK', 'RALL', 'TARR', 'COL3', 'PNR4']
                                if any(term in aircraft_str for term in skip_terms):
                                    continue

                                # Skip if it has dots (like HA.LYG)
                                if '.' in aircraft_str:
                                    continue

                                # Skip invalid aircraft types (header/summary labels)
                                invalid_aircraft = ['BROJ LETOVA', 'ISTOVARENO', 'PUTNICI', 'TERET', 'UKRCANO', 'UTOVARENO']
                                if aircraft_str in invalid_aircraft:
                                    continue

                                aircraft_types.add(aircraft_str)

            wb.close()
            print(f"  ✓ {month_folder}")

        except Exception as e:
            print(f"  ❌ {month_folder}: {str(e)}")

    return airlines, aircraft_types

def main():
    years = sys.argv[1:] if len(sys.argv) > 1 else ["2024", "2025"]

    print("🚀 Extracting Airlines and Aircraft Types\n")
    print(f"Years: {', '.join(years)}\n")

    all_airlines = set()
    all_aircraft = set()

    for year in years:
        airlines, aircraft = extract_data_from_year(year)
        all_airlines.update(airlines)
        all_aircraft.update(aircraft)
        print()

    # Sort and save results
    all_airlines = sorted(list(all_airlines))
    all_aircraft = sorted(list(all_aircraft))

    print("=" * 80)
    print("📋 RESULTS")
    print("=" * 80)
    print(f"\n✈️  Total unique airlines found: {len(all_airlines)}")
    print(f"🛩️  Total unique aircraft types found: {len(all_aircraft)}")

    # Save to JSON
    output = {
        'airlines': all_airlines,
        'aircraftTypes': all_aircraft,
        'extractedAt': str(__import__('datetime').datetime.now()),
        'years': years
    }

    os.makedirs('output', exist_ok=True)
    output_file = 'output/extracted-master-data.json'

    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump(output, f, indent=2, ensure_ascii=False)

    print(f"\n💾 Saved to: {output_file}")

    # Print preview
    print("\n" + "=" * 80)
    print("AIRLINES (first 20):")
    print("=" * 80)
    for airline in all_airlines[:20]:
        print(f"  - {airline}")
    if len(all_airlines) > 20:
        print(f"  ... and {len(all_airlines) - 20} more")

    print("\n" + "=" * 80)
    print("AIRCRAFT TYPES (first 20):")
    print("=" * 80)
    for aircraft in all_aircraft[:20]:
        print(f"  - {aircraft}")
    if len(all_aircraft) > 20:
        print(f"  ... and {len(all_aircraft) - 20} more")

    print("\n💡 Next step: Run comparison script to check what's missing in database")

if __name__ == '__main__':
    main()
