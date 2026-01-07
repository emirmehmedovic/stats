#!/usr/bin/env python3

import openpyxl
import os
import re

STATS_DIR = "/Users/emir_mw/stats/STATS/2025/Dnevni izvještaji"
month_folder = "01. JANUAR"
month_path = os.path.join(STATS_DIR, month_folder)

files = [
    f for f in os.listdir(month_path)
    if 'Dnevni izvještaj o saobraćaju' in f and f.endswith('.xlsx')
]

file_path = os.path.join(month_path, files[0])

def parse_passengers(value):
    """Parse passenger string"""
    if not value or value == '-' or value == 'N/A' or str(value).strip() == '':
        return 0, 0

    str_val = str(value).strip()

    # Pattern: "110+6 INF" or "110+6"
    match = re.match(r'^(\d+)\s*\+\s*(\d+)', str_val)
    if match:
        adults = int(match.group(1))
        infants = int(match.group(2))
        return adults, infants

    # Pattern: just a number "110"
    match = re.match(r'^(\d+)$', str_val)
    if match:
        adults = int(match.group(1))
        return adults, 0

    return 0, 0

wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)

# Aircraft types that are missing
missing_types = ['SA-342J', '737-800', 'IAR-330']

skipped_flights = []
skipped_pax_total = 0

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]

    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if i == 1:  # Skip header
            continue

        if not row[0]:  # Skip empty rows
            continue

        aircraft_model = row[3] if len(row) > 3 else None
        aircraft_model_str = str(aircraft_model).strip() if aircraft_model else ''

        # Check if this is a skipped flight
        if aircraft_model_str in missing_types:
            arrival_pax_str = row[13] if len(row) > 13 else None
            departure_pax_str = row[14] if len(row) > 14 else None

            arr_adults, arr_infants = parse_passengers(arrival_pax_str)
            dep_adults, dep_infants = parse_passengers(departure_pax_str)

            flight_total = arr_adults + arr_infants + dep_adults + dep_infants
            skipped_pax_total += flight_total

            skipped_flights.append({
                'date': row[0],
                'airline': row[1],
                'route': row[2],
                'aircraft': aircraft_model_str,
                'arr_adults': arr_adults,
                'arr_infants': arr_infants,
                'dep_adults': dep_adults,
                'dep_infants': dep_infants,
                'total': flight_total,
            })

wb.close()

print("="*80)
print("📋 SKIPPED FLIGHTS (Missing Aircraft Types)")
print("="*80)
print(f"\nTotal skipped flights: {len(skipped_flights)}")
print(f"Total passengers in skipped flights: {skipped_pax_total}\n")

for f in skipped_flights:
    print(f"{f['date']} - {f['airline']}")
    print(f"  Route: {f['route']}")
    print(f"  Aircraft: {f['aircraft']}")
    print(f"  Arrival: {f['arr_adults']} adults + {f['arr_infants']} infants")
    print(f"  Departure: {f['dep_adults']} adults + {f['dep_infants']} infants")
    print(f"  Total: {f['total']} passengers")
    print()

print("="*80)
print(f"Missing passengers from skipped flights: {skipped_pax_total}")
print(f"Current database total: 20,584")
print(f"Excel total: 20,817")
print(f"Difference: {20817 - 20584}")
print(f"Expected after adding skipped: {20584 + skipped_pax_total}")
print("="*80)
