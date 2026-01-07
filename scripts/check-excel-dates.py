#!/usr/bin/env python3

import openpyxl
import os

STATS_DIR = "/Users/emir_mw/stats/STATS/2025/Dnevni izvještaji"
month_folder = "01. JANUAR"
month_path = os.path.join(STATS_DIR, month_folder)

files = [
    f for f in os.listdir(month_path)
    if 'Dnevni izvještaj o saobraćaju' in f and f.endswith('.xlsx')
]

file_path = os.path.join(month_path, files[0])

wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)

print(f"📊 Checking dates in: {files[0]}\n")
print(f"Total sheets (days): {len(wb.sheetnames)}\n")

# Count dates
date_counts = {}

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]

    print(f"\n📅 Sheet: {sheet_name}")

    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if i == 1:  # Skip header
            continue

        if not row[0]:  # Skip empty
            continue

        date_val = row[0]
        date_str = str(date_val).split(' ')[0] if date_val else 'None'

        # First row in each sheet
        if i == 2:
            print(f"   First date in sheet: {date_str}")

        date_counts[date_str] = date_counts.get(date_str, 0) + 1

wb.close()

print("\n" + "="*80)
print("Date Distribution:")
print("="*80)
for date, count in sorted(date_counts.items()):
    print(f"{date}: {count} flights")
