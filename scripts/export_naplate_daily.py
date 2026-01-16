import json
import sys
from copy import copy
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def find_header_row(ws):
    for row in range(1, 15):
        values = [ws.cell(row=row, column=col).value for col in range(1, ws.max_column + 1)]
        if "Fee types" in values:
            return row
    return None


def build_blocks(ws, header_row):
    blocks = []
    for col in range(1, ws.max_column + 1):
        if ws.cell(row=header_row, column=col).value == "Fee types":
            blocks.append(
                {
                    "start": col,
                    "fee_types": col,
                    "code": col + 1,
                    "charged": col + 2,
                    "price": col + 3,
                    "qty": col + 4,
                    "amount": col + 5,
                    "valute": col + 6,
                }
            )
    return blocks


def clone_block(ws, source_start, target_start, block_width, max_row, source_merges):
    for offset in range(block_width):
        src_col = source_start + offset
        dst_col = target_start + offset
        src_letter = get_column_letter(src_col)
        dst_letter = get_column_letter(dst_col)
        if src_letter in ws.column_dimensions:
            ws.column_dimensions[dst_letter].width = ws.column_dimensions[src_letter].width

    for row in range(1, max_row + 1):
        for offset in range(block_width):
            src_col = source_start + offset
            dst_col = target_start + offset
            src_cell = ws.cell(row=row, column=src_col)
            dst_cell = ws.cell(row=row, column=dst_col)
            dst_cell.value = src_cell.value
            if src_cell.has_style:
                dst_cell.font = copy(src_cell.font)
                dst_cell.border = copy(src_cell.border)
                dst_cell.fill = copy(src_cell.fill)
                dst_cell.number_format = src_cell.number_format
                dst_cell.protection = copy(src_cell.protection)
                dst_cell.alignment = copy(src_cell.alignment)

    for rng in source_merges:
        col_offset = target_start - source_start
        ws.merge_cells(
            start_column=rng.min_col + col_offset,
            start_row=rng.min_row,
            end_column=rng.max_col + col_offset,
            end_row=rng.max_row,
        )


def scan_service_rows(ws, block, header_row):
    rows = []
    other_rows = []
    code_rows = {}
    for row in range(header_row + 2, ws.max_row + 1):
        label = ws.cell(row=row, column=block["fee_types"]).value
        code = ws.cell(row=row, column=block["code"]).value
        if isinstance(label, str) and label.strip().lower().startswith("total"):
            break
        if label is None and code is None:
            continue
        rows.append(row)
        if isinstance(label, str) and label.strip().lower().startswith("ostalo"):
            other_rows.append(row)
        if code:
            code_rows[str(code).strip()] = row
    return rows, other_rows, code_rows


def find_label_row(ws, label):
    for row in range(1, ws.max_row + 1):
        cell_value = ws.cell(row=row, column=2).value
        if isinstance(cell_value, str) and cell_value.strip().lower().startswith(label.lower()):
            return row
    return None


def write_amount_formula(ws, row, price_col, qty_col, amount_col):
    price_letter = get_column_letter(price_col)
    qty_letter = get_column_letter(qty_col)
    ws.cell(row=row, column=amount_col).value = f"={qty_letter}{row}*{price_letter}{row}"


def main():
    if len(sys.argv) < 4:
        print("Usage: export_naplate_daily.py <input_json> <template_xlsx> <output_xlsx>")
        sys.exit(1)

    input_path = Path(sys.argv[1])
    template_path = Path(sys.argv[2])
    output_path = Path(sys.argv[3])

    data = json.loads(input_path.read_text(encoding="utf-8"))

    wb = load_workbook(template_path)
    ws = wb.active

    header_row = find_header_row(ws)
    if header_row is None:
        print("Missing header row", file=sys.stderr)
        sys.exit(1)

    blocks = build_blocks(ws, header_row)
    carriers = data.get("carrierOrder") or list((data.get("carriers") or {}).keys())
    carriers = [carrier for carrier in carriers if carrier in (data.get("carriers") or {})]
    block_width = 7
    block_gap = 1
    if len(blocks) > 1:
        block_gap = max(1, blocks[1]["start"] - blocks[0]["start"] - block_width)

    if len(blocks) < len(carriers) and blocks:
        source_start = blocks[-1]["start"]
        source_end = source_start + block_width - 1
        source_merges = [
            rng for rng in ws.merged_cells.ranges
            if rng.min_col >= source_start and rng.max_col <= source_end
        ]
        max_row = ws.max_row
        for i in range(len(blocks), len(carriers)):
            target_start = blocks[0]["start"] + (block_width + block_gap) * i
            clone_block(ws, source_start, target_start, block_width, max_row, source_merges)
        blocks = build_blocks(ws, header_row)

    if len(blocks) > len(carriers):
        extras = blocks[len(carriers):]
        for block in reversed(extras):
            delete_start = block["start"] - block_gap
            ws.delete_cols(delete_start, block_width + block_gap)
        blocks = build_blocks(ws, header_row)
    airport_services = data.get("airportServices", [])
    adjustments_amount = float(data.get("adjustmentsAmount") or 0)

    for idx, carrier in enumerate(carriers):
        if idx >= len(blocks):
            continue
        block = blocks[idx]
        carrier_data = data["carriers"][carrier]
        services = carrier_data["services"]
        label = str(carrier_data.get("label") or carrier).strip()
        title = f"1. OTHER {label} SERVICES SOLD TO THE PASSENGERS AT AIRPORT"
        ws.cell(row=header_row - 1, column=block["fee_types"]).value = title

        rows, other_rows, code_rows = scan_service_rows(ws, block, header_row)

        # Reset service rows
        for row in rows:
            ws.cell(row=row, column=block["qty"]).value = 0
            if ws.cell(row=row, column=block["amount"]).data_type != "f":
                ws.cell(row=row, column=block["amount"]).value = 0

        other_iter = iter(other_rows)
        used_other = set()

        for service in services:
            code = str(service.get("code", "")).strip()
            target_row = code_rows.get(code)
            if target_row is None:
                try:
                    target_row = next(other_iter)
                    used_other.add(target_row)
                except StopIteration:
                    continue

            ws.cell(row=target_row, column=block["fee_types"]).value = service.get("label") or "Ostalo"
            ws.cell(row=target_row, column=block["code"]).value = code or ""
            ws.cell(row=target_row, column=block["charged"]).value = service.get("unit") or ""
            ws.cell(row=target_row, column=block["price"]).value = float(service.get("price") or 0)
            ws.cell(row=target_row, column=block["qty"]).value = float(service.get("qty") or 0)
            ws.cell(row=target_row, column=block["valute"]).value = service.get("currency") or "EUR"

            if service.get("price"):
                write_amount_formula(ws, target_row, block["price"], block["qty"], block["amount"])
            else:
                ws.cell(row=target_row, column=block["amount"]).value = float(service.get("amountOverride") or 0)

        # Clear unused "Ostalo" rows
        for row in other_rows:
            if row in used_other:
                continue
            ws.cell(row=row, column=block["fee_types"]).value = "Ostalo"
            ws.cell(row=row, column=block["code"]).value = ""
            ws.cell(row=row, column=block["charged"]).value = ""
            ws.cell(row=row, column=block["price"]).value = 0
            ws.cell(row=row, column=block["qty"]).value = 0
            ws.cell(row=row, column=block["amount"]).value = 0

        # Airport remuneration (dodatni servis)
        for row in range(header_row, ws.max_row + 1):
            label = ws.cell(row=row, column=block["fee_types"]).value
            if isinstance(label, str) and label.strip().lower().startswith("airport remunerations"):
                total_airport_rem = sum(
                    float(txn.get("airportRemunerationKm") or 0)
                    for txn in data["carriers"][carrier]["bookings"].get("transactions", [])
                )
                ws.cell(row=row, column=block["qty"]).value = total_airport_rem
                break

        # Bookings section
        booking_header_row = None
        for row in range(header_row, ws.max_row + 1):
            label = ws.cell(row=row, column=block["fee_types"]).value
            if isinstance(label, str) and label.strip().lower().startswith("bookings sold"):
                booking_header_row = row
                break
        if booking_header_row:
            start_row = booking_header_row + 1
            end_row = start_row
            for row in range(start_row, ws.max_row + 1):
                label = ws.cell(row=row, column=block["fee_types"]).value
                if isinstance(label, str) and label.strip().lower().startswith("total amount for"):
                    end_row = row - 1
                    break
            for row in range(start_row, end_row + 1):
                ws.cell(row=row, column=block["fee_types"]).value = None
                ws.cell(row=row, column=block["fee_types"] + 2).value = None
            transactions = data["carriers"][carrier]["bookings"].get("transactions", [])
            for offset, txn in enumerate(transactions):
                row = start_row + offset
                if row > end_row:
                    break
                ws.cell(row=row, column=block["fee_types"]).value = float(txn.get("amountEur") or 0)
                ws.cell(row=row, column=block["fee_types"] + 1).value = txn.get("pnr") or ""
                ws.cell(row=row, column=block["fee_types"] + 2).value = float(txn.get("pax") or 0)

    # Ticket commission row
    commission_row = find_label_row(ws, "Airport remuneration (Provizija na kartu)")
    if commission_row:
        total_commission = 0.0
        for carrier in carriers:
            for txn in data["carriers"][carrier]["bookings"].get("transactions", []):
                total_commission += float(txn.get("commissionKm") or 0)
        ws.cell(row=commission_row, column=5).value = total_commission

    # Airport services rows
    service_lookup = {item.get("id"): item for item in airport_services}

    def service_amount(service):
        price = float(service.get("price") or 0)
        qty = float(service.get("qty") or 0)
        if price:
            return price * qty
        return float(service.get("amountOverride") or 0)

    mapping = {
        "3. PVC ZIP vrecice prodate": "airport_pvc",
        "4. Higijenske Maske": "airport_masks",
        "5. Internet kodovi": "airport_internet",
        "6. Dječija nedelja": "airport_donation",
        "7. Dodatni Aerodromski servis": "airport_total",
    }

    airport_total = 0.0
    for key in ("airport_pvc", "airport_masks", "airport_internet", "airport_donation"):
        svc = service_lookup.get(key)
        if svc:
            airport_total += service_amount(svc)

    for label, key in mapping.items():
        row = find_label_row(ws, label)
        if not row:
            continue
        if key == "airport_total":
            ws.cell(row=row, column=5).value = airport_total + adjustments_amount
            continue
        svc = service_lookup.get(key)
        if not svc:
            continue
        if key == "airport_internet":
            ws.cell(row=row, column=4).value = float(svc.get("qty") or 0)
        ws.cell(row=row, column=5).value = service_amount(svc)

    korekcije_row = find_label_row(ws, "Korekcije")
    if korekcije_row:
        ws.cell(row=korekcije_row, column=5).value = adjustments_amount

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


if __name__ == "__main__":
    main()
