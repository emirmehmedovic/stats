import json
import sys
from copy import copy
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import MergedCell


def find_header_row(ws):
    for row in range(1, 20):
        values = [ws.cell(row=row, column=col).value for col in range(1, ws.max_column + 1)]
        if "Fee types" in values:
            return row
    return None


def find_label_row(ws, label):
    for row in range(1, ws.max_row + 1):
        cell_value = ws.cell(row=row, column=2).value
        if isinstance(cell_value, str) and label.lower() in cell_value.lower():
            return row
    return None


def find_service_total_row(ws, fee_col, header_row):
    for row in range(header_row + 2, ws.max_row + 1):
        label = ws.cell(row=row, column=fee_col).value
        if isinstance(label, str) and label.strip().lower().startswith("total"):
            return row
    return None


def find_value_column(ws, row, preferred):
    for col in preferred:
        cell = ws.cell(row=row, column=col)
        if cell.data_type == "f" or isinstance(cell.value, (int, float)):
            return col
    return preferred[0]

def find_row_value_column(ws, row, fallback):
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=row, column=col)
        if cell.data_type == "f":
            return col
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=row, column=col)
        if isinstance(cell.value, (int, float)):
            return col
    return fallback

def set_cell_value_safe(ws, row, col, value):
    cell = ws.cell(row=row, column=col)
    if cell.__class__.__name__ == "MergedCell":
        for merged in ws.merged_cells.ranges:
            if merged.min_row <= row <= merged.max_row and merged.min_col <= col <= merged.max_col:
                ws.cell(row=merged.min_row, column=merged.min_col).value = value
                return
    cell.value = value


def build_header_map(ws, header_row):
    headers = {}
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=header_row, column=col).value
        if isinstance(val, str):
            headers[val.strip().lower()] = col
    return headers


def scan_service_rows(ws, fee_col, code_col, header_row, price_col=None):
    rows = []
    other_rows = []
    code_rows = {}
    label_rows = {}
    price_rows = {}  # Map price -> list of rows for matching by price
    for row in range(header_row + 2, ws.max_row + 1):
        label = ws.cell(row=row, column=fee_col).value
        code = ws.cell(row=row, column=code_col).value
        if isinstance(label, str) and label.strip().lower().startswith("total"):
            break
        if label is None and code is None:
            continue
        rows.append(row)
        if isinstance(label, str) and label.strip().lower().startswith(("other", "ostalo")):
            other_rows.append(row)
        if isinstance(label, str) and label.strip():
            label_rows[label.strip()] = row
        if code:
            code_rows[str(code).strip()] = row
        # Store price mapping for matching by price (support multiple rows per price)
        if price_col:
            price_val = ws.cell(row=row, column=price_col).value
            if isinstance(price_val, (int, float)) and price_val > 0:
                price_key = float(price_val)
                if price_key not in price_rows:
                    price_rows[price_key] = []
                price_rows[price_key].append(row)
    return rows, other_rows, code_rows, label_rows, price_rows


def write_amount_formula(ws, row, price_col, qty_col, amount_col):
    price_letter = get_column_letter(price_col)
    qty_letter = get_column_letter(qty_col)
    ws.cell(row=row, column=amount_col).value = f"={qty_letter}{row}*{price_letter}{row}"

def copy_row_style(ws, src_row, dst_row, start_col, end_col):
    for col in range(start_col, end_col + 1):
        src = ws.cell(row=src_row, column=col)
        dst = ws.cell(row=dst_row, column=col)
        if src.has_style:
            dst.font = copy(src.font)
            dst.border = copy(src.border)
            dst.fill = copy(src.fill)
            dst.number_format = src.number_format
            dst.protection = copy(src.protection)
            dst.alignment = copy(src.alignment)
        dst.value = None


def export_sky_speed(template_path, output_path, carrier_data):
    wb = load_workbook(template_path)
    ws = wb.active

    header_row = find_header_row(ws)
    if header_row is None:
        raise RuntimeError("Missing header row in template")

    headers = build_header_map(ws, header_row)
    fee_col = headers.get("fee types", 2)
    code_col = headers.get("fees code", 3)
    price_col = headers.get("eur", 4)
    qty_col = headers.get("qty", 5)
    amount_col = headers.get("amount", 6)
    valute_col = headers.get("valute", 7)

    rows, other_rows, code_rows, label_rows, price_rows = scan_service_rows(ws, fee_col, code_col, header_row, price_col)
    template_row = other_rows[0] if other_rows else (rows[-1] if rows else None)

    for row in rows:
        ws.cell(row=row, column=qty_col).value = 0
        if ws.cell(row=row, column=amount_col).data_type != "f":
            ws.cell(row=row, column=amount_col).value = 0

    other_iter = iter(other_rows)
    used_other = set()
    used_rows = set()
    template_row = other_rows[0] if other_rows else (rows[-1] if rows else None)
    template_row = other_rows[0] if other_rows else (rows[-1] if rows else None)
    template_row = other_rows[0] if other_rows else (rows[-1] if rows else None)

    for service in carrier_data.get("services", []):
        label = str(service.get("label", "")).strip()
        code = str(service.get("code", "")).strip()
        price = float(service.get("price", 0))

        target_row = None

        # Try to find by price first (most specific for services with same label/code)
        if price > 0 and price in price_rows and len(price_rows[price]) > 0:
            # Pop the first unused row with this price
            for potential_row in price_rows[price]:
                if potential_row not in used_rows:
                    target_row = potential_row
                    price_rows[price].remove(potential_row)
                    used_rows.add(target_row)
                    break

        # If not found by price, try by label
        if target_row is None:
            target_row = label_rows.get(label)
            if target_row and target_row in used_rows:
                target_row = None  # Already used

        # If not found by label, try by code (but skip BAGEXC as it's not unique)
        if target_row is None and code and code != "BAGEXC":
            target_row = code_rows.get(code)
            if target_row and target_row in used_rows:
                target_row = None  # Already used

        # If still not found, use "Other" row
        if target_row is None:
            try:
                target_row = next(other_iter)
                used_other.add(target_row)
            except StopIteration:
                total_row = find_service_total_row(ws, fee_col, header_row)
                if total_row is None or template_row is None:
                    continue
                ws.insert_rows(total_row)
                copy_row_style(ws, template_row, total_row, fee_col, valute_col)
                target_row = total_row
                other_rows.append(target_row)
                used_other.add(target_row)

        if target_row:
            used_rows.add(target_row)

        ws.cell(row=target_row, column=fee_col).value = service.get("label") or "Other"
        ws.cell(row=target_row, column=code_col).value = code or ""
        ws.cell(row=target_row, column=price_col).value = float(service.get("price") or 0)
        ws.cell(row=target_row, column=qty_col).value = float(service.get("qty") or 0)
        ws.cell(row=target_row, column=valute_col).value = service.get("currency") or "EUR"
        if service.get("price"):
            write_amount_formula(ws, target_row, price_col, qty_col, amount_col)
        else:
            ws.cell(row=target_row, column=amount_col).value = float(service.get("amountOverride") or 0)

    for row in other_rows:
        if row in used_other:
            continue
        ws.cell(row=row, column=fee_col).value = "Other"
        ws.cell(row=row, column=code_col).value = ""
        ws.cell(row=row, column=price_col).value = 0
        ws.cell(row=row, column=qty_col).value = 0
        ws.cell(row=row, column=amount_col).value = 0

    booking_totals = carrier_data.get("bookings", {}).get("transactions", [])
    total_booking = sum(float(txn.get("amountEur") or 0) for txn in booking_totals)
    booking_row = find_label_row(ws, "booking")
    if booking_row:
        value_col = find_value_column(ws, booking_row, [amount_col, qty_col, price_col])
        ws.cell(row=booking_row, column=value_col).value = total_booking

    wb.save(output_path)


def export_wizz_airport(template_path, output_path, carrier_data, airport_services, adjustments_amount):
    wb = load_workbook(template_path)
    ws = wb.active

    header_row = find_header_row(ws)
    if header_row is None:
        raise RuntimeError("Missing header row in template")

    headers = build_header_map(ws, header_row)
    fee_col = headers.get("fee types", 2)
    code_col = headers.get("fees code", 3)
    charged_col = headers.get("charged", 4)
    price_col = headers.get("eur", 5)
    qty_col = headers.get("qty", 6)
    amount_col = headers.get("amount", 7)
    valute_col = headers.get("valute", 8)

    title_row = header_row - 1
    label = str(carrier_data.get("label") or "AIRLINE").strip()
    ws.cell(row=title_row, column=fee_col).value = f"1. OTHER {label} SERVICES SOLD TO THE PASSENGERS AT AIRPORT"

    rows, other_rows, code_rows, label_rows, price_rows = scan_service_rows(ws, fee_col, code_col, header_row, price_col)

    for row in rows:
        ws.cell(row=row, column=qty_col).value = 0
        if ws.cell(row=row, column=amount_col).data_type != "f":
            ws.cell(row=row, column=amount_col).value = 0

    other_iter = iter(other_rows)
    used_other = set()
    used_rows = set()

    for service in carrier_data.get("services", []):
        label = str(service.get("label", "")).strip()
        code = str(service.get("code", "")).strip()
        price = float(service.get("price", 0))

        target_row = None

        # Try to find by price first (most specific for services with same label/code)
        if price > 0 and price in price_rows and len(price_rows[price]) > 0:
            # Pop the first unused row with this price
            for potential_row in price_rows[price]:
                if potential_row not in used_rows:
                    target_row = potential_row
                    price_rows[price].remove(potential_row)
                    used_rows.add(target_row)
                    break

        # If not found by price, try by label
        if target_row is None:
            target_row = label_rows.get(label)
            if target_row and target_row in used_rows:
                target_row = None  # Already used

        # If not found by label, try by code (but skip BAGEXC as it's not unique)
        if target_row is None and code and code != "BAGEXC":
            target_row = code_rows.get(code)
            if target_row and target_row in used_rows:
                target_row = None  # Already used

        # If still not found, use "Other" row
        if target_row is None:
            try:
                target_row = next(other_iter)
                used_other.add(target_row)
            except StopIteration:
                total_row = find_service_total_row(ws, fee_col, header_row)
                if total_row is None or template_row is None:
                    continue
                ws.insert_rows(total_row)
                copy_row_style(ws, template_row, total_row, fee_col, valute_col)
                target_row = total_row
                used_other.add(target_row)

        if target_row:
            used_rows.add(target_row)

        ws.cell(row=target_row, column=fee_col).value = service.get("label") or "Ostalo"
        ws.cell(row=target_row, column=code_col).value = code or ""
        ws.cell(row=target_row, column=charged_col).value = service.get("unit") or ""
        ws.cell(row=target_row, column=price_col).value = float(service.get("price") or 0)
        ws.cell(row=target_row, column=qty_col).value = float(service.get("qty") or 0)
        ws.cell(row=target_row, column=valute_col).value = service.get("currency") or "EUR"
        if service.get("price"):
            write_amount_formula(ws, target_row, price_col, qty_col, amount_col)
        else:
            ws.cell(row=target_row, column=amount_col).value = float(service.get("amountOverride") or 0)

    for row in other_rows:
        if row in used_other:
            continue
        ws.cell(row=row, column=fee_col).value = "Ostalo"
        ws.cell(row=row, column=code_col).value = ""
        ws.cell(row=row, column=charged_col).value = ""
        ws.cell(row=row, column=price_col).value = 0
        ws.cell(row=row, column=qty_col).value = 0
        ws.cell(row=row, column=amount_col).value = 0

    # Get booking transactions (needed for multiple calculations below)
    booking_txns = carrier_data.get("bookings", {}).get("transactions", [])

    # airport remuneration - calculate total service items (qty) × 10 KM
    total_service_items = sum(float(service.get("qty") or 0) for service in carrier_data.get("services", []))
    airport_rem_value = total_service_items * 10  # 10 KM per service item
    airport_rem_row = find_label_row(ws, "airport remunerations")
    if airport_rem_row:
        value_col = find_value_column(ws, airport_rem_row, [amount_col, qty_col, price_col])
        ws.cell(row=airport_rem_row, column=value_col).value = airport_rem_value

    # bookings section
    booking_header_row = find_label_row(ws, "bookings sold")
    total_booking_row = find_label_row(ws, "total amount for")
    if booking_header_row and total_booking_row:
        start_row = booking_header_row + 1
        end_row = total_booking_row - 1
        for row in range(start_row, end_row + 1):
            ws.cell(row=row, column=fee_col).value = None
            ws.cell(row=row, column=fee_col + 1).value = None

        total_amount = sum(float(txn.get("amountEur") or 0) for txn in booking_txns)
        total_pax = sum(float(txn.get("pax") or 0) for txn in booking_txns)
        ws.cell(row=start_row, column=fee_col).value = total_amount
        ws.cell(row=start_row, column=fee_col + 1).value = total_pax

        value_col = find_value_column(ws, total_booking_row, [amount_col, qty_col, price_col])
        ws.cell(row=total_booking_row, column=value_col).value = total_amount

    commission_row = find_label_row(ws, "provizija")
    if commission_row:
        total_commission = sum(float(txn.get("commissionKm") or 0) for txn in booking_txns)
        value_col = find_value_column(ws, commission_row, [amount_col, qty_col, price_col])
        ws.cell(row=commission_row, column=value_col).value = total_commission

    booking_rem_row = find_label_row(ws, "Airport remuneration (Provizija na kartu)")
    total_booking_rem = 0.0
    if booking_rem_row:
        total_booking_rem = sum(float(txn.get("airportRemunerationKm") or 0) for txn in booking_txns)
        value_col = find_value_column(ws, booking_rem_row, [amount_col, qty_col, price_col])
        ws.cell(row=booking_rem_row, column=value_col).value = total_booking_rem

    # Airport services rows
    service_lookup = {item.get("id"): item for item in airport_services}

    def service_amount(service):
        price = float(service.get("price") or 0)
        qty = float(service.get("qty") or 0)
        if price:
            return price * qty
        return float(service.get("amountOverride") or 0)

    mapping = {
        "pvc zip": "airport_pvc",
        "higijenske": "airport_masks",
        "internet": "airport_internet",
        "dječija": "airport_donation",
    }

    airport_total = 0.0
    for key in ("airport_pvc", "airport_masks", "airport_internet", "airport_donation"):
        svc = service_lookup.get(key)
        if svc:
            airport_total += service_amount(svc)

    remove_row = find_label_row(ws, "7. Dodatni Aerodromski servis")
    if remove_row:
        for col in range(fee_col, valute_col + 1):
            cell = ws.cell(row=remove_row, column=col)
            if not isinstance(cell, MergedCell):
                cell.value = None

    for row in range(1, ws.max_row + 1):
        label = ws.cell(row=row, column=fee_col).value
        if not isinstance(label, str):
            continue
        lowered = label.lower()
        for marker, key in mapping.items():
            if marker in lowered:
                svc = service_lookup.get(key)
                # Airport services block uses fixed columns (qty in D, amount in E/F area)
                if not svc:
                    set_cell_value_safe(ws, row, 4, 0)
                    value_col = find_row_value_column(ws, row, 5)
                    set_cell_value_safe(ws, row, value_col, 0)
                    break
                if key == "airport_donation":
                    # Donation is a total amount; qty should stay empty/0.
                    donation_amount = float(svc.get("amountOverride") or 0)
                    if donation_amount == 0:
                        donation_amount = float(svc.get("qty") or 0)
                    set_cell_value_safe(ws, row, 4, 0)
                    value_col = find_row_value_column(ws, row, 5)
                    set_cell_value_safe(ws, row, 5, donation_amount)
                    if value_col != 5:
                        set_cell_value_safe(ws, row, value_col, 0)
                else:
                    set_cell_value_safe(ws, row, 4, float(svc.get("qty") or 0))
                    value_col = find_row_value_column(ws, row, 5)
                    set_cell_value_safe(ws, row, value_col, service_amount(svc))
                break

    # Ensure donation row is cleared if no data (template may contain a fixed value)
    donation_row = find_label_row(ws, "Dječija")
    if donation_row:
        svc = service_lookup.get("airport_donation")
        qty = float((svc or {}).get("qty") or 0)
        amount_override = float((svc or {}).get("amountOverride") or 0)
        if qty == 0 and amount_override == 0:
            set_cell_value_safe(ws, donation_row, 4, 0)
            set_cell_value_safe(ws, donation_row, 5, 0)

    amount_airport_row = find_label_row(ws, "Amount for Airport Tuzla")
    if amount_airport_row:
        total_airport_amount = airport_rem_value + total_booking_rem + airport_total
        value_col = find_row_value_column(ws, amount_airport_row, amount_col)
        ws.cell(row=amount_airport_row, column=value_col).value = total_airport_amount

    wb.save(output_path)


def export_general(template_path, output_path, report):
    wb = load_workbook(template_path)
    ws = wb.active

    start_row = 4
    row = start_row
    airport_services_total = sum(
        (float(item.get("qty") or 0) * float(item.get("price") or 0)) if item.get("price") else float(item.get("amountOverride") or 0)
        for item in report.get("airportServices", [])
    )
    adjustments = float(report.get("adjustmentsAmount") or 0)

    carriers = report.get("carriers") or {}
    order = report.get("carrierOrder") or list(carriers.keys())

    total_services = total_bookings = total_eur = 0.0
    total_airport_rem = total_commission = 0.0

    for carrier in order:
        carrier_data = carriers.get(carrier)
        if not carrier_data:
            continue
        label = carrier_data.get("label") or carrier
        services_eur = sum(float(s.get("qty") or 0) * float(s.get("price") or 0) if s.get("price") else float(s.get("amountOverride") or 0) for s in carrier_data.get("services", []))
        booking_txns = carrier_data.get("bookings", {}).get("transactions", [])
        bookings_eur = sum(float(txn.get("amountEur") or 0) for txn in booking_txns)
        airport_rem = sum(float(txn.get("airportRemunerationKm") or 0) for txn in booking_txns)
        commission = sum(float(txn.get("commissionKm") or 0) for txn in booking_txns)
        total = services_eur + bookings_eur

        ws.cell(row=row, column=1).value = label
        ws.cell(row=row, column=2).value = services_eur
        ws.cell(row=row, column=3).value = bookings_eur
        ws.cell(row=row, column=4).value = total
        ws.cell(row=row, column=5).value = airport_rem
        ws.cell(row=row, column=6).value = commission
        ws.cell(row=row, column=7).value = 0
        ws.cell(row=row, column=8).value = 0
        ws.cell(row=row, column=9).value = airport_rem + commission
        ws.cell(row=row, column=10).value = 0

        total_services += services_eur
        total_bookings += bookings_eur
        total_eur += total
        total_airport_rem += airport_rem
        total_commission += commission
        row += 1

    ws.cell(row=row, column=1).value = "Airport Tuzla"
    ws.cell(row=row, column=7).value = airport_services_total
    ws.cell(row=row, column=8).value = adjustments
    ws.cell(row=row, column=9).value = airport_services_total + adjustments
    row += 1

    fx_rate = float(report.get("fxRateEurToKm") or 1.95583)
    total_airport_km = total_airport_rem + total_commission + airport_services_total + adjustments
    total_km = total_eur * fx_rate + total_airport_km

    ws.cell(row=row, column=1).value = "Ukupno"
    ws.cell(row=row, column=2).value = total_services
    ws.cell(row=row, column=3).value = total_bookings
    ws.cell(row=row, column=4).value = total_eur
    ws.cell(row=row, column=5).value = total_airport_rem
    ws.cell(row=row, column=6).value = total_commission
    ws.cell(row=row, column=7).value = airport_services_total
    ws.cell(row=row, column=8).value = adjustments
    ws.cell(row=row, column=9).value = total_airport_km
    ws.cell(row=row, column=10).value = total_km

    wb.save(output_path)


def main():
    if len(sys.argv) < 5:
        print("Usage: export_naplate_monthly.py <mode> <input_json> <template_xlsx> <output_xlsx>")
        sys.exit(1)

    mode = sys.argv[1]
    input_path = Path(sys.argv[2])
    template_path = Path(sys.argv[3])
    output_path = Path(sys.argv[4])

    data = json.loads(input_path.read_text(encoding="utf-8"))
    carrier = data.get("carrier")
    report = data.get("report") or {}

    output_path.parent.mkdir(parents=True, exist_ok=True)

    if mode == "sky-speed":
        if not carrier:
            raise RuntimeError("Missing carrier for sky-speed export")
        export_sky_speed(template_path, output_path, report["carriers"][carrier])
    elif mode == "carrier-airport":
        if not carrier:
            raise RuntimeError("Missing carrier for carrier-airport export")
        export_wizz_airport(
            template_path,
            output_path,
            report["carriers"][carrier],
            report.get("airportServices", []),
            report.get("adjustmentsAmount", 0),
        )
    elif mode == "general":
        export_general(template_path, output_path, report)
    else:
        raise RuntimeError("Unknown mode")


if __name__ == "__main__":
    main()
