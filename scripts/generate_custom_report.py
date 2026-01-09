#!/usr/bin/env python3
"""
Custom Advanced Report Generator

Generiše prilagođeni izvještaj sa naprednim filterima:
- Tip saobraćaja (multi-select)
- Aviokompanije (multi-select)
- Rute (multi-select)
- Period (od-do)
- Tip putnika (odlazni/dolazni/svi/bebe)
"""

import sys
import os
import json
from pathlib import Path
from datetime import datetime
import psycopg2
from psycopg2.extras import RealDictCursor
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

SCRIPT_DIR = Path(__file__).parent
PROJECT_ROOT = SCRIPT_DIR.parent
OUTPUT_DIR = PROJECT_ROOT / "izvještaji" / "generated"


def get_db_connection():
    """Konekcija na PostgreSQL bazu"""
    DATABASE_URL = os.getenv("DATABASE_URL")
    if not DATABASE_URL:
        raise ValueError("DATABASE_URL environment variable not set")
    return psycopg2.connect(DATABASE_URL, cursor_factory=RealDictCursor)


def get_flight_data(date_from, date_to, operation_types, airlines, routes):
    """
    Povuči letove sa filterima
    """
    conn = get_db_connection()
    cursor = conn.cursor()

    # Build WHERE conditions
    where_conditions = ["f.date >= %s", "f.date <= %s"]
    params = [date_from, date_to]

    if operation_types:
        placeholders = ','.join(['%s'] * len(operation_types))
        where_conditions.append(f'f."operationTypeId" IN ({placeholders})')
        params.extend(operation_types)

    if airlines:
        placeholders = ','.join(['%s'] * len(airlines))
        where_conditions.append(f'f."airlineId" IN ({placeholders})')
        params.extend(airlines)

    if routes:
        placeholders = ','.join(['%s'] * len(routes))
        where_conditions.append(f'f.route IN ({placeholders})')
        params.extend(routes)

    where_clause = " AND ".join(where_conditions)

    query = f"""
        SELECT
            f.id,
            f.date,
            f.route,
            f.registration,
            
            -- Airline
            a.name as airline_name,
            a."icaoCode" as airline_icao,
            
            -- Aircraft Type
            at.model as aircraft_model,
            
            -- Operation Type
            ot.name as operation_type_name,
            ot.code as operation_type_code,
            
            -- Arrival info
            f."arrivalFlightNumber",
            f."arrivalScheduledTime",
            f."arrivalActualTime",
            f."arrivalPassengers",
            f."arrivalMalePassengers",
            f."arrivalFemalePassengers",
            f."arrivalChildren",
            f."arrivalInfants",
            
            -- Departure info
            f."departureFlightNumber",
            f."departureScheduledTime",
            f."departureActualTime",
            f."departurePassengers",
            f."departureMalePassengers",
            f."departureFemalePassengers",
            f."departureChildren",
            f."departureInfants"
            
        FROM "Flight" f
        LEFT JOIN "Airline" a ON f."airlineId" = a.id
        LEFT JOIN "AircraftType" at ON f."aircraftTypeId" = at.id
        LEFT JOIN "OperationType" ot ON f."operationTypeId" = ot.id
        WHERE {where_clause}
        ORDER BY f.date ASC, f."departureScheduledTime" ASC
    """

    cursor.execute(query, params)
    flights = cursor.fetchall()
    cursor.close()
    conn.close()

    return flights


def format_datetime(dt):
    """Format datetime za prikaz"""
    if not dt:
        return "-"
    if isinstance(dt, str):
        dt = datetime.fromisoformat(dt.replace('Z', '+00:00'))
    return dt.strftime("%d.%m.%Y %H:%M")


def format_date(dt):
    """Format date za prikaz"""
    if not dt:
        return "-"
    if isinstance(dt, str):
        dt = datetime.fromisoformat(dt.split('T')[0])
    return dt.strftime("%d.%m.%Y")


def create_merged_cell(ws, row, start_col, end_col, value, font=None, fill=None, alignment=None, border=None):
    """
    Helper function to properly create merged cells with consistent formatting.
    Applies formatting to ALL cells in the range before merging to avoid Excel corruption.
    """
    for col in range(start_col, end_col + 1):
        cell = ws.cell(row=row, column=col)
        if col == start_col and value is not None:
            cell.value = value
        if font:
            cell.font = font
        if fill:
            cell.fill = fill
        if alignment:
            cell.alignment = alignment
        if border:
            cell.border = border
    
    if start_col != end_col:
        ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=end_col)


def create_excel_report(flights, passenger_type, date_from, date_to):
    """
    Kreira Excel izvještaj sa filteriranim podacima
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Custom izvještaj"

    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Define columns based on passenger type
    base_headers = ["Datum", "Aviokompanija", "Tip aviona", "Registracija", "Ruta", "Tip operacije"]
    
    if passenger_type == "departure":
        headers = base_headers + [
            "Broj leta (odlazak)", "Planirano vrijeme", "Stvarno vrijeme",
            "Putnici", "Muškarci", "Žene", "Djeca", "Bebe"
        ]
    elif passenger_type == "arrival":
        headers = base_headers + [
            "Broj leta (dolazak)", "Planirano vrijeme", "Stvarno vrijeme",
            "Putnici", "Muškarci", "Žene", "Djeca", "Bebe"
        ]
    elif passenger_type == "infants":
        headers = base_headers + [
            "Broj leta (dolazak)", "Broj leta (odlazak)",
            "Bebe (dolazak)", "Bebe (odlazak)", "Ukupno beba"
        ]
    else:  # all
        headers = base_headers + [
            "Broj leta (dolazak)", "Putnici (dolazak)",
            "Broj leta (odlazak)", "Putnici (odlazak)",
            "Ukupno putnika"
        ]

    # Write headers
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Set column widths
    column_widths = [12, 25, 15, 12, 15, 15] + [18] * (len(headers) - 6)
    for col_idx, width in enumerate(column_widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

    # Write data rows
    row_idx = 2
    total_passengers = 0
    total_departure_passengers = 0
    total_arrival_passengers = 0
    total_infants = 0
    total_operations = 0

    for flight in flights:
        base_data = [
            format_date(flight['date']),
            flight['airline_name'] or "-",
            flight['aircraft_model'] or "-",
            flight['registration'] or "-",
            flight['route'] or "-",
            flight['operation_type_name'] or "-"
        ]

        if passenger_type == "departure":
            passenger_data = [
                flight['departureFlightNumber'] or "-",
                format_datetime(flight['departureScheduledTime']),
                format_datetime(flight['departureActualTime']),
                flight['departurePassengers'] or 0,
                flight['departureMalePassengers'] or 0,
                flight['departureFemalePassengers'] or 0,
                flight['departureChildren'] or 0,
                flight['departureInfants'] or 0
            ]
            dep_pax = flight['departurePassengers'] or 0
            total_passengers += dep_pax
            total_departure_passengers += dep_pax
            if flight['departureFlightNumber']:
                total_operations += 1
        elif passenger_type == "arrival":
            passenger_data = [
                flight['arrivalFlightNumber'] or "-",
                format_datetime(flight['arrivalScheduledTime']),
                format_datetime(flight['arrivalActualTime']),
                flight['arrivalPassengers'] or 0,
                flight['arrivalMalePassengers'] or 0,
                flight['arrivalFemalePassengers'] or 0,
                flight['arrivalChildren'] or 0,
                flight['arrivalInfants'] or 0
            ]
            arr_pax = flight['arrivalPassengers'] or 0
            total_passengers += arr_pax
            total_arrival_passengers += arr_pax
            if flight['arrivalFlightNumber']:
                total_operations += 1
        elif passenger_type == "infants":
            arr_infants = flight['arrivalInfants'] or 0
            dep_infants = flight['departureInfants'] or 0
            passenger_data = [
                flight['arrivalFlightNumber'] or "-",
                flight['departureFlightNumber'] or "-",
                arr_infants,
                dep_infants,
                arr_infants + dep_infants
            ]
            total_infants += arr_infants + dep_infants
        else:  # all
            arr_pax = flight['arrivalPassengers'] or 0
            dep_pax = flight['departurePassengers'] or 0
            passenger_data = [
                flight['arrivalFlightNumber'] or "-",
                arr_pax,
                flight['departureFlightNumber'] or "-",
                dep_pax,
                arr_pax + dep_pax
            ]
            total_passengers += arr_pax + dep_pax
            total_arrival_passengers += arr_pax
            total_departure_passengers += dep_pax
            if flight['arrivalFlightNumber']:
                total_operations += 1
            if flight['departureFlightNumber']:
                total_operations += 1

        row_data = base_data + passenger_data

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            if col_idx > 6:  # Numeric columns
                cell.alignment = Alignment(horizontal="center")

        row_idx += 1

    # Create comprehensive summary table
    row_idx += 2
    
    # Summary header
    summary_header_fill = PatternFill(start_color="F59E0B", end_color="F59E0B", fill_type="solid")
    create_merged_cell(
        ws, row_idx, 1, len(headers),
        "SAŽETAK IZVJEŠTAJA",
        font=Font(bold=True, size=14, color="FFFFFF"),
        fill=summary_header_fill,
        alignment=Alignment(horizontal="center", vertical="center"),
        border=thin_border
    )
    row_idx += 1
    
    # Period info
    period_cell = ws.cell(row=row_idx, column=1, value="Period:")
    period_cell.font = Font(bold=True)
    period_cell.border = thin_border
    
    create_merged_cell(
        ws, row_idx, 2, 4,
        f"{date_from} do {date_to}",
        border=thin_border
    )
    row_idx += 1
    
    # Basic statistics
    stats_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
    
    # Total flights
    flights_label = ws.cell(row=row_idx, column=1, value="Ukupno letova:")
    flights_label.font = Font(bold=True)
    flights_label.fill = stats_fill
    flights_label.border = thin_border
    flights_value = ws.cell(row=row_idx, column=2, value=len(flights))
    flights_value.font = Font(bold=True, size=12)
    flights_value.fill = stats_fill
    flights_value.alignment = Alignment(horizontal="center")
    flights_value.border = thin_border
    row_idx += 1
    
    # Total operations
    ops_label = ws.cell(row=row_idx, column=1, value="Ukupno operacija:")
    ops_label.font = Font(bold=True)
    ops_label.fill = stats_fill
    ops_label.border = thin_border
    ops_value = ws.cell(row=row_idx, column=2, value=total_operations)
    ops_value.font = Font(bold=True, size=12)
    ops_value.fill = stats_fill
    ops_value.alignment = Alignment(horizontal="center")
    ops_value.border = thin_border
    row_idx += 1
    
    # Total passengers/infants
    if passenger_type == "infants":
        pax_label = ws.cell(row=row_idx, column=1, value="Ukupno beba:")
        pax_label.font = Font(bold=True)
        pax_label.fill = stats_fill
        pax_label.border = thin_border
        pax_value = ws.cell(row=row_idx, column=2, value=total_infants)
        pax_value.font = Font(bold=True, size=12, color="D97706")
        pax_value.fill = stats_fill
        pax_value.alignment = Alignment(horizontal="center")
        pax_value.border = thin_border
    else:
        pax_label = ws.cell(row=row_idx, column=1, value="Ukupno putnika:")
        pax_label.font = Font(bold=True)
        pax_label.fill = stats_fill
        pax_label.border = thin_border
        pax_value = ws.cell(row=row_idx, column=2, value=total_passengers)
        pax_value.font = Font(bold=True, size=12, color="D97706")
        pax_value.fill = stats_fill
        pax_value.alignment = Alignment(horizontal="center")
        pax_value.border = thin_border
    row_idx += 1
    
    # Departure passengers (only if not infants type)
    if passenger_type != "infants" and total_departure_passengers > 0:
        dep_pax_label = ws.cell(row=row_idx, column=1, value="  - Odlazeći putnici:")
        dep_pax_label.font = Font(bold=False, italic=True)
        dep_pax_label.fill = stats_fill
        dep_pax_label.border = thin_border
        dep_pax_value = ws.cell(row=row_idx, column=2, value=total_departure_passengers)
        dep_pax_value.font = Font(bold=True)
        dep_pax_value.fill = stats_fill
        dep_pax_value.alignment = Alignment(horizontal="center")
        dep_pax_value.border = thin_border
        row_idx += 1
    
    # Arrival passengers (only if not infants type)
    if passenger_type != "infants" and total_arrival_passengers > 0:
        arr_pax_label = ws.cell(row=row_idx, column=1, value="  - Dolazeći putnici:")
        arr_pax_label.font = Font(bold=False, italic=True)
        arr_pax_label.fill = stats_fill
        arr_pax_label.border = thin_border
        arr_pax_value = ws.cell(row=row_idx, column=2, value=total_arrival_passengers)
        arr_pax_value.font = Font(bold=True)
        arr_pax_value.fill = stats_fill
        arr_pax_value.alignment = Alignment(horizontal="center")
        arr_pax_value.border = thin_border
        row_idx += 1
    
    # Average passengers per flight
    if passenger_type != "infants" and len(flights) > 0:
        avg_label = ws.cell(row=row_idx, column=1, value="Prosječno putnika po letu:")
        avg_label.font = Font(bold=True)
        avg_label.fill = stats_fill
        avg_label.border = thin_border
        avg_value = ws.cell(row=row_idx, column=2, value=round(total_passengers / len(flights), 1))
        avg_value.font = Font(bold=True)
        avg_value.fill = stats_fill
        avg_value.alignment = Alignment(horizontal="center")
        avg_value.border = thin_border
        row_idx += 1
    
    # Average passengers per operation
    if passenger_type != "infants" and total_operations > 0:
        avg_op_label = ws.cell(row=row_idx, column=1, value="Prosječno putnika po operaciji:")
        avg_op_label.font = Font(bold=True)
        avg_op_label.fill = stats_fill
        avg_op_label.border = thin_border
        avg_op_value = ws.cell(row=row_idx, column=2, value=round(total_passengers / total_operations, 1))
        avg_op_value.font = Font(bold=True)
        avg_op_value.fill = stats_fill
        avg_op_value.alignment = Alignment(horizontal="center")
        avg_op_value.border = thin_border
        row_idx += 1
    
    row_idx += 1
    
    # Breakdown by airline
    airline_counts = {}
    airline_passengers = {}
    airline_departure_passengers = {}
    airline_arrival_passengers = {}
    airline_operations = {}
    for flight in flights:
        airline = flight['airline_name'] or "Nepoznato"
        airline_counts[airline] = airline_counts.get(airline, 0) + 1
        
        # Count operations
        ops = 0
        if flight['arrivalFlightNumber']:
            ops += 1
        if flight['departureFlightNumber']:
            ops += 1
        airline_operations[airline] = airline_operations.get(airline, 0) + ops
        
        # Track departure and arrival passengers separately
        dep_pax = flight['departurePassengers'] or 0
        arr_pax = flight['arrivalPassengers'] or 0
        airline_departure_passengers[airline] = airline_departure_passengers.get(airline, 0) + dep_pax
        airline_arrival_passengers[airline] = airline_arrival_passengers.get(airline, 0) + arr_pax
        
        if passenger_type == "departure":
            airline_passengers[airline] = airline_passengers.get(airline, 0) + dep_pax
        elif passenger_type == "arrival":
            airline_passengers[airline] = airline_passengers.get(airline, 0) + arr_pax
        elif passenger_type == "all":
            airline_passengers[airline] = airline_passengers.get(airline, 0) + arr_pax + dep_pax
    
    if airline_counts:
        breakdown_fill = PatternFill(start_color="64748B", end_color="64748B", fill_type="solid")
        # Determine header span based on passenger type
        header_end_col = 7 if passenger_type == "all" else 5
        create_merged_cell(
            ws, row_idx, 1, header_end_col,
            "PREGLED PO AVIOKOMPANIJAMA",
            font=Font(bold=True, size=11, color="FFFFFF"),
            fill=breakdown_fill,
            alignment=Alignment(horizontal="center"),
            border=thin_border
        )
        row_idx += 1
        
        # Headers
        ws.cell(row=row_idx, column=1, value="Aviokompanija").font = Font(bold=True)
        ws.cell(row=row_idx, column=2, value="Letova").font = Font(bold=True)
        ws.cell(row=row_idx, column=3, value="Operacija").font = Font(bold=True)
        if passenger_type != "infants":
            if passenger_type == "all":
                ws.cell(row=row_idx, column=4, value="Odlazeći").font = Font(bold=True)
                ws.cell(row=row_idx, column=5, value="Dolazeći").font = Font(bold=True)
                ws.cell(row=row_idx, column=6, value="Ukupno").font = Font(bold=True)
                ws.cell(row=row_idx, column=7, value="Prosječno").font = Font(bold=True)
            else:
                ws.cell(row=row_idx, column=4, value="Putnici").font = Font(bold=True)
                ws.cell(row=row_idx, column=5, value="Prosječno").font = Font(bold=True)
        row_idx += 1
        
        # Sort airlines by passenger count (highest first)
        sorted_airlines = sorted(airline_counts.keys(), key=lambda x: airline_passengers.get(x, 0), reverse=True)
        for airline in sorted_airlines:
            ws.cell(row=row_idx, column=1, value=airline)
            ws.cell(row=row_idx, column=2, value=airline_counts[airline]).alignment = Alignment(horizontal="center")
            ws.cell(row=row_idx, column=3, value=airline_operations.get(airline, 0)).alignment = Alignment(horizontal="center")
            if passenger_type != "infants":
                if passenger_type == "all":
                    ws.cell(row=row_idx, column=4, value=airline_departure_passengers.get(airline, 0)).alignment = Alignment(horizontal="center")
                    ws.cell(row=row_idx, column=5, value=airline_arrival_passengers.get(airline, 0)).alignment = Alignment(horizontal="center")
                    ws.cell(row=row_idx, column=6, value=airline_passengers.get(airline, 0)).alignment = Alignment(horizontal="center")
                    avg = round(airline_passengers.get(airline, 0) / airline_operations.get(airline, 1), 1) if airline_operations.get(airline, 0) > 0 else 0
                    ws.cell(row=row_idx, column=7, value=avg).alignment = Alignment(horizontal="center")
                else:
                    ws.cell(row=row_idx, column=4, value=airline_passengers.get(airline, 0)).alignment = Alignment(horizontal="center")
                    avg = round(airline_passengers.get(airline, 0) / airline_operations.get(airline, 1), 1) if airline_operations.get(airline, 0) > 0 else 0
                    ws.cell(row=row_idx, column=5, value=avg).alignment = Alignment(horizontal="center")
            row_idx += 1
        
        row_idx += 1
    
    # Breakdown by route (ALL routes)
    route_counts = {}
    route_passengers = {}
    route_departure_passengers = {}
    route_arrival_passengers = {}
    route_operations = {}
    for flight in flights:
        route = flight['route'] or "Nepoznato"
        route_counts[route] = route_counts.get(route, 0) + 1
        
        # Count operations
        ops = 0
        if flight['arrivalFlightNumber']:
            ops += 1
        if flight['departureFlightNumber']:
            ops += 1
        route_operations[route] = route_operations.get(route, 0) + ops
        
        # Track departure and arrival passengers separately
        dep_pax = flight['departurePassengers'] or 0
        arr_pax = flight['arrivalPassengers'] or 0
        route_departure_passengers[route] = route_departure_passengers.get(route, 0) + dep_pax
        route_arrival_passengers[route] = route_arrival_passengers.get(route, 0) + arr_pax
        
        if passenger_type == "departure":
            route_passengers[route] = route_passengers.get(route, 0) + dep_pax
        elif passenger_type == "arrival":
            route_passengers[route] = route_passengers.get(route, 0) + arr_pax
        elif passenger_type == "all":
            route_passengers[route] = route_passengers.get(route, 0) + arr_pax + dep_pax
    
    if route_counts:
        route_fill = PatternFill(start_color="64748B", end_color="64748B", fill_type="solid")
        # Determine header span based on passenger type
        header_end_col = 7 if passenger_type == "all" else 5
        create_merged_cell(
            ws, row_idx, 1, header_end_col,
            "PREGLED PO RUTAMA",
            font=Font(bold=True, size=11, color="FFFFFF"),
            fill=route_fill,
            alignment=Alignment(horizontal="center"),
            border=thin_border
        )
        row_idx += 1
        
        # Headers
        ws.cell(row=row_idx, column=1, value="Ruta").font = Font(bold=True)
        ws.cell(row=row_idx, column=2, value="Letova").font = Font(bold=True)
        ws.cell(row=row_idx, column=3, value="Operacija").font = Font(bold=True)
        if passenger_type != "infants":
            if passenger_type == "all":
                ws.cell(row=row_idx, column=4, value="Odlazeći").font = Font(bold=True)
                ws.cell(row=row_idx, column=5, value="Dolazeći").font = Font(bold=True)
                ws.cell(row=row_idx, column=6, value="Ukupno").font = Font(bold=True)
                ws.cell(row=row_idx, column=7, value="Prosječno").font = Font(bold=True)
            else:
                ws.cell(row=row_idx, column=4, value="Putnici").font = Font(bold=True)
                ws.cell(row=row_idx, column=5, value="Prosječno").font = Font(bold=True)
        row_idx += 1
        
        # Sort by flight count and show ALL routes
        all_routes = sorted(route_counts.items(), key=lambda x: x[1], reverse=True)
        for route, count in all_routes:
            ws.cell(row=row_idx, column=1, value=route)
            ws.cell(row=row_idx, column=2, value=count).alignment = Alignment(horizontal="center")
            ws.cell(row=row_idx, column=3, value=route_operations.get(route, 0)).alignment = Alignment(horizontal="center")
            if passenger_type != "infants":
                if passenger_type == "all":
                    ws.cell(row=row_idx, column=4, value=route_departure_passengers.get(route, 0)).alignment = Alignment(horizontal="center")
                    ws.cell(row=row_idx, column=5, value=route_arrival_passengers.get(route, 0)).alignment = Alignment(horizontal="center")
                    ws.cell(row=row_idx, column=6, value=route_passengers.get(route, 0)).alignment = Alignment(horizontal="center")
                    avg = round(route_passengers.get(route, 0) / route_operations.get(route, 1), 1) if route_operations.get(route, 0) > 0 else 0
                    ws.cell(row=row_idx, column=7, value=avg).alignment = Alignment(horizontal="center")
                else:
                    ws.cell(row=row_idx, column=4, value=route_passengers.get(route, 0)).alignment = Alignment(horizontal="center")
                    avg = round(route_passengers.get(route, 0) / route_operations.get(route, 1), 1) if route_operations.get(route, 0) > 0 else 0
                    ws.cell(row=row_idx, column=5, value=avg).alignment = Alignment(horizontal="center")
            row_idx += 1
        
        row_idx += 1
    
    # Breakdown by operation type
    optype_counts = {}
    optype_passengers = {}
    optype_departure_passengers = {}
    optype_arrival_passengers = {}
    optype_operations = {}
    for flight in flights:
        optype = flight['operation_type_name'] or "Nepoznato"
        optype_counts[optype] = optype_counts.get(optype, 0) + 1
        
        # Count operations
        ops = 0
        if flight['arrivalFlightNumber']:
            ops += 1
        if flight['departureFlightNumber']:
            ops += 1
        optype_operations[optype] = optype_operations.get(optype, 0) + ops
        
        # Track departure and arrival passengers separately
        dep_pax = flight['departurePassengers'] or 0
        arr_pax = flight['arrivalPassengers'] or 0
        optype_departure_passengers[optype] = optype_departure_passengers.get(optype, 0) + dep_pax
        optype_arrival_passengers[optype] = optype_arrival_passengers.get(optype, 0) + arr_pax
        
        if passenger_type == "departure":
            optype_passengers[optype] = optype_passengers.get(optype, 0) + dep_pax
        elif passenger_type == "arrival":
            optype_passengers[optype] = optype_passengers.get(optype, 0) + arr_pax
        elif passenger_type == "all":
            optype_passengers[optype] = optype_passengers.get(optype, 0) + arr_pax + dep_pax
    
    if optype_counts:
        optype_fill = PatternFill(start_color="64748B", end_color="64748B", fill_type="solid")
        # Determine header span based on passenger type
        header_end_col = 7 if passenger_type == "all" else 5
        create_merged_cell(
            ws, row_idx, 1, header_end_col,
            "PREGLED PO TIPU SAOBRAĆAJA",
            font=Font(bold=True, size=11, color="FFFFFF"),
            fill=optype_fill,
            alignment=Alignment(horizontal="center"),
            border=thin_border
        )
        row_idx += 1
        
        # Headers
        ws.cell(row=row_idx, column=1, value="Tip saobraćaja").font = Font(bold=True)
        ws.cell(row=row_idx, column=2, value="Letova").font = Font(bold=True)
        ws.cell(row=row_idx, column=3, value="Operacija").font = Font(bold=True)
        if passenger_type != "infants":
            if passenger_type == "all":
                ws.cell(row=row_idx, column=4, value="Odlazeći").font = Font(bold=True)
                ws.cell(row=row_idx, column=5, value="Dolazeći").font = Font(bold=True)
                ws.cell(row=row_idx, column=6, value="Ukupno").font = Font(bold=True)
                ws.cell(row=row_idx, column=7, value="Prosječno").font = Font(bold=True)
            else:
                ws.cell(row=row_idx, column=4, value="Putnici").font = Font(bold=True)
                ws.cell(row=row_idx, column=5, value="Prosječno").font = Font(bold=True)
        row_idx += 1
        
        for optype in sorted(optype_counts.keys()):
            ws.cell(row=row_idx, column=1, value=optype)
            ws.cell(row=row_idx, column=2, value=optype_counts[optype]).alignment = Alignment(horizontal="center")
            ws.cell(row=row_idx, column=3, value=optype_operations.get(optype, 0)).alignment = Alignment(horizontal="center")
            if passenger_type != "infants":
                if passenger_type == "all":
                    ws.cell(row=row_idx, column=4, value=optype_departure_passengers.get(optype, 0)).alignment = Alignment(horizontal="center")
                    ws.cell(row=row_idx, column=5, value=optype_arrival_passengers.get(optype, 0)).alignment = Alignment(horizontal="center")
                    ws.cell(row=row_idx, column=6, value=optype_passengers.get(optype, 0)).alignment = Alignment(horizontal="center")
                    avg = round(optype_passengers.get(optype, 0) / optype_operations.get(optype, 1), 1) if optype_operations.get(optype, 0) > 0 else 0
                    ws.cell(row=row_idx, column=7, value=avg).alignment = Alignment(horizontal="center")
                else:
                    ws.cell(row=row_idx, column=4, value=optype_passengers.get(optype, 0)).alignment = Alignment(horizontal="center")
                    avg = round(optype_passengers.get(optype, 0) / optype_operations.get(optype, 1), 1) if optype_operations.get(optype, 0) > 0 else 0
                    ws.cell(row=row_idx, column=5, value=avg).alignment = Alignment(horizontal="center")
            row_idx += 1

    # Set row height for header
    ws.row_dimensions[1].height = 30

    return wb


def main():
    """Main function"""
    if len(sys.argv) < 2:
        print(json.dumps({
            "success": False,
            "error": "Missing filters argument"
        }))
        sys.exit(1)

    try:
        # Parse filters from command line argument
        filters = json.loads(sys.argv[1])
        
        date_from = filters.get('dateFrom')
        date_to = filters.get('dateTo')
        operation_types = filters.get('operationTypes', [])
        airlines = filters.get('airlines', [])
        routes = filters.get('routes', [])
        passenger_type = filters.get('passengerType', 'all')

        if not date_from or not date_to:
            print(json.dumps({
                "success": False,
                "error": "dateFrom and dateTo are required"
            }))
            sys.exit(1)

        # Fetch data
        flights = get_flight_data(date_from, date_to, operation_types, airlines, routes)

        # Generate Excel
        wb = create_excel_report(flights, passenger_type, date_from, date_to)

        # Save file
        OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d")
        filename = f"Custom_izvjestaj_{timestamp}.xlsx"
        output_path = OUTPUT_DIR / filename

        wb.save(output_path)

        print(json.dumps({
            "success": True,
            "fileName": filename,
            "message": f"Custom izvještaj uspješno generisan ({len(flights)} letova)"
        }))

    except Exception as e:
        print(json.dumps({
            "success": False,
            "error": str(e)
        }))
        sys.exit(1)


if __name__ == "__main__":
    main()
