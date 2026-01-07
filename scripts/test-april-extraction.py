#!/usr/bin/env python3

import openpyxl
import os

STATS_DIR = "/Users/emir_mw/stats/STATS/2025/Dnevni izvještaji"
month_folder = "04. APRIL"
month_path = os.path.join(STATS_DIR, month_folder)

print(f"Checking folder: {month_path}")
print(f"Folder exists: {os.path.exists(month_path)}")
print(f"Is directory: {os.path.isdir(month_path)}")

files = os.listdir(month_path)
print(f"\nAll files: {files}")

# Check with pattern
pattern_files = [
    f for f in files
    if 'Dnevni izvještaj o saobraćaju' in f and f.endswith('.xlsx')
]
print(f"Pattern matched files: {pattern_files}")

if pattern_files:
    file_path = os.path.join(month_path, pattern_files[0])
    print(f"\nTrying to open: {file_path}")

    try:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        print(f"Successfully opened! Sheets: {len(wb.sheetnames)}")
        print(f"Sheet names: {wb.sheetnames[:5]}...")  # First 5 sheets
        wb.close()
    except Exception as e:
        print(f"Error opening: {e}")
