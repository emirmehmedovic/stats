#!/usr/bin/env python3
"""
Statistika za direktora - Director Statistics Report Generator

Generiše godišnji izvještaj sa mjesečnim tabelama od januara do odabranog mjeseca,
uz završni summary ukupnih statistika.
"""

import sys
import os
import re
import calendar
from pathlib import Path
import psycopg2
from psycopg2.extras import RealDictCursor
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# Putanja
SCRIPT_DIR = Path(__file__).parent
PROJECT_ROOT = SCRIPT_DIR.parent
OUTPUT_DIR = PROJECT_ROOT / "izvještaji" / "generated"

# Nazivi mjeseci
MONTH_NAMES = {
    1: "Januar", 2: "Februar", 3: "Mart", 4: "April",
    5: "Maj", 6: "Juni", 7: "Juli", 8: "Avgust",
    9: "Septembar", 10: "Oktobar", 11: "Novembar", 12: "Decembar"
}


def get_db_connection():
    """Konekcija na PostgreSQL bazu"""
    DATABASE_URL = os.getenv("DATABASE_URL")
    if not DATABASE_URL:
        raise ValueError("DATABASE_URL environment variable not set")

    return psycopg2.connect(DATABASE_URL, cursor_factory=RealDictCursor)


def sanitize_text(value):
    """
    Gentle sanitization that preserves unicode (ć, č, š, ž, đ) but removes Excel-problematic characters.
    openpyxl 3.1.5+ properly handles unicode, so we only need to remove control characters.
    """
    if value is None:
        return None
    if not isinstance(value, str):
        return value

    # Step 1: Remove control characters (0x00-0x1F, 0x7F-0x9F)
    value = re.sub(r'[\x00-\x1F\x7F-\x9F]', '', value)

    # Step 2: Use openpyxl's built-in cleaner for illegal Excel characters
    try:
        from openpyxl.utils.cell import ILLEGAL_CHARACTERS_RE
        value = ILLEGAL_CHARACTERS_RE.sub('', value)
    except Exception:
        pass

    # Step 3: Remove zero-width spaces and BOM
    value = value.replace('\ufeff', '').replace('\u200b', '').replace('\u200c', '').replace('\u200d', '')

    # Step 4: Replace non-breaking space with regular space
    value = value.replace('\xa0', ' ')

    # Step 5: Strip whitespace
    return value.strip()


def create_merged_cell(ws, row, start_col, end_col, value, font=None, fill=None, alignment=None, border=None):
    """
    Helper function to properly create merged cells with consistent formatting.
    Applies formatting to ALL cells in the range before merging to avoid Excel corruption.
    """
    # Convert column letters to numbers if needed
    if isinstance(start_col, str):
        start_col = openpyxl.utils.column_index_from_string(start_col)
    if isinstance(end_col, str):
        end_col = openpyxl.utils.column_index_from_string(end_col)

    for col in range(start_col, end_col + 1):
        cell = ws.cell(row=row, column=col)
        if col == start_col and value is not None:
            cell.value = sanitize_text(value) if isinstance(value, str) else value
        if font:
            cell.font = font
        if fill:
            cell.fill = fill
        if alignment:
            cell.alignment = alignment
        if border:
            cell.border = border

    if start_col != end_col:
        ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=end_col)


def get_flight_data(year: int, month: int):
    """Povuči sve letove za zadati mjesec"""
    conn = get_db_connection()
    cursor = conn.cursor()

    first_day = f"{year}-{month:02d}-01"
    last_day = calendar.monthrange(year, month)[1]
    last_date = f"{year}-{month:02d}-{last_day:02d}"

    query = """
        SELECT
            f.id,
            f.date,

            -- Airline
            a.name as airline_name,
            a."icaoCode" as airline_icao,
            a."iataCode" as airline_iata,

            -- Operation Type
            ot.code as operation_type_code,

            -- Departure info
            f."departurePassengers",
            f."departureCargo",
            f."departureMail",
            f."departureStatus",
            f."departureFerryOut",

            -- Arrival info
            f."arrivalPassengers",
            f."arrivalCargo",
            f."arrivalMail",
            f."arrivalStatus",
            f."arrivalFerryIn"

        FROM "Flight" f
        INNER JOIN "Airline" a ON f."airlineId" = a.id
        INNER JOIN "OperationType" ot ON f."operationTypeId" = ot.id
        WHERE f.date >= %s AND f.date <= %s
        ORDER BY a.name
    """

    cursor.execute(query, (first_day, last_date))
    flights = cursor.fetchall()

    cursor.close()
    conn.close()

    return flights


def normalize_airline_name(flight):
    """Group Wizz Air entities under a single name."""
    name = (flight.get('airline_name') or '').strip()
    icao = (flight.get('airline_icao') or '').strip().upper()
    iata = (flight.get('airline_iata') or '').strip().upper()

    if 'WIZZ' in name.upper() or icao in {'WZZ', 'WMT', 'WUK'} or iata in {'W6', 'W4', 'W9'}:
        return 'Wizz Air'

    return name


def aggregate_customs_data(flights):
    """Agregira podatke za statistiku carinskog izvještaja"""
    data = {
        'scheduled': {},
        'non_scheduled': {'flights': 0, 'embarked': 0, 'disembarked': 0},
        'other_landings': {'flights': 0, 'embarked': 0, 'disembarked': 0},
        'freight': {'loaded': 0.0, 'unloaded': 0.0}
    }

    for flight in flights:
        operation_code = flight.get('operation_type_code', '').upper()
        is_scheduled = operation_code == 'SCHEDULED'
        airline_name = normalize_airline_name(flight)

        is_ferry = flight.get('arrivalFerryIn') or flight.get('departureFerryOut')

        has_arrival = flight.get('arrivalPassengers') is not None or flight.get('arrivalStatus') == 'OPERATED'
        has_departure = flight.get('departurePassengers') is not None or flight.get('departureStatus') == 'OPERATED'

        if is_ferry:
            category = 'other_landings'
        elif is_scheduled:
            category = 'scheduled'
        else:
            category = 'non_scheduled'

        if category == 'scheduled':
            if airline_name not in data['scheduled']:
                data['scheduled'][airline_name] = {'flights': 0, 'embarked': 0, 'disembarked': 0}

            if has_arrival:
                data['scheduled'][airline_name]['flights'] += 1
            if has_departure and not has_arrival:
                data['scheduled'][airline_name]['flights'] += 1

            if flight.get('departurePassengers'):
                data['scheduled'][airline_name]['embarked'] += flight['departurePassengers']
            if flight.get('arrivalPassengers'):
                data['scheduled'][airline_name]['disembarked'] += flight['arrivalPassengers']
        else:
            target = data[category]
            if has_arrival:
                target['flights'] += 1
            if has_departure and not has_arrival:
                target['flights'] += 1

            if flight.get('departurePassengers'):
                target['embarked'] += flight['departurePassengers']
            if flight.get('arrivalPassengers'):
                target['disembarked'] += flight['arrivalPassengers']

        if flight.get('departureCargo'):
            data['freight']['loaded'] += flight['departureCargo'] / 1000.0
        if flight.get('arrivalCargo'):
            data['freight']['unloaded'] += flight['arrivalCargo'] / 1000.0

        if flight.get('departureMail'):
            data['freight']['loaded'] += flight['departureMail'] / 1000.0
        if flight.get('arrivalMail'):
            data['freight']['unloaded'] += flight['arrivalMail'] / 1000.0

    data['freight']['loaded'] = round(data['freight']['loaded'], 2)
    data['freight']['unloaded'] = round(data['freight']['unloaded'], 2)

    return data


def combine_customs_data(total, current):
    """Saberi mjesečne podatke u ukupne podatke"""
    for airline_name, values in current['scheduled'].items():
        if airline_name not in total['scheduled']:
            total['scheduled'][airline_name] = {'flights': 0, 'embarked': 0, 'disembarked': 0}
        total['scheduled'][airline_name]['flights'] += values['flights']
        total['scheduled'][airline_name]['embarked'] += values['embarked']
        total['scheduled'][airline_name]['disembarked'] += values['disembarked']

    for key in ['non_scheduled', 'other_landings']:
        total[key]['flights'] += current[key]['flights']
        total[key]['embarked'] += current[key]['embarked']
        total[key]['disembarked'] += current[key]['disembarked']

    total['freight']['loaded'] += current['freight']['loaded']
    total['freight']['unloaded'] += current['freight']['unloaded']


def add_customs_block(ws, row, title, data, styles):
    """Dodaj jedan block statistike u worksheet"""
    title_font, section_font, header_font, regular_font, indent_font, bold_font = styles
    center_align, right_align = Alignment(horizontal='center', vertical='center'), Alignment(horizontal='right', vertical='center')
    header_fill = PatternFill(start_color="E5E7EB", end_color="E5E7EB", fill_type="solid")
    total_fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
    zebra_fill = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")

    def apply_outline_border(start_row, end_row, start_col, end_col):
        side = Side(style='thin')
        for r in range(start_row, end_row + 1):
            for c in range(start_col, end_col + 1):
                cell = ws.cell(row=r, column=c)
                border = Border(
                    top=side if r == start_row else None,
                    bottom=side if r == end_row else None,
                    left=side if c == start_col else None,
                    right=side if c == end_col else None
                )
                cell.border = border

    create_merged_cell(
        ws, row, 'A', 'E',
        title,
        font=title_font,
        alignment=center_align
    )
    row += 2

    create_merged_cell(
        ws, row, 'B', 'E',
        "PUTNICI",
        font=section_font,
        alignment=center_align
    )
    row += 1

    ws[f'B{row}'] = "Broj letova"
    ws[f'C{row}'] = "Ukrcano"
    ws[f'D{row}'] = "Iskrcano"
    ws[f'E{row}'] = "Ukupno"
    for col in ['B', 'C', 'D', 'E']:
        ws[f'{col}{row}'].font = header_font
        ws[f'{col}{row}'].alignment = center_align
        ws[f'{col}{row}'].fill = header_fill
    passengers_header_row = row
    row += 1

    scheduled_total_flights = sum(airline['flights'] for airline in data['scheduled'].values())
    scheduled_total_embarked = sum(airline['embarked'] for airline in data['scheduled'].values())
    scheduled_total_disembarked = sum(airline['disembarked'] for airline in data['scheduled'].values())
    scheduled_total = scheduled_total_embarked + scheduled_total_disembarked

    ws[f'A{row}'] = "Redovni promet"
    ws[f'A{row}'].font = bold_font
    ws[f'B{row}'] = scheduled_total_flights
    ws[f'C{row}'] = scheduled_total_embarked
    ws[f'D{row}'] = scheduled_total_disembarked
    ws[f'E{row}'] = scheduled_total
    for col in ['B', 'C', 'D', 'E']:
        ws[f'{col}{row}'].font = regular_font
        ws[f'{col}{row}'].alignment = right_align
        ws[f'{col}{row}'].number_format = '#,##0'
    for col in ['A', 'B', 'C', 'D', 'E']:
        ws[f'{col}{row}'].fill = total_fill
    row += 1

    airline_rows = sorted(
        data['scheduled'].items(),
        key=lambda item: (item[1]['embarked'] + item[1]['disembarked'], item[0]),
        reverse=True
    )

    for idx, (airline_name, airline_data) in enumerate(airline_rows):
        ws[f'A{row}'] = f"  {airline_name}"
        ws[f'A{row}'].font = indent_font
        ws[f'B{row}'] = airline_data['flights']
        ws[f'C{row}'] = airline_data['embarked']
        ws[f'D{row}'] = airline_data['disembarked']
        ws[f'E{row}'] = airline_data['embarked'] + airline_data['disembarked']
        for col in ['B', 'C', 'D', 'E']:
            ws[f'{col}{row}'].font = regular_font
            ws[f'{col}{row}'].alignment = right_align
            ws[f'{col}{row}'].number_format = '#,##0'
        if idx % 2 == 0:
            for col in ['A', 'B', 'C', 'D', 'E']:
                ws[f'{col}{row}'].fill = zebra_fill
        row += 1

    ws[f'A{row}'] = "Vanredni promet"
    ws[f'A{row}'].font = bold_font
    ws[f'B{row}'] = data['non_scheduled']['flights']
    ws[f'C{row}'] = data['non_scheduled']['embarked']
    ws[f'D{row}'] = data['non_scheduled']['disembarked']
    ws[f'E{row}'] = data['non_scheduled']['embarked'] + data['non_scheduled']['disembarked']
    for col in ['B', 'C', 'D', 'E']:
        ws[f'{col}{row}'].font = regular_font
        ws[f'{col}{row}'].alignment = right_align
        ws[f'{col}{row}'].number_format = '#,##0'
    for col in ['A', 'B', 'C', 'D', 'E']:
        ws[f'{col}{row}'].fill = total_fill
    row += 1

    ws[f'A{row}'] = "Ostala slijetanja"
    ws[f'A{row}'].font = bold_font
    ws[f'B{row}'] = data['other_landings']['flights']
    ws[f'C{row}'] = data['other_landings']['embarked']
    ws[f'D{row}'] = data['other_landings']['disembarked']
    ws[f'E{row}'] = data['other_landings']['embarked'] + data['other_landings']['disembarked']
    for col in ['B', 'C', 'D', 'E']:
        ws[f'{col}{row}'].font = regular_font
        ws[f'{col}{row}'].alignment = right_align
        ws[f'{col}{row}'].number_format = '#,##0'
    for col in ['A', 'B', 'C', 'D', 'E']:
        ws[f'{col}{row}'].fill = total_fill
    row += 2

    total_flights = scheduled_total_flights + data['non_scheduled']['flights'] + data['other_landings']['flights']
    total_embarked = scheduled_total_embarked + data['non_scheduled']['embarked'] + data['other_landings']['embarked']
    total_disembarked = scheduled_total_disembarked + data['non_scheduled']['disembarked'] + data['other_landings']['disembarked']
    total_passengers = total_embarked + total_disembarked

    ws[f'A{row}'] = "UKUPNO"
    ws[f'A{row}'].font = bold_font
    ws[f'B{row}'] = total_flights
    ws[f'C{row}'] = total_embarked
    ws[f'D{row}'] = total_disembarked
    ws[f'E{row}'] = total_passengers
    for col in ['B', 'C', 'D', 'E']:
        ws[f'{col}{row}'].font = bold_font
        ws[f'{col}{row}'].alignment = right_align
        ws[f'{col}{row}'].number_format = '#,##0'
    for col in ['A', 'B', 'C', 'D', 'E']:
        ws[f'{col}{row}'].fill = total_fill
    passengers_end_row = row
    row += 3

    create_merged_cell(
        ws, row, 'B', 'D',
        "TERET",
        font=section_font,
        alignment=center_align
    )
    row += 1

    ws[f'B{row}'] = "Utovareno"
    ws[f'C{row}'] = "Istovareno"
    ws[f'D{row}'] = "Ukupno"
    for col in ['B', 'C', 'D']:
        ws[f'{col}{row}'].font = header_font
        ws[f'{col}{row}'].alignment = center_align
        ws[f'{col}{row}'].fill = header_fill
    freight_header_row = row
    row += 1

    ws[f'B{row}'] = data['freight']['loaded']
    ws[f'C{row}'] = data['freight']['unloaded']
    ws[f'D{row}'] = data['freight']['loaded'] + data['freight']['unloaded']
    for col in ['B', 'C', 'D']:
        ws[f'{col}{row}'].font = regular_font
        ws[f'{col}{row}'].alignment = right_align
        ws[f'{col}{row}'].number_format = '#,##0.00'
    freight_end_row = row

    apply_outline_border(passengers_header_row, passengers_end_row, 1, 5)
    apply_outline_border(freight_header_row, freight_end_row, 2, 4)

    row += 3
    return row


def generate_director_stats(year: int, month_to: int, output_path: Path = None):
    """Glavni metod za generisanje Director statistike"""
    print(f"Generišem statistiku za direktora za period Januar-{MONTH_NAMES[month_to]} {year}...")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Statistika"

    title_font = Font(name='Arial', size=14, bold=True)
    section_font = Font(name='Arial', size=11, bold=True)
    header_font = Font(name='Arial', size=10, bold=True)
    regular_font = Font(name='Arial', size=10)
    indent_font = Font(name='Arial', size=10, italic=True)
    bold_font = Font(name='Arial', size=10, bold=True)

    styles = (title_font, section_font, header_font, regular_font, indent_font, bold_font)

    total_data = {
        'scheduled': {},
        'non_scheduled': {'flights': 0, 'embarked': 0, 'disembarked': 0},
        'other_landings': {'flights': 0, 'embarked': 0, 'disembarked': 0},
        'freight': {'loaded': 0.0, 'unloaded': 0.0}
    }

    total_flights_all = 0
    row = 1

    for month in range(1, month_to + 1):
        print(f"Povlačim podatke za {MONTH_NAMES[month]} {year}...")
        flights = get_flight_data(year, month)
        total_flights_all += len(flights)
        data = aggregate_customs_data(flights)
        combine_customs_data(total_data, data)
        row = add_customs_block(ws, row, f"{MONTH_NAMES[month]} {year}", data, styles)

    if total_flights_all == 0:
        print("UPOZORENJE: Nema letova za zadati period!")
        return None

    total_data['freight']['loaded'] = round(total_data['freight']['loaded'], 2)
    total_data['freight']['unloaded'] = round(total_data['freight']['unloaded'], 2)

    summary_title = f"UKUPNO (Januar - {MONTH_NAMES[month_to]} {year})"
    row = add_customs_block(ws, row, summary_title, total_data, styles)

    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15

    if output_path is None:
        OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
        if month_to == 1:
            file_suffix = f"{MONTH_NAMES[month_to]}_{year}"
        else:
            file_suffix = f"Januar-{MONTH_NAMES[month_to]}_{year}"
        output_path = OUTPUT_DIR / f"Statistika_za_direktora_{file_suffix}.xlsx"

    print(f"Čuvam izvještaj u: {output_path}")
    wb.iso_dates = True
    wb.save(output_path)
    wb.close()
    print("✅ Statistika za direktora uspješno generisana!")

    return str(output_path)


if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("Usage: python generate_director_stats.py <year> <month_to>")
        print("Example: python generate_director_stats.py 2025 3")
        sys.exit(1)

    year = int(sys.argv[1])
    month_to = int(sys.argv[2])

    if month_to < 1 or month_to > 12:
        print("Mjesec mora biti između 1 i 12")
        sys.exit(1)

    try:
        output_file = generate_director_stats(year, month_to)
        print(f"\nIzvještaj sačuvan: {output_file}")
    except Exception as e:
        print(f"GREŠKA: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
