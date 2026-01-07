#!/usr/bin/env python3

import openpyxl
import os

STATS_DIR = "/Users/emir_mw/stats/STATS/2025/Dnevni izvještaji"
month_folder = "01. JANUAR"
month_path = os.path.join(STATS_DIR, month_folder)

files = [
    f for f in os.listdir(month_path)
    if 'Dnevni izvještaj o saobraćaju' in f and f.endswith('.xlsx')
]

file_path = os.path.join(month_path, files[0])

wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)

# Specific dates to check
check_dates = [
    '2024-12-31',  # FLYNAS JED-TZL
    '2024-01-03',  # Three flights
    '2025-01-16',  # WIZZAIR MALTA LTD PAD-TZL
    '2025-01-17',  # ELITAVIA
    '2025-01-21',  # FLYNAS
]

print("🔍 Checking flights with missing passenger data in Excel:\n")

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]

    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if i == 1:  # Skip header
            continue

        if not row[0]:  # Skip empty rows
            continue

        date_value = row[0]
        airline_name = row[1]
        route = row[2]
        arrival_pax = row[13] if len(row) > 13 else None
        departure_pax = row[14] if len(row) > 14 else None

        # Convert date
        date_str = str(date_value).split(' ')[0] if date_value else ''

        # Check if this is one of our problem flights
        if any(check_date in date_str for check_date in check_dates):
            print(f"Date: {date_str}")
            print(f"  Airline: {airline_name}")
            print(f"  Route: {route}")
            print(f"  Arrival Passengers (col 13): {arrival_pax}")
            print(f"  Departure Passengers (col 14): {departure_pax}")
            print()

wb.close()

# Also check the JSON file
print("\n" + "="*80)
print("Checking extracted JSON data:")
print("="*80 + "\n")

import json

with open('output/2025-flights-data.json', 'r') as f:
    data = json.load(f)

flights = data['flights']
print(f"Total flights in JSON: {len(flights)}")

# Count flights with passenger data
with_arrival = sum(1 for f in flights if f.get('arrivalPassengers') is not None)
with_departure = sum(1 for f in flights if f.get('departurePassengers') is not None)
without_any = sum(1 for f in flights if f.get('arrivalPassengers') is None and f.get('departurePassengers') is None)

print(f"Flights with arrival data: {with_arrival}")
print(f"Flights with departure data: {with_departure}")
print(f"Flights WITHOUT any passenger data: {without_any}")

# Show the flights without data
if without_any > 0:
    print(f"\nFlights without passenger data in JSON:")
    for f in flights:
        if f.get('arrivalPassengers') is None and f.get('departurePassengers') is None:
            print(f"  {f['date'][:10]} - {f['airline']} - {f['route']}")
