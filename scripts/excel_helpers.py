#!/usr/bin/env python3
"""
Ultra-safe Excel helper functions for all report generators
"""
import re
import unicodedata
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill


def sanitize_text(value):
    """
    Remove ALL characters that can corrupt Excel XML.
    """
    if value is None:
        return None
    if not isinstance(value, str):
        return value

    # Normalize unicode
    try:
        value = unicodedata.normalize('NFC', value)
    except Exception:
        pass

    # Remove control characters except tab, newline, carriage return
    value = re.sub(r'[\x00-\x08\x0B\x0C\x0E-\x1F\x7F-\x9F]', '', value)

    # Try openpyxl's built-in cleaner
    try:
        from openpyxl.utils.cell import ILLEGAL_CHARACTERS_RE
        value = ILLEGAL_CHARACTERS_RE.sub('', value)
    except Exception:
        pass

    # Remove null bytes and BOM
    value = value.replace('\x00', '').replace('\ufeff', '')

    # Ensure valid UTF-8
    try:
        value = value.encode('utf-8', errors='ignore').decode('utf-8')
    except Exception:
        pass

    return value


def safe_set_cell_value(cell, value):
    """
    Safely set cell value with proper sanitization.
    """
    if value is None:
        cell.value = None
        return

    if isinstance(value, str):
        cell.value = sanitize_text(value)
    elif isinstance(value, (int, float)):
        import math
        if isinstance(value, float) and (math.isnan(value) or math.isinf(value)):
            cell.value = 0
        else:
            cell.value = value
    elif hasattr(value, 'strftime'):  # datetime
        cell.value = value
    elif isinstance(value, bool):
        cell.value = value
    else:
        cell.value = sanitize_text(str(value))


def create_merged_cell_safe(ws, row, start_col, end_col, value, font=None, fill=None, alignment=None, border=None):
    """
    Ultra-safe merged cell creation.
    """
    # Convert column letters to numbers if needed
    if isinstance(start_col, str):
        start_col = openpyxl.utils.column_index_from_string(start_col)
    if isinstance(end_col, str):
        end_col = openpyxl.utils.column_index_from_string(end_col)

    # First, apply formatting and value to ALL cells BEFORE merging
    for col in range(start_col, end_col + 1):
        cell = ws.cell(row=row, column=col)

        # Set value only in first cell
        if col == start_col:
            safe_set_cell_value(cell, value)
        else:
            # Other cells should be empty before merging
            cell.value = None

        # Apply formatting to ALL cells
        if font:
            cell.font = font
        if fill:
            cell.fill = fill
        if alignment:
            cell.alignment = alignment
        if border:
            cell.border = border

    # NOW merge the cells
    if start_col != end_col:
        try:
            ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=end_col)
        except Exception as e:
            print(f"[WARNING] Could not merge cells row={row}, cols={start_col}:{end_col}: {e}")


def save_workbook_safe(workbook, output_path):
    """
    Safely save workbook with proper settings.
    """
    try:
        # Remove any existing file
        import os
        if os.path.exists(output_path):
            os.remove(output_path)

        # Save with explicit ISO dates
        workbook.iso_dates = True

        # Save the workbook
        workbook.save(output_path)

        # Close workbook to free memory
        workbook.close()

        return True
    except Exception as e:
        print(f"[ERROR] Could not save workbook: {e}")
        return False
