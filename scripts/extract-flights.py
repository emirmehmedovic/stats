#!/usr/bin/env python3

import openpyxl
import os
import sys
import json
import re
from datetime import datetime

# Get year from command line argument or use default
YEAR = sys.argv[1] if len(sys.argv) > 1 else "2025"

# Try both directory structures
STATS_DIR = f"/Users/emir_mw/stats/STATS/{YEAR}/Dnevni izvještaji"
if not os.path.exists(STATS_DIR):
    STATS_DIR = f"/Users/emir_mw/stats/STATS/{YEAR}/Mjesečni izvještaji"

def parse_passengers(value):
    """Parse passenger string like '165+6 INF' or '110' or '-'"""
    if not value or value == '-' or value == 'N/A' or str(value).strip() == '':
        return None

    str_val = str(value).strip()

    # Pattern: "110+6 INF" or "110+6" (with infants)
    match = re.match(r'^(\d+)\s*\+\s*(\d+)', str_val)
    if match:
        adults = int(match.group(1))
        infants = int(match.group(2))
        return {
            'adults': adults,
            'infants': infants,
            'total': adults + infants
        }

    # Pattern: just a number "110" (no infants)
    match = re.match(r'^(\d+)$', str_val)
    if match:
        adults = int(match.group(1))
        return {
            'adults': adults,
            'infants': 0,
            'total': adults
        }

    return None

def parse_route(route):
    """Parse route like 'TZL-AYT' and return [departure, arrival]"""
    if not route or route == '-' or route == 'N/A':
        return None

    parts = [p.strip() for p in route.split('-') if p.strip()]

    if len(parts) >= 2:
        return {
            'departure': parts[0],
            'arrival': parts[-1],
        }

    return None

def detect_format(header_row):
    """
    Detect Excel format based on header row.
    Returns format info dict with column mappings and format type.
    """

    format_info = {}

    # Check for 2023 format: has "Datum" in column 0, aircraft type in column 2, and passengers at 10/17
    # Note: December 2023 has "datum" and "kompanija" but "ruta" in col 2, so it's NOT Format 5
    if (len(header_row) > 17 and
        header_row[0] and 'datum' in str(header_row[0]).lower() and
        header_row[1] and 'kompanija' in str(header_row[1]).lower() and
        header_row[2] and 'tip' in str(header_row[2]).lower() and 'a/c' in str(header_row[2]).lower() and
        header_row[10] and 'putnik' in str(header_row[10]).lower()):
        # Format 5: 2023 Jan-Nov (has "Datum" in first column, "Tip A/C" in col 2, passengers in format "171+6")
        format_info['type'] = 5
        format_info['has_date_column'] = True
        format_info['airline_col'] = 1
        format_info['aircraft_col'] = 2
        format_info['available_seats_col'] = None
        format_info['registration_col'] = 3
        format_info['mtow_col'] = 4
        format_info['operation_type_col'] = 5
        format_info['route_col'] = [6, 13]  # "Dolazak iz" at 6, "Odlazak za" at 13
        format_info['arrival_flight_number_col'] = 7
        format_info['departure_flight_number_col'] = 14
        format_info['scheduled_arrival_time_col'] = 8
        format_info['actual_arrival_time_col'] = 9
        format_info['scheduled_departure_time_col'] = 15
        format_info['actual_departure_time_col'] = 16
        format_info['arrival_pax_col'] = 10
        format_info['departure_pax_col'] = 17
        format_info['arrival_inf_col'] = None  # Combined in pax col as "171+6"
        format_info['departure_inf_col'] = None  # Combined in pax col
        format_info['arrival_baggage_col'] = 11
        format_info['departure_baggage_col'] = 18
        format_info['arrival_cargo_col'] = 12
        format_info['departure_cargo_col'] = 19
        format_info['arrival_mail_col'] = None
        format_info['departure_mail_col'] = None
        format_info['use_parse_passengers'] = True
        return format_info

    # Find route column
    route_col = None
    for i, cell in enumerate(header_row):
        if cell and 'ruta' in str(cell).lower():
            route_col = i
            break

    # Check if ICAO column exists
    icao_col = None
    for i, cell in enumerate(header_row):
        if cell and 'icao' in str(cell).lower():
            icao_col = i
            break

    # Check for duplicate date/company columns (July format)
    has_duplicate_cols = False
    if (len(header_row) > 3 and
        header_row[0] and header_row[2] and
        'datum' in str(header_row[0]).lower() and
        'datum' in str(header_row[2]).lower()):
        has_duplicate_cols = True

    # Find passenger columns
    pax_cols = []
    inf_cols = []

    for i, cell in enumerate(header_row):
        cell_str = str(cell).lower() if cell else ''
        if 'putnik' in cell_str:
            pax_cols.append(i)
        if 'beb' in cell_str:
            inf_cols.append(i)

    format_info = {'route_col': route_col}

    # Format 1: January, February (no ICAO, passengers in [13], [14] as "165+6 INF")
    if icao_col is None:
        format_info['type'] = 1
        format_info['airline_col'] = 1
        format_info['aircraft_col'] = 3
        format_info['available_seats_col'] = None  # No available seats column in Format 1
        format_info['registration_col'] = 4
        format_info['operation_type_col'] = 5
        format_info['mtow_col'] = 6
        format_info['arrival_flight_number_col'] = 7
        format_info['departure_flight_number_col'] = 8
        format_info['scheduled_arrival_time_col'] = 9   # pl vrijeme dol
        format_info['actual_arrival_time_col'] = 10     # st vrijeme dol
        format_info['scheduled_departure_time_col'] = 11  # pl vrijeme odl
        format_info['actual_departure_time_col'] = 12   # st vrijeme odl
        format_info['arrival_pax_col'] = 13
        format_info['departure_pax_col'] = 14
        format_info['arrival_inf_col'] = None  # Combined in pax col
        format_info['departure_inf_col'] = None  # Combined in pax col
        format_info['arrival_baggage_col'] = 15
        format_info['departure_baggage_col'] = 16
        format_info['arrival_cargo_col'] = 17
        format_info['departure_cargo_col'] = 18
        format_info['arrival_mail_col'] = 19
        format_info['departure_mail_col'] = 20
        format_info['use_parse_passengers'] = True
        return format_info

    # Format 4: July (duplicate columns, ICAO at [4], route at [5])
    if has_duplicate_cols:
        format_info['type'] = 4
        format_info['airline_col'] = 3  # Use the second "Kompanija" column
        format_info['aircraft_col'] = 6
        format_info['available_seats_col'] = 7  # Rasp. mjesta
        format_info['registration_col'] = 8
        format_info['operation_type_col'] = 9
        format_info['mtow_col'] = 10
        format_info['arrival_flight_number_col'] = 11
        format_info['departure_flight_number_col'] = 19
        format_info['scheduled_arrival_time_col'] = 12   # pl vrijeme dol
        format_info['actual_arrival_time_col'] = 13     # st vrijeme dol
        format_info['scheduled_departure_time_col'] = 20  # pl vrijeme odl
        format_info['actual_departure_time_col'] = 21   # st vrijeme odl
        format_info['arrival_pax_col'] = 14  # Adults only
        format_info['departure_pax_col'] = 22  # Adults only
        format_info['arrival_inf_col'] = 15  # Infants
        format_info['departure_inf_col'] = 23  # Infants
        format_info['arrival_baggage_col'] = 16
        format_info['departure_baggage_col'] = 24
        format_info['arrival_cargo_col'] = 17
        format_info['departure_cargo_col'] = 25
        format_info['arrival_mail_col'] = 18
        format_info['departure_mail_col'] = 26
        format_info['use_parse_passengers'] = False
        return format_info

    # Format 2: March (has ICAO, passengers in [15], [16] as "165+6 INF")
    if len(pax_cols) >= 2 and pax_cols[0] == 15:
        format_info['type'] = 2
        format_info['airline_col'] = 1
        format_info['aircraft_col'] = 4
        format_info['available_seats_col'] = 5  # Rasp. mjesta
        format_info['registration_col'] = 6
        format_info['operation_type_col'] = 7
        format_info['mtow_col'] = 8
        format_info['arrival_flight_number_col'] = 9
        format_info['departure_flight_number_col'] = 10
        format_info['scheduled_arrival_time_col'] = 11   # pl vrijeme dol
        format_info['actual_arrival_time_col'] = 12     # st vrijeme dol
        format_info['scheduled_departure_time_col'] = 13  # pl vrijeme odl
        format_info['actual_departure_time_col'] = 14   # st vrijeme odl
        format_info['arrival_pax_col'] = 15
        format_info['departure_pax_col'] = 16
        format_info['arrival_inf_col'] = None  # Combined in pax col
        format_info['departure_inf_col'] = None  # Combined in pax col
        format_info['arrival_baggage_col'] = 17
        format_info['departure_baggage_col'] = 18
        format_info['arrival_cargo_col'] = 19
        format_info['departure_cargo_col'] = 20
        format_info['arrival_mail_col'] = 21
        format_info['departure_mail_col'] = 22
        format_info['use_parse_passengers'] = True
        return format_info

    # Format 3: April-June, August-November (has ICAO, passengers split)
    # [12] Putnici u avionu (arrival), [13] Bebe (arrival infants)
    # [20] Putnici u avionu (departure), [21] Bebe (departure infants)
    format_info['type'] = 3
    format_info['airline_col'] = 1
    format_info['aircraft_col'] = 4
    format_info['available_seats_col'] = 5  # Rasp. mjesta
    format_info['registration_col'] = 6
    format_info['operation_type_col'] = 7
    format_info['mtow_col'] = 8
    format_info['arrival_flight_number_col'] = 9
    format_info['departure_flight_number_col'] = 17
    format_info['scheduled_arrival_time_col'] = 10   # pl vrijeme dol
    format_info['actual_arrival_time_col'] = 11     # st vrijeme dol
    format_info['scheduled_departure_time_col'] = 18  # pl vrijeme odl
    format_info['actual_departure_time_col'] = 19   # st vrijeme odl
    format_info['arrival_pax_col'] = 12  # Adults only
    format_info['departure_pax_col'] = 20  # Adults only
    format_info['arrival_inf_col'] = 13  # Infants
    format_info['departure_inf_col'] = 21  # Infants
    format_info['arrival_baggage_col'] = 14
    format_info['departure_baggage_col'] = 22
    format_info['arrival_cargo_col'] = 15
    format_info['departure_cargo_col'] = 23
    format_info['arrival_mail_col'] = 16
    format_info['departure_mail_col'] = 24
    format_info['use_parse_passengers'] = False

    return format_info

def safe_get(row, col):
    """Safely get a value from row at column index"""
    if col is None:
        return None
    if len(row) > col:
        return row[col]
    return None

def extract_month(month_folder):
    """Extract flight data from one month"""
    month_path = os.path.join(STATS_DIR, month_folder)

    # Find Excel file (match both 2025 and 2024 naming conventions)
    files = [
        f for f in os.listdir(month_path)
        if f.endswith('.xlsx') and not f.startswith('~')  # Exclude temp files
    ]

    if not files:
        print(f"   ⚠️  No Excel file found in {month_folder}")
        return []

    # Prefer files without "STATISTIKA" in name (those are summary files, not flight data)
    non_stat_files = [f for f in files if 'STATISTIKA' not in f.upper()]
    files_to_use = non_stat_files if non_stat_files else files

    file_path = os.path.join(month_path, files_to_use[0])
    flights = []

    try:
        print(f"📊 {month_folder}: ", end="", flush=True)

        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)

        # Extract year and month from folder name
        month_number = month_folder.split('.')[0].strip().zfill(2)
        year = YEAR

        # Iterate through all sheets (days)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            # Detect format for THIS sheet (each sheet might have different structure!)
            sheet_rows = list(ws.iter_rows(values_only=True))
            if len(sheet_rows) == 0:
                continue

            header_row = sheet_rows[0]
            fmt = detect_format(header_row)

            # Extract day from sheet name (2024/2025 format) or use None (2023 format - date in first column)
            day_str = re.match(r'^(\d+)', sheet_name.strip())
            sheet_has_day = day_str is not None
            day_from_sheet = day_str.group(1).zfill(2) if sheet_has_day else None

            for i, row in enumerate(sheet_rows, 1):
                # Skip header
                if i == 1:
                    continue

                # Skip empty rows
                if not row[0]:
                    continue

                try:
                    # Determine proper_date_str
                    if sheet_has_day:
                        # 2024/2025 format: day is in sheet name
                        proper_date_str = f"{year}-{month_number}-{day_from_sheet}"
                    else:
                        # 2023 format: date is in first column (datetime object)
                        date_obj = row[0]
                        if isinstance(date_obj, datetime):
                            proper_date_str = date_obj.strftime('%Y-%m-%d')
                        else:
                            # Skip rows without valid date
                            continue
                    # Extract data using format-specific columns
                    airline_name = safe_get(row, fmt['airline_col'])

                    # Extract route (can be single column or list of [arrival_from, departure_to] columns)
                    route_col = fmt['route_col']
                    if isinstance(route_col, list):
                        # 2023 format: separate "Dolazak iz" and "Odlazak za" columns
                        arrival_from = safe_get(row, route_col[0])
                        departure_to = safe_get(row, route_col[1])
                        if arrival_from and departure_to:
                            route_str = f"{arrival_from}-{departure_to}"
                        elif arrival_from:
                            route_str = arrival_from
                        elif departure_to:
                            route_str = departure_to
                        else:
                            route_str = None
                    else:
                        # Other formats: single "Ruta" column
                        route_str = safe_get(row, route_col)

                    aircraft_model = safe_get(row, fmt['aircraft_col'])
                    available_seats = safe_get(row, fmt.get('available_seats_col'))
                    registration = safe_get(row, fmt['registration_col'])
                    operation_type_str = safe_get(row, fmt['operation_type_col'])
                    mtow = safe_get(row, fmt['mtow_col'])
                    arrival_flight_number = safe_get(row, fmt['arrival_flight_number_col'])
                    departure_flight_number = safe_get(row, fmt['departure_flight_number_col'])

                    # Times
                    scheduled_arrival_time = safe_get(row, fmt.get('scheduled_arrival_time_col'))
                    actual_arrival_time = safe_get(row, fmt.get('actual_arrival_time_col'))
                    scheduled_departure_time = safe_get(row, fmt.get('scheduled_departure_time_col'))
                    actual_departure_time = safe_get(row, fmt.get('actual_departure_time_col'))

                    # Passenger data
                    arrival_pax_val = safe_get(row, fmt['arrival_pax_col'])
                    departure_pax_val = safe_get(row, fmt['departure_pax_col'])
                    arrival_inf_val = safe_get(row, fmt['arrival_inf_col'])
                    departure_inf_val = safe_get(row, fmt['departure_inf_col'])

                    # Baggage, cargo, mail
                    arrival_baggage = safe_get(row, fmt['arrival_baggage_col'])
                    departure_baggage = safe_get(row, fmt['departure_baggage_col'])
                    arrival_cargo = safe_get(row, fmt['arrival_cargo_col'])
                    departure_cargo = safe_get(row, fmt['departure_cargo_col'])
                    arrival_mail = safe_get(row, fmt['arrival_mail_col'])
                    departure_mail = safe_get(row, fmt['departure_mail_col'])

                    # Parse date
                    try:
                        flight_date = datetime.strptime(proper_date_str, '%Y-%m-%d').isoformat()
                    except:
                        continue

                    # Parse route
                    route = parse_route(str(route_str))
                    if not route:
                        continue

                    # Handle passenger data based on format
                    if fmt['use_parse_passengers']:
                        # Format 1 & 2: Passengers in "165+6 INF" format
                        arrival_pax = parse_passengers(arrival_pax_val)
                        departure_pax = parse_passengers(departure_pax_val)
                        arrival_adults = arrival_pax['adults'] if arrival_pax else None
                        arrival_infants = arrival_pax['infants'] if arrival_pax else None
                        departure_adults = departure_pax['adults'] if departure_pax else None
                        departure_infants = departure_pax['infants'] if departure_pax else None
                    else:
                        # Format 3 & 4: Adults and infants in separate columns
                        arrival_adults = int(arrival_pax_val) if arrival_pax_val and str(arrival_pax_val).strip() not in ['-', 'N/A', ''] else None
                        departure_adults = int(departure_pax_val) if departure_pax_val and str(departure_pax_val).strip() not in ['-', 'N/A', ''] else None
                        arrival_infants = int(arrival_inf_val) if arrival_inf_val and str(arrival_inf_val).strip() not in ['-', 'N/A', ''] else None
                        departure_infants = int(departure_inf_val) if departure_inf_val and str(departure_inf_val).strip() not in ['-', 'N/A', ''] else None

                    # Create flight object
                    flight = {
                        'date': flight_date,
                        'airline': str(airline_name).strip().upper(),
                        'route': str(route_str),
                        'departureAirport': route['departure'],
                        'arrivalAirport': route['arrival'],
                        'aircraftModel': str(aircraft_model).strip(),
                        'registration': str(registration or ''),
                        'operationType': str(operation_type_str).strip().upper() if operation_type_str else 'N/A',

                        # Additional data
                        'availableSeats': int(available_seats) if available_seats and str(available_seats).strip() not in ['-', 'N/A', ''] else None,
                        'mtow': int(mtow) if mtow and str(mtow).strip() and str(mtow).strip() != '-' else None,
                        'arrivalFlightNumber': str(arrival_flight_number) if arrival_flight_number else None,
                        'departureFlightNumber': str(departure_flight_number) if departure_flight_number else None,

                        # Times
                        'scheduledArrivalTime': str(scheduled_arrival_time) if scheduled_arrival_time and str(scheduled_arrival_time).strip() not in ['-', 'N/A', ''] else None,
                        'actualArrivalTime': str(actual_arrival_time) if actual_arrival_time and str(actual_arrival_time).strip() not in ['-', 'N/A', ''] else None,
                        'scheduledDepartureTime': str(scheduled_departure_time) if scheduled_departure_time and str(scheduled_departure_time).strip() not in ['-', 'N/A', ''] else None,
                        'actualDepartureTime': str(actual_departure_time) if actual_departure_time and str(actual_departure_time).strip() not in ['-', 'N/A', ''] else None,

                        # Passengers
                        'arrivalPassengers': arrival_adults,
                        'arrivalInfants': arrival_infants,
                        'departurePassengers': departure_adults,
                        'departureInfants': departure_infants,

                        # Baggage, cargo, mail
                        'arrivalBaggage': int(arrival_baggage) if arrival_baggage and str(arrival_baggage).strip() != '-' else None,
                        'departureBaggage': int(departure_baggage) if departure_baggage and str(departure_baggage).strip() != '-' else None,
                        'arrivalCargo': int(arrival_cargo) if arrival_cargo and str(arrival_cargo).strip() != '-' else None,
                        'departureCargo': int(departure_cargo) if departure_cargo and str(departure_cargo).strip() != '-' else None,
                        'arrivalMail': int(arrival_mail) if arrival_mail and str(arrival_mail).strip() != '-' else None,
                        'departureMail': int(departure_mail) if departure_mail and str(departure_mail).strip() != '-' else None,

                        'sourceFile': files[0],
                        'sheet': sheet_name,
                    }

                    flights.append(flight)

                except Exception as e:
                    # Skip problematic rows silently
                    pass

        print(f"✅ {len(flights)} flights")
        wb.close()
        return flights

    except Exception as e:
        print(f"❌ Error: {str(e)}")
        return []

def main():
    """Main function"""
    # First argument is year (already handled in YEAR variable)
    # Second argument (if present) is month
    args = sys.argv[2:]  # Skip year argument
    month_arg = None

    for arg in args:
        if not arg.startswith('--'):
            month_arg = arg
            break

    print(f'🚀 Extracting {YEAR} Flight Data to JSON (v2 - Multi-Format)...\\n')

    all_flights = []

    if month_arg:
        # Extract specific month
        print(f"📅 Extracting single month: {month_arg}\\n")
        flights = extract_month(month_arg)
        all_flights.extend(flights)
    else:
        # Extract all months
        month_folders = sorted([
            f for f in os.listdir(STATS_DIR)
            if os.path.isdir(os.path.join(STATS_DIR, f))
        ])

        print(f"📅 Found {len(month_folders)} months\\n")

        for month_folder in month_folders:
            flights = extract_month(month_folder)
            all_flights.extend(flights)

    # Save to JSON
    output_dir = 'output'
    os.makedirs(output_dir, exist_ok=True)
    output_file = os.path.join(output_dir, f'{YEAR}-flights-data.json')

    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump({
            'flights': all_flights,
            'totalCount': len(all_flights),
            'extractedAt': datetime.now().isoformat(),
        }, f, indent=2, ensure_ascii=False)

    print('\\n' + '='*80)
    print('✅ EXTRACTION COMPLETED')
    print('='*80)
    print(f'\\n📋 Summary:')
    print(f'   - Total flights: {len(all_flights)}')
    print(f'   - Output file: {output_file}\\n')

if __name__ == '__main__':
    main()
