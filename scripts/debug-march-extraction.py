#!/usr/bin/env python3

import openpyxl
import os
import re

STATS_DIR = "/Users/emir_mw/stats/STATS/2025/Dnevni izvještaji"
month_folder = "03. MART"
month_path = os.path.join(STATS_DIR, month_folder)

files = [
    f for f in os.listdir(month_path)
    if 'Dnevni izvještaj o saobraćaju' in f and f.endswith('.xlsx')
]

file_path = os.path.join(month_path, files[0])
wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)

print(f"📊 Debugging March Extraction\n")
print(f"Total sheets: {len(wb.sheetnames)}")
print(f"All sheets: {wb.sheetnames}\n")

# Count how many sheets match the day pattern
valid_sheets = []
for sheet_name in wb.sheetnames:
    day_match = re.match(r'^(\d+)', sheet_name.strip())
    if day_match:
        valid_sheets.append((sheet_name, day_match.group(1)))
        print(f"  ✅ '{sheet_name}' -> day {day_match.group(1)}")
    else:
        print(f"  ❌ '{sheet_name}' -> NO MATCH (skipped)")

print(f"\n Valid sheets: {len(valid_sheets)}")

# Check first valid sheet
if valid_sheets:
    sheet_name = valid_sheets[0][0]
    ws = wb[sheet_name]

    print(f"\n Checking sheet '{sheet_name}':")
    row_count = 0
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if i == 1:
            continue  # Skip header

        if row[0]:
            row_count += 1
            if row_count <= 3:
                route = row[3] if len(row) > 3 else None
                print(f"  Row {i}: Route = '{route}'")

    print(f"\n  Total data rows in '{sheet_name}': {row_count}")

wb.close()
