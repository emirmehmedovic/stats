#!/usr/bin/env python3

import openpyxl
import os

STATS_DIR = "/Users/emir_mw/stats/STATS/2025/Dnevni izvještaji"
month_folder = "04. APRIL"
month_path = os.path.join(STATS_DIR, month_folder)

files = [
    f for f in os.listdir(month_path)
    if 'Dnevni izvještaj o saobraćaju' in f and f.endswith('.xlsx')
]

file_path = os.path.join(month_path, files[0])
wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)

print(f"📊 Analyzing April 2025 Excel file\n")
print(f"Total sheets: {len(wb.sheetnames)}")
print(f"Sheet names: {wb.sheetnames}\n")

# Check first sheet with data
for sheet_name in wb.sheetnames[:5]:
    print(f"\n{'='*80}")
    print(f"Sheet: {sheet_name}")
    print('='*80)

    ws = wb[sheet_name]

    row_count = 0
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if i == 1:
            print(f"\nHeader row:")
            for j, cell in enumerate(row):
                if cell:
                    print(f"  [{j}] {cell}")
            continue

        # Show first 3 data rows
        if row[0] and row_count < 3:
            row_count += 1
            print(f"\nRow {i}:")
            print(f"  [0] Date: {row[0]}")
            print(f"  [1] Airline: {row[1]}")
            print(f"  [2] Route: {row[2]}")
            print(f"  [3] Aircraft: {row[3]}")
            print(f"  [4] Registration: {row[4]}")
            print(f"  [13] Arrival pax: {row[13] if len(row) > 13 else 'N/A'}")
            print(f"  [14] Departure pax: {row[14] if len(row) > 14 else 'N/A'}")

wb.close()
