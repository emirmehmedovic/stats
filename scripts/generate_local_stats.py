#!/usr/bin/env python3
"""
Lokalna statistika - Local Statistics Report Generator

Generiše mjesečni Excel izvještaj sa:
- Redovni promet (dinamične destinacije po aviokompanijama)
- Vanredni promet
- Domestic flight
- Ostala slijetanja
- Izvještaji za fakturisanje
- UKUPNO rezime
"""

import sys
import os
import calendar
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Optional
import psycopg2
from psycopg2.extras import RealDictCursor
import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.cell import MergedCell

SCRIPT_DIR = Path(__file__).parent
PROJECT_ROOT = SCRIPT_DIR.parent
OUTPUT_DIR = PROJECT_ROOT / "izvještaji" / "generated"

HOME_AIRPORT_IATA = "TZL"

MONTH_NAMES = {
    1: "Januar", 2: "Februar", 3: "Mart", 4: "April",
    5: "Maj", 6: "Juni", 7: "Juli", 8: "Avgust",
    9: "Septembar", 10: "Oktobar", 11: "Novembar", 12: "Decembar"
}

DAY_NAMES = {
    0: "ponedjeljak",
    1: "utorak",
    2: "srijeda",
    3: "četvrtak",
    4: "petak",
    5: "subota",
    6: "nedjelja",
}


def get_db_connection():
    database_url = os.getenv("DATABASE_URL")
    if not database_url:
        raise ValueError("DATABASE_URL environment variable not set")
    return psycopg2.connect(database_url, cursor_factory=RealDictCursor)


def get_operation_type_code(operation_type_code):
    if not operation_type_code:
        return ""
    code = operation_type_code.upper()
    if code == "SCHEDULED":
        return "R"
    if code == "CHARTER":
        return "H"
    return code[0] if code else ""


def parse_route_codes(route):
    if not route:
        return []
    cleaned = route.replace('–', '-').replace('—', '-')
    parts = [part.strip().upper() for part in cleaned.split('-') if part.strip()]
    return parts


def get_route_other_iata(route):
    codes = parse_route_codes(route)
    for code in codes:
        if code != HOME_AIRPORT_IATA:
            return code
    return None


def get_flight_data(year: int, month: int):
    conn = get_db_connection()
    cursor = conn.cursor()

    first_day = f"{year}-{month:02d}-01"
    last_day = calendar.monthrange(year, month)[1]
    last_date = f"{year}-{month:02d}-{last_day:02d}"

    query = """
        SELECT
            f.id,
            f.date,
            f.route,

            a.name as airline_name,
            a."iataCode" as airline_iata,
            a."icaoCode" as airline_icao,

            ot.code as operation_type_code,

            -- Arrival info
            f."arrivalPassengers",
            f."arrivalInfants",
            f."arrivalCargo",
            f."arrivalMail",
            f."arrivalStatus",
            f."arrivalFerryIn",

            -- Departure info
            f."departurePassengers",
            f."departureInfants",
            f."departureCargo",
            f."departureMail",
            f."departureStatus",
            f."departureFerryOut",

            arr_airport."iataCode" as arrival_airport_iata,
            arr_airport."country" as arrival_airport_country,
            arr_airport."name" as arrival_airport_name,

            dep_airport."iataCode" as departure_airport_iata,
            dep_airport."country" as departure_airport_country,
            dep_airport."name" as departure_airport_name
        FROM "Flight" f
        INNER JOIN "Airline" a ON f."airlineId" = a.id
        INNER JOIN "OperationType" ot ON f."operationTypeId" = ot.id
        LEFT JOIN "Airport" arr_airport ON f."arrivalAirportId" = arr_airport.id
        LEFT JOIN "Airport" dep_airport ON f."departureAirportId" = dep_airport.id
        WHERE f.date >= %s AND f.date <= %s
        ORDER BY f.date
    """

    cursor.execute(query, (first_day, last_date))
    flights = cursor.fetchall()
    cursor.close()
    conn.close()
    return flights


def get_airline_display_name(airline_name: str, iata_code: Optional[str]):
    if iata_code and iata_code not in airline_name:
        return f"{airline_name} ({iata_code})"
    return airline_name


def get_airport_display_name(airport_name: Optional[str], iata_code: Optional[str]):
    if not iata_code:
        return ""
    if airport_name:
        return f"{airport_name} ({iata_code})"
    return iata_code


def is_domestic_flight(flight):
    arr_country = flight.get('arrival_airport_country')
    dep_country = flight.get('departure_airport_country')
    if arr_country and dep_country:
        return arr_country == 'Bosnia and Herzegovina' and dep_country == 'Bosnia and Herzegovina'
    return False


def count_movement(flight):
    movements = 0
    if flight.get('arrivalPassengers') is not None or flight.get('arrivalStatus') == 'OPERATED':
        movements += 1
    if flight.get('departurePassengers') is not None or flight.get('departureStatus') == 'OPERATED':
        movements += 1
    return movements


def build_redovni_data(flights):
    data = defaultdict(lambda: defaultdict(lambda: defaultdict(lambda: {
        'dep_pax': 0,
        'dep_inf': 0,
        'arr_pax': 0,
        'arr_inf': 0,
    })))
    destinations = defaultdict(set)
    airline_meta = {}
    dest_meta = {}
    summary = defaultdict(lambda: {
        'dep_flights': 0,
        'dep_pax': 0,
        'dep_inf': 0,
        'arr_pax': 0,
        'arr_inf': 0,
    })
    route_summary = defaultdict(lambda: defaultdict(lambda: {
        'dep_flights': 0,
        'dep_pax': 0,
        'dep_inf': 0,
        'arr_flights': 0,
        'arr_pax': 0,
        'arr_inf': 0,
    }))

    for flight in flights:
        operation_code = (flight.get('operation_type_code') or '').upper()
        if operation_code != 'SCHEDULED':
            continue

        if flight.get('arrivalFerryIn') or flight.get('departureFerryOut'):
            continue

        if is_domestic_flight(flight):
            continue

        airline_name = flight.get('airline_name') or ''
        airline_iata = flight.get('airline_iata')
        airline_key = (airline_name, airline_iata)
        airline_meta[airline_key] = get_airline_display_name(airline_name, airline_iata)

        date_key = flight['date'].date()

        dep_iata = flight.get('departure_airport_iata')
        arr_iata = flight.get('arrival_airport_iata')
        route_other = get_route_other_iata(flight.get('route'))
        route_codes = parse_route_codes(flight.get('route'))
        has_tzl_in_route = HOME_AIRPORT_IATA in route_codes

        if not dep_iata and has_tzl_in_route:
            dep_iata = HOME_AIRPORT_IATA if route_codes and route_codes[0] == HOME_AIRPORT_IATA else dep_iata
        if not arr_iata and has_tzl_in_route:
            arr_iata = HOME_AIRPORT_IATA if route_codes and route_codes[-1] == HOME_AIRPORT_IATA else arr_iata

        if dep_iata == HOME_AIRPORT_IATA and arr_iata:
            dest = arr_iata
            destinations[airline_key].add(dest)
            dest_meta[dest] = get_airport_display_name(flight.get('arrival_airport_name'), dest)

            data[date_key][airline_key][dest]['dep_pax'] += flight.get('departurePassengers') or 0
            data[date_key][airline_key][dest]['dep_inf'] += flight.get('departureInfants') or 0

            summary[airline_key]['dep_flights'] += 1
            summary[airline_key]['dep_pax'] += flight.get('departurePassengers') or 0
            summary[airline_key]['dep_inf'] += flight.get('departureInfants') or 0

            route_summary[airline_key][dest]['dep_flights'] += 1
            route_summary[airline_key][dest]['dep_pax'] += flight.get('departurePassengers') or 0
            route_summary[airline_key][dest]['dep_inf'] += flight.get('departureInfants') or 0

        if arr_iata == HOME_AIRPORT_IATA and dep_iata:
            dest = dep_iata
            destinations[airline_key].add(dest)
            dest_meta[dest] = get_airport_display_name(flight.get('departure_airport_name'), dest)

            data[date_key][airline_key][dest]['arr_pax'] += flight.get('arrivalPassengers') or 0
            data[date_key][airline_key][dest]['arr_inf'] += flight.get('arrivalInfants') or 0

            summary[airline_key]['arr_pax'] += flight.get('arrivalPassengers') or 0
            summary[airline_key]['arr_inf'] += flight.get('arrivalInfants') or 0

            route_summary[airline_key][dest]['arr_flights'] += 1
            route_summary[airline_key][dest]['arr_pax'] += flight.get('arrivalPassengers') or 0
            route_summary[airline_key][dest]['arr_inf'] += flight.get('arrivalInfants') or 0

        if has_tzl_in_route and route_other and dep_iata is None and arr_iata is None:
            dest = route_other
            destinations[airline_key].add(dest)
            dest_meta[dest] = dest_meta.get(dest, get_airport_display_name(None, dest))

            if flight.get('departurePassengers') or flight.get('departureInfants'):
                data[date_key][airline_key][dest]['dep_pax'] += flight.get('departurePassengers') or 0
                data[date_key][airline_key][dest]['dep_inf'] += flight.get('departureInfants') or 0

                summary[airline_key]['dep_flights'] += 1
                summary[airline_key]['dep_pax'] += flight.get('departurePassengers') or 0
                summary[airline_key]['dep_inf'] += flight.get('departureInfants') or 0

                route_summary[airline_key][dest]['dep_flights'] += 1
                route_summary[airline_key][dest]['dep_pax'] += flight.get('departurePassengers') or 0
                route_summary[airline_key][dest]['dep_inf'] += flight.get('departureInfants') or 0

            if flight.get('arrivalPassengers') or flight.get('arrivalInfants'):
                data[date_key][airline_key][dest]['arr_pax'] += flight.get('arrivalPassengers') or 0
                data[date_key][airline_key][dest]['arr_inf'] += flight.get('arrivalInfants') or 0

                summary[airline_key]['arr_pax'] += flight.get('arrivalPassengers') or 0
                summary[airline_key]['arr_inf'] += flight.get('arrivalInfants') or 0

                route_summary[airline_key][dest]['arr_flights'] += 1
                route_summary[airline_key][dest]['arr_pax'] += flight.get('arrivalPassengers') or 0
                route_summary[airline_key][dest]['arr_inf'] += flight.get('arrivalInfants') or 0

    return data, destinations, airline_meta, dest_meta, summary, route_summary


def sort_airlines(airlines):
    def key(item):
        name = item[0][0].upper()
        if 'WIZZ' in name:
            return (0, name)
        if 'PEGASUS' in name:
            return (1, name)
        if 'RYANAIR' in name:
            return (2, name)
        if 'AJET' in name:
            return (3, name)
        return (4, name)
    return sorted(airlines, key=key)


def aggregate_daily(flights, filter_fn):
    daily = defaultdict(lambda: {
        'operations': 0,
        'dep_pax': 0,
        'dep_inf': 0,
        'arr_pax': 0,
        'arr_inf': 0,
        'cargo_loaded': 0.0,
        'cargo_unloaded': 0.0,
        'routes': set(),
        'companies': set(),
        'operation_types': set(),
        'transit': 0,
    })

    for flight in flights:
        if not filter_fn(flight):
            continue

        day_key = flight['date'].date()
        daily[day_key]['operations'] += count_movement(flight)
        daily[day_key]['dep_pax'] += flight.get('departurePassengers') or 0
        daily[day_key]['dep_inf'] += flight.get('departureInfants') or 0
        daily[day_key]['arr_pax'] += flight.get('arrivalPassengers') or 0
        daily[day_key]['arr_inf'] += flight.get('arrivalInfants') or 0

        daily[day_key]['cargo_loaded'] += (flight.get('departureCargo') or 0) / 1000.0
        daily[day_key]['cargo_loaded'] += (flight.get('departureMail') or 0) / 1000.0
        daily[day_key]['cargo_unloaded'] += (flight.get('arrivalCargo') or 0) / 1000.0
        daily[day_key]['cargo_unloaded'] += (flight.get('arrivalMail') or 0) / 1000.0

        if flight.get('route'):
            daily[day_key]['routes'].add(flight.get('route'))
        else:
            dep = flight.get('departure_airport_iata')
            arr = flight.get('arrival_airport_iata')
            if dep and arr:
                daily[day_key]['routes'].add(f"{dep}-{arr}")

        airline_name = flight.get('airline_name')
        if airline_name:
            daily[day_key]['companies'].add(airline_name)

        daily[day_key]['operation_types'].add(get_operation_type_code(flight.get('operation_type_code')))

    return daily


TEMPLATE_PATH = PROJECT_ROOT / "izvještaji" / "10. LOKALNA STATISTIKA - Oktobar 2025.xlsx"


def clear_range(ws, start_row, end_row, start_col, end_col):
    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            cell = ws.cell(row=row, column=col)
            if isinstance(cell, MergedCell):
                continue
            cell.value = None


def normalize_name(value):
    return (value or "").upper()


def match_airline_key(block_name, airline_keys):
    name = normalize_name(block_name)
    def find_with(*keywords):
        for key in airline_keys:
            key_name = normalize_name(key[0])
            if all(k in key_name for k in keywords):
                return key
        return None

    if "WIZZ" in name and ("HUNGARY" in name or "W6" in name):
        return find_with("WIZZ", "HUNGARY") or find_with("WIZZ", "W6") or find_with("WIZZ")
    if "WIZZ" in name and ("MALTA" in name or "W4" in name):
        return find_with("WIZZ", "MALTA") or find_with("WIZZ", "W4") or find_with("WIZZ")
    if "PEGASUS" in name:
        return find_with("PEGASUS")
    if "RYANAIR" in name:
        return find_with("RYANAIR")
    if "AJET" in name:
        return find_with("AJET")
    if "WIZZ" in name:
        return find_with("WIZZ")
    return None


def get_template_airline_blocks(ws):
    airline_blocks = []
    dest_ranges = []
    for merged in ws.merged_cells.ranges:
        if merged.min_row == 2 and merged.max_row == 2:
            airline_blocks.append({
                'name': ws.cell(merged.min_row, merged.min_col).value,
                'start_col': merged.min_col,
                'end_col': merged.max_col,
                'dest_cols': []
            })
        if merged.min_row == 4 and merged.max_row == 4:
            dest_ranges.append(merged)

    airline_blocks.sort(key=lambda b: b['start_col'])
    dest_ranges = sorted(dest_ranges, key=lambda r: r.min_col)

    for block in airline_blocks:
        for r in dest_ranges:
            if block['start_col'] <= r.min_col <= block['end_col']:
                block['dest_cols'].append(r.min_col)

    return airline_blocks


def set_value_if_nonzero(ws, row, col, value):
    ws.cell(row=row, column=col).value = value if value else None


def populate_redovni_sheet(ws, year, month, redovni_data, destinations, airline_keys):
    last_day = calendar.monthrange(year, month)[1]
    start_row = 6
    last_template_day_row = 36

    airline_blocks = get_template_airline_blocks(ws)
    column_map = {}

    for block in airline_blocks:
        airline_key = match_airline_key(block['name'], airline_keys)
        dest_list = sorted(destinations.get(airline_key, [])) if airline_key else []
        dest_slots = block['dest_cols']
        if len(dest_list) > len(dest_slots):
            print(f"UPOZORENJE: Previše destinacija za {block['name']}. Prikazujem prvih {len(dest_slots)}.")
            dest_list = dest_list[:len(dest_slots)]

        for idx, col in enumerate(dest_slots):
            dest_value = dest_list[idx] if idx < len(dest_list) else None
            ws.cell(row=4, column=col).value = dest_value
            column_map[(airline_key, dest_value)] = {
                'dep_pax': col,
                'dep_inf': col + 1,
                'arr_pax': col + 2,
                'arr_inf': col + 3,
            }

            clear_range(ws, start_row, last_template_day_row, col, col + 3)
            if dest_value is None:
                ws.cell(row=4, column=col).value = None

    for day in range(1, last_day + 1):
        row_idx = start_row + day - 1
        date_value = datetime(year, month, day)
        ws.cell(row=row_idx, column=2).value = date_value
        ws.cell(row=row_idx, column=3).value = DAY_NAMES[date_value.weekday()]

        day_data = redovni_data.get(date_value.date(), {})
        for airline_key, dests in day_data.items():
            for dest, values in dests.items():
                cols = column_map.get((airline_key, dest))
                if not cols:
                    continue
                set_value_if_nonzero(ws, row_idx, cols['dep_pax'], values['dep_pax'])
                set_value_if_nonzero(ws, row_idx, cols['dep_inf'], values['dep_inf'])
                set_value_if_nonzero(ws, row_idx, cols['arr_pax'], values['arr_pax'])
                set_value_if_nonzero(ws, row_idx, cols['arr_inf'], values['arr_inf'])

    for row_idx in range(start_row + last_day, last_template_day_row + 1):
        ws.cell(row=row_idx, column=2).value = None
        ws.cell(row=row_idx, column=3).value = None

    ws['V43'] = None


def populate_daily_sheet(ws, year, month, daily_data, layout):
    last_day = calendar.monthrange(year, month)[1]
    start_row = layout['start_row']
    last_template_day_row = layout['last_row']

    clear_range(ws, start_row, last_template_day_row, layout['clear_start_col'], layout['clear_end_col'])

    for day in range(1, last_day + 1):
        row_idx = start_row + day - 1
        date_value = datetime(year, month, day)
        ws.cell(row=row_idx, column=layout['date_col']).value = date_value

        values = daily_data.get(date_value.date())
        if not values:
            continue

        ws.cell(row=row_idx, column=layout['op_col']).value = values['operations'] or None
        ws.cell(row=row_idx, column=layout['dep_pax_col']).value = values['dep_pax'] or None
        if layout.get('dep_inf_col'):
            ws.cell(row=row_idx, column=layout['dep_inf_col']).value = values['dep_inf'] or None
        ws.cell(row=row_idx, column=layout['arr_pax_col']).value = values['arr_pax'] or None
        if layout.get('arr_inf_col'):
            ws.cell(row=row_idx, column=layout['arr_inf_col']).value = values['arr_inf'] or None
        ws.cell(row=row_idx, column=layout['cargo_load_col']).value = round(values['cargo_loaded'], 2) or None
        ws.cell(row=row_idx, column=layout['cargo_unload_col']).value = round(values['cargo_unloaded'], 2) or None

        route_value = ", ".join(sorted(v for v in values['routes'] if v))
        company_value = ", ".join(sorted(values['companies']))
        op_value = ", ".join(sorted(v for v in values['operation_types'] if v))

        if layout.get('route_col'):
            ws.cell(row=row_idx, column=layout['route_col']).value = route_value or None
        if layout.get('company_col'):
            ws.cell(row=row_idx, column=layout['company_col']).value = company_value or None
        if layout.get('type_col'):
            ws.cell(row=row_idx, column=layout['type_col']).value = op_value or None
        if layout.get('transit_col'):
            ws.cell(row=row_idx, column=layout['transit_col']).value = values['transit'] or None

    for row_idx in range(start_row + last_day, last_template_day_row + 1):
        ws.cell(row=row_idx, column=layout['date_col']).value = None


def populate_wizz_sheet(ws, year, month, route_summary, dest_meta, airline_keys):
    ws['B9'] = f"Izvještaj o broju odlazećih putnika WIZZAIR                          za mjesec {MONTH_NAMES[month].upper()} {year}. godine"

    clear_range(ws, 16, 28, 2, 8)

    wizz_hungary = next((k for k in airline_keys if "WIZZ" in normalize_name(k[0]) and ("HUNGARY" in normalize_name(k[0]) or "W6" in normalize_name(k[0]))), None)
    wizz_malta = next((k for k in airline_keys if "WIZZ" in normalize_name(k[0]) and ("MALTA" in normalize_name(k[0]) or "W4" in normalize_name(k[0]))), None)

    ws.cell(row=16, column=2).value = wizz_hungary[0] if wizz_hungary else "WIZZ AIR HUNGARY LTD"
    ws.cell(row=22, column=2).value = wizz_malta[0] if wizz_malta else "WIZZ AIR MALTA"

    def fill_section(start_row, airline_key):
        if not airline_key:
            return 0, 0
        destinations = sorted(route_summary[airline_key].keys())
        total_flights = 0
        total_pax = 0
        for idx, dest in enumerate(destinations[:3], start=1):
            stats = route_summary[airline_key][dest]
            flights = stats['dep_flights']
            pax = stats['dep_pax']
            ws.cell(row=start_row + idx - 1, column=2).value = f"{idx}."
            ws.cell(row=start_row + idx - 1, column=3).value = dest_meta.get(dest, dest)
            ws.cell(row=start_row + idx - 1, column=6).value = flights
            ws.cell(row=start_row + idx - 1, column=8).value = pax
            total_flights += flights
            total_pax += pax
        return total_flights, total_pax

    hungary_totals = fill_section(17, wizz_hungary)
    ws.cell(row=20, column=6).value = hungary_totals[0]
    ws.cell(row=20, column=8).value = hungary_totals[1]

    malta_totals = fill_section(23, wizz_malta)
    ws.cell(row=26, column=6).value = malta_totals[0]
    ws.cell(row=26, column=8).value = malta_totals[1]

    ws.cell(row=28, column=3).value = "UKUPNO:"
    ws.cell(row=28, column=6).value = hungary_totals[0] + malta_totals[0]
    ws.cell(row=28, column=8).value = hungary_totals[1] + malta_totals[1]


def populate_simple_invoice(ws, year, month, route_summary, dest_meta, airline_keyword):
    ws['B14'] = f"Izvještaj o broju odlazećih putnika za mjesec {MONTH_NAMES[month].upper()} {year}. godine"

    airline_key = next((k for k in route_summary.keys() if airline_keyword in normalize_name(k[0])), None)
    if airline_key:
        ws['C9'] = airline_key[0]

    clear_range(ws, 23, 25, 2, 8)

    if not airline_key:
        return

    destinations = sorted(route_summary[airline_key].keys())
    total_flights = 0
    total_pax = 0
    for idx, dest in enumerate(destinations[:1], start=1):
        stats = route_summary[airline_key][dest]
        flights = stats['dep_flights']
        pax = stats['dep_pax']
        ws.cell(row=23, column=2).value = f"{idx}."
        ws.cell(row=23, column=3).value = dest_meta.get(dest, dest)
        ws.cell(row=23, column=6).value = flights
        ws.cell(row=23, column=8).value = pax
        total_flights += flights
        total_pax += pax

    ws.cell(row=25, column=3).value = "UKUPNO:"
    ws.cell(row=25, column=6).value = total_flights
    ws.cell(row=25, column=8).value = total_pax


def populate_ryanair_invoice(ws, year, month, route_summary, dest_meta, airline_keys):
    ws['B14'] = f"Izvještaj o broju odlazećih putnika za mjesec {MONTH_NAMES[month].upper()} {year}. godine"
    ws['M14'] = f"Izvještaj o broju dolazećih putnika za mjesec {MONTH_NAMES[month].upper()} {year}. godine"

    clear_range(ws, 23, 28, 2, 19)

    airline_key = next((k for k in airline_keys if "RYANAIR" in normalize_name(k[0])), None)
    if not airline_key:
        return

    ws['C9'] = airline_key[0]

    destinations = sorted(route_summary[airline_key].keys())
    dep_rows = []
    arr_rows = []
    for dest in destinations:
        stats = route_summary[airline_key][dest]
        if stats['dep_flights'] or stats['dep_pax']:
            dep_rows.append((dest, stats['dep_flights'], stats['dep_pax']))
        if stats['arr_flights'] or stats['arr_pax']:
            arr_rows.append((dest, stats['arr_flights'], stats['arr_pax']))

    total_dep_flights = total_dep_pax = 0
    total_arr_flights = total_arr_pax = 0

    for idx, (dest, flights, pax) in enumerate(dep_rows[:3], start=1):
        ws.cell(row=22 + idx, column=2).value = f"{idx}."
        ws.cell(row=22 + idx, column=3).value = dest_meta.get(dest, dest)
        ws.cell(row=22 + idx, column=6).value = flights
        ws.cell(row=22 + idx, column=8).value = pax
        total_dep_flights += flights
        total_dep_pax += pax

    for idx, (dest, flights, pax) in enumerate(arr_rows[:3], start=1):
        ws.cell(row=22 + idx, column=13).value = f"{idx}."
        ws.cell(row=22 + idx, column=14).value = dest_meta.get(dest, dest)
        ws.cell(row=22 + idx, column=17).value = flights
        ws.cell(row=22 + idx, column=19).value = pax
        total_arr_flights += flights
        total_arr_pax += pax

    ws.cell(row=25, column=3).value = "UKUPNO:"
    ws.cell(row=25, column=6).value = total_dep_flights
    ws.cell(row=25, column=8).value = total_dep_pax
    ws.cell(row=28, column=14).value = "UKUPNO:"
    ws.cell(row=28, column=17).value = total_arr_flights
    ws.cell(row=28, column=19).value = total_arr_pax


def generate_local_stats(year: int, month: int, output_path: Path = None):
    print(f"Generišem lokalnu statistiku za {MONTH_NAMES[month]} {year}...")
    flights = get_flight_data(year, month)
    if not flights:
        print("UPOZORENJE: Nema letova za zadati period!")
        return None

    redovni_data, destinations, airline_meta, dest_meta, redovni_summary, route_summary = build_redovni_data(flights)

    if not TEMPLATE_PATH.exists():
        raise FileNotFoundError(f"Template nije pronađen: {TEMPLATE_PATH}")

    wb = openpyxl.load_workbook(TEMPLATE_PATH, data_only=False)

    populate_redovni_sheet(
        wb['Redovni promet'],
        year,
        month,
        redovni_data,
        destinations,
        list(redovni_summary.keys())
    )

    vanredni_daily = aggregate_daily(
        flights,
        lambda f: (
            (f.get('operation_type_code') or '').upper() != 'SCHEDULED'
            and not is_domestic_flight(f)
            and not f.get('arrivalFerryIn')
            and not f.get('departureFerryOut')
        )
    )
    domestic_daily = aggregate_daily(flights, is_domestic_flight)
    ostala_daily = aggregate_daily(
        flights,
        lambda f: f.get('arrivalFerryIn') or f.get('departureFerryOut')
    )

    populate_daily_sheet(
        wb['Vanredni promet'],
        year,
        month,
        vanredni_daily,
        {
            'start_row': 6,
            'last_row': 39,
            'date_col': 3,
            'op_col': 4,
            'dep_pax_col': 5,
            'dep_inf_col': 6,
            'arr_pax_col': 7,
            'arr_inf_col': 8,
            'cargo_load_col': 9,
            'cargo_unload_col': 10,
            'route_col': 11,
            'company_col': 13,
            'type_col': 14,
            'transit_col': 15,
            'clear_start_col': 3,
            'clear_end_col': 15,
        }
    )

    populate_daily_sheet(
        wb['DOMESTIC FLIGHT'],
        year,
        month,
        domestic_daily,
        {
            'start_row': 6,
            'last_row': 36,
            'date_col': 2,
            'op_col': 3,
            'dep_pax_col': 4,
            'dep_inf_col': 5,
            'arr_pax_col': 6,
            'arr_inf_col': 7,
            'cargo_load_col': 8,
            'cargo_unload_col': 9,
            'route_col': 10,
            'company_col': 12,
            'type_col': 13,
            'clear_start_col': 2,
            'clear_end_col': 13,
        }
    )

    populate_daily_sheet(
        wb['Ostala slijetanja'],
        year,
        month,
        ostala_daily,
        {
            'start_row': 7,
            'last_row': 37,
            'date_col': 3,
            'op_col': 4,
            'dep_pax_col': 5,
            'arr_pax_col': 6,
            'cargo_load_col': 7,
            'cargo_unload_col': 8,
            'route_col': 9,
            'company_col': 11,
            'type_col': 12,
            'clear_start_col': 3,
            'clear_end_col': 12,
        }
    )

    airline_keys = list(redovni_summary.keys())
    populate_wizz_sheet(wb['Izvještaj za fakturisanje WZZ'], year, month, route_summary, dest_meta, airline_keys)
    populate_simple_invoice(wb['Izvještaj za fakturisanje PGT'], year, month, route_summary, dest_meta, "PEGASUS")
    populate_ryanair_invoice(wb['Izvještaj za fakturisanje Ryana'], year, month, route_summary, dest_meta, airline_keys)
    populate_simple_invoice(wb['Izvještaj za fakturisanje TKJ'], year, month, route_summary, dest_meta, "AJET")

    if output_path is None:
        OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
        output_path = OUTPUT_DIR / f"{month:02d}. LOKALNA STATISTIKA - {MONTH_NAMES[month]} {year}.xlsx"

    print(f"Čuvam izvještaj u: {output_path}")
    wb.save(output_path)
    print("✅ Lokalna statistika uspješno generisana!")
    return str(output_path)


if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("Usage: python generate_local_stats.py <year> <month>")
        print("Example: python generate_local_stats.py 2025 10")
        sys.exit(1)

    year = int(sys.argv[1])
    month = int(sys.argv[2])
    if month < 1 or month > 12:
        print("Mjesec mora biti između 1 i 12")
        sys.exit(1)

    try:
        output_file = generate_local_stats(year, month)
        print(f"\nIzvještaj sačuvan: {output_file}")
    except Exception as e:
        print(f"GREŠKA: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
