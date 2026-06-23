#!/usr/bin/env python3
"""
Load Factor Report Generator

Generiše detaljni izvještaj o load factoru sa:
- Sažetak
- Breakdown po destinacijama
- Mjesečni breakdown po destinacijama (pivot tabela)
- Dnevni trend
"""

import sys
import os
import json
from pathlib import Path
from datetime import datetime
from calendar import month_name
import psycopg2
from psycopg2.extras import RealDictCursor
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

SCRIPT_DIR = Path(__file__).parent
PROJECT_ROOT = SCRIPT_DIR.parent
OUTPUT_DIR = PROJECT_ROOT / "izvještaji" / "generated"


def get_db_connection():
    """Konekcija na PostgreSQL bazu"""
    DATABASE_URL = os.getenv("DATABASE_URL")
    if not DATABASE_URL:
        raise ValueError("DATABASE_URL environment variable not set")
    return psycopg2.connect(DATABASE_URL, cursor_factory=RealDictCursor)


def get_load_factor_data(date_from, date_to, airlines, routes, operation_type_id, direction):
    """
    Povuci load factor podatke sa filterima
    """
    conn = get_db_connection()
    cursor = conn.cursor()

    # Build WHERE conditions
    where_conditions = ["f.date >= %s", "f.date <= %s"]
    params = [date_from, date_to]

    if airlines:
        airline_placeholders = ','.join(['%s'] * len(airlines))
        where_conditions.append(f'a."icaoCode" IN ({airline_placeholders})')
        params.extend(airlines)

    if routes:
        route_placeholders = ','.join(['%s'] * len(routes))
        where_conditions.append(f'f.route IN ({route_placeholders})')
        params.extend(routes)

    if operation_type_id and operation_type_id != 'ALL':
        where_conditions.append('f."operationTypeId" = %s')
        params.append(operation_type_id)

    if direction == 'arrival':
        where_conditions.append('f."arrivalFlightNumber" IS NOT NULL')
    elif direction == 'departure':
        where_conditions.append('f."departureFlightNumber" IS NOT NULL')

    where_clause = " AND ".join(where_conditions)

    # Passenger and seats expressions based on direction
    if direction == 'arrival':
        passenger_expr = "CASE WHEN f.\"arrivalFerryIn\" THEN 0 ELSE COALESCE(f.\"arrivalPassengers\", 0) + COALESCE(f.\"arrivalInfants\", 0) END"
        seats_expr = "CASE WHEN f.\"arrivalFerryIn\" THEN 0 ELSE COALESCE(f.\"availableSeats\", at.seats, 0) END"
    elif direction == 'departure':
        passenger_expr = "CASE WHEN f.\"departureFerryOut\" THEN 0 ELSE COALESCE(f.\"departurePassengers\", 0) + COALESCE(f.\"departureInfants\", 0) END"
        seats_expr = "CASE WHEN f.\"departureFerryOut\" THEN 0 ELSE COALESCE(f.\"availableSeats\", at.seats, 0) END"
    else:  # all
        passenger_expr = """
            CASE WHEN f."arrivalFerryIn" THEN 0 ELSE COALESCE(f."arrivalPassengers", 0) + COALESCE(f."arrivalInfants", 0) END +
            CASE WHEN f."departureFerryOut" THEN 0 ELSE COALESCE(f."departurePassengers", 0) + COALESCE(f."departureInfants", 0) END
        """
        seats_expr = """
            CASE WHEN f."arrivalFerryIn" THEN 0 ELSE COALESCE(f."availableSeats", at.seats, 0) END +
            CASE WHEN f."departureFerryOut" THEN 0 ELSE COALESCE(f."availableSeats", at.seats, 0) END
        """

    # Summary query
    summary_query = f"""
        WITH flight_stats AS (
            SELECT
                ({passenger_expr})::int AS total_passengers,
                ({seats_expr})::int AS total_seats
            FROM "Flight" f
            LEFT JOIN "AircraftType" at ON f."aircraftTypeId" = at.id
            LEFT JOIN "Airline" a ON f."airlineId" = a.id
            WHERE {where_clause}
        )
        SELECT
            COUNT(*) FILTER (WHERE total_seats > 0)::int AS total_flights,
            SUM(total_passengers)::int AS total_passengers,
            SUM(total_seats)::int AS total_seats,
            AVG(CASE WHEN total_seats > 0 THEN (total_passengers::float / total_seats) * 100 END) AS avg_load_factor
        FROM flight_stats;
    """

    # By route query
    route_query = f"""
        WITH flight_stats AS (
            SELECT
                f.route,
                ({passenger_expr})::int AS total_passengers,
                ({seats_expr})::int AS total_seats
            FROM "Flight" f
            LEFT JOIN "AircraftType" at ON f."aircraftTypeId" = at.id
            LEFT JOIN "Airline" a ON f."airlineId" = a.id
            WHERE {where_clause}
        )
        SELECT
            route,
            COUNT(*) FILTER (WHERE total_seats > 0)::int AS flights,
            SUM(total_passengers)::int AS total_passengers,
            SUM(total_seats)::int AS total_seats,
            AVG(CASE WHEN total_seats > 0 THEN (total_passengers::float / total_seats) * 100 END) AS avg_load_factor
        FROM flight_stats
        GROUP BY route
        ORDER BY avg_load_factor DESC;
    """

    # Monthly by route query (for pivot table)
    monthly_route_query = f"""
        WITH flight_stats AS (
            SELECT
                f.route,
                date_trunc('month', f.date) AS month,
                ({passenger_expr})::int AS total_passengers,
                ({seats_expr})::int AS total_seats
            FROM "Flight" f
            LEFT JOIN "AircraftType" at ON f."aircraftTypeId" = at.id
            LEFT JOIN "Airline" a ON f."airlineId" = a.id
            WHERE {where_clause}
        )
        SELECT
            route,
            month,
            COUNT(*) FILTER (WHERE total_seats > 0)::int AS flights,
            SUM(total_passengers)::int AS total_passengers,
            SUM(total_seats)::int AS total_seats,
            AVG(CASE WHEN total_seats > 0 THEN (total_passengers::float / total_seats) * 100 END) AS avg_load_factor
        FROM flight_stats
        GROUP BY route, month
        ORDER BY route, month;
    """

    # Daily trend query
    daily_query = f"""
        WITH flight_stats AS (
            SELECT
                date_trunc('day', f.date) AS day,
                ({passenger_expr})::int AS total_passengers,
                ({seats_expr})::int AS total_seats
            FROM "Flight" f
            LEFT JOIN "AircraftType" at ON f."aircraftTypeId" = at.id
            LEFT JOIN "Airline" a ON f."airlineId" = a.id
            WHERE {where_clause}
        )
        SELECT
            day,
            COUNT(*) FILTER (WHERE total_seats > 0)::int AS flights,
            SUM(total_passengers)::int AS total_passengers,
            SUM(total_seats)::int AS total_seats,
            AVG(CASE WHEN total_seats > 0 THEN (total_passengers::float / total_seats) * 100 END) AS avg_load_factor
        FROM flight_stats
        GROUP BY day
        ORDER BY day ASC;
    """

    cursor.execute(summary_query, params)
    summary = cursor.fetchone()

    cursor.execute(route_query, params)
    by_route = cursor.fetchall()

    cursor.execute(monthly_route_query, params)
    monthly_by_route = cursor.fetchall()

    cursor.execute(daily_query, params)
    daily_trend = cursor.fetchall()

    cursor.close()
    conn.close()

    return {
        'summary': summary,
        'by_route': by_route,
        'monthly_by_route': monthly_by_route,
        'daily_trend': daily_trend,
    }


def create_excel_report(data, filters, date_from, date_to):
    """
    Kreira Excel izvještaj sa formatiranim podacima
    """
    wb = openpyxl.Workbook()

    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    subheader_font = Font(bold=True, color="FFFFFF", size=10)
    subheader_fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")

    data_alignment = Alignment(horizontal="center", vertical="center")

    thin_border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )

    # Sheet 1: Summary
    ws_summary = wb.active
    ws_summary.title = "Sažetak"

    summary = data['summary']
    avg_lf = round(summary['avg_load_factor']) if summary['avg_load_factor'] else 0

    summary_data = [
        ['LOAD FACTOR ANALIZA - Aerodrom Tuzla'],
        [],
        ['Period:', f"{date_from} do {date_to}"],
        ['Aviokompanije:', ', '.join(filters['airlines']) if filters['airlines'] else 'SVE'],
        ['Rute:', ', '.join(filters['routes']) if filters['routes'] else 'SVE'],
        ['Smjer:', filters['direction_label']],
        [],
        ['SAŽETAK'],
        ['Ukupno letova:', summary['total_flights'] or 0],
        ['Prosječna popunjenost:', f"{avg_lf}%"],
        ['Ukupno putnika:', summary['total_passengers'] or 0],
        ['Ukupno sjedišta:', summary['total_seats'] or 0],
    ]

    for row_idx, row_data in enumerate(summary_data, start=1):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws_summary.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            if row_idx == 1:
                cell.font = Font(bold=True, size=14, color="1E40AF")
            elif row_idx == 8:
                cell.font = Font(bold=True, size=12)
                cell.fill = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
            elif col_idx == 1:
                cell.font = Font(bold=True)

    ws_summary.column_dimensions['A'].width = 25
    ws_summary.column_dimensions['B'].width = 40

    # Sheet 2: By Route (Po destinacijama)
    ws_route = wb.create_sheet("Po destinacijama")

    headers = ['Destinacija', 'Letovi', 'Load Factor (%)', 'Putnici', 'Sjedišta']
    for col_idx, header in enumerate(headers, start=1):
        cell = ws_route.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    for row_idx, route_data in enumerate(data['by_route'], start=2):
        lf = round(route_data['avg_load_factor']) if route_data['avg_load_factor'] else 0
        row_values = [
            route_data['route'] or 'Unknown',
            route_data['flights'] or 0,
            lf,
            route_data['total_passengers'] or 0,
            route_data['total_seats'] or 0,
        ]
        for col_idx, value in enumerate(row_values, start=1):
            cell = ws_route.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            if col_idx > 1:
                cell.alignment = data_alignment
            # Color code load factor
            if col_idx == 3:
                if lf >= 85:
                    cell.font = Font(bold=True, color="10B981")
                elif lf >= 70:
                    cell.font = Font(bold=True, color="3B82F6")
                else:
                    cell.font = Font(bold=True, color="EF4444")

    ws_route.column_dimensions['A'].width = 20
    for col in range(2, 6):
        ws_route.column_dimensions[get_column_letter(col)].width = 15

    # Sheet 3: Monthly Breakdown by Route (Pivot Table)
    ws_monthly = wb.create_sheet("Mjesečni breakdown")

    # Build pivot data structure
    routes_set = set()
    months_set = set()
    pivot_data = {}

    for row in data['monthly_by_route']:
        route = row['route'] or 'Unknown'
        month = row['month']
        lf = round(row['avg_load_factor']) if row['avg_load_factor'] else 0

        routes_set.add(route)
        months_set.add(month)

        if route not in pivot_data:
            pivot_data[route] = {}
        pivot_data[route][month] = lf

    routes_sorted = sorted(routes_set)
    months_sorted = sorted(months_set)

    # Write headers
    ws_monthly.cell(row=1, column=1, value='Destinacija').font = header_font
    ws_monthly.cell(row=1, column=1).fill = header_fill
    ws_monthly.cell(row=1, column=1).alignment = header_alignment
    ws_monthly.cell(row=1, column=1).border = thin_border

    for col_idx, month in enumerate(months_sorted, start=2):
        month_str = month.strftime('%m/%Y')
        cell = ws_monthly.cell(row=1, column=col_idx, value=month_str)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Write data
    for row_idx, route in enumerate(routes_sorted, start=2):
        ws_monthly.cell(row=row_idx, column=1, value=route).border = thin_border

        for col_idx, month in enumerate(months_sorted, start=2):
            lf = pivot_data.get(route, {}).get(month, 0)
            cell = ws_monthly.cell(row=row_idx, column=col_idx, value=f"{lf}%" if lf > 0 else "-")
            cell.alignment = data_alignment
            cell.border = thin_border

            # Color code
            if lf >= 85:
                cell.font = Font(bold=True, color="10B981")
            elif lf >= 70:
                cell.font = Font(bold=True, color="3B82F6")
            elif lf > 0:
                cell.font = Font(bold=True, color="EF4444")

    ws_monthly.column_dimensions['A'].width = 20
    for col in range(2, len(months_sorted) + 2):
        ws_monthly.column_dimensions[get_column_letter(col)].width = 12

    # Sheet 4: Daily Trend
    ws_daily = wb.create_sheet("Dnevni trend")

    daily_headers = ['Datum', 'Letovi', 'Load Factor (%)', 'Putnici', 'Sjedišta']
    for col_idx, header in enumerate(daily_headers, start=1):
        cell = ws_daily.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    for row_idx, daily_data in enumerate(data['daily_trend'], start=2):
        lf = round(daily_data['avg_load_factor']) if daily_data['avg_load_factor'] else 0
        date_str = daily_data['day'].strftime('%d.%m.%Y')

        row_values = [
            date_str,
            daily_data['flights'] or 0,
            lf,
            daily_data['total_passengers'] or 0,
            daily_data['total_seats'] or 0,
        ]

        for col_idx, value in enumerate(row_values, start=1):
            cell = ws_daily.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            if col_idx > 1:
                cell.alignment = data_alignment
            if col_idx == 3 and lf >= 85:
                cell.font = Font(bold=True, color="10B981")

    ws_daily.column_dimensions['A'].width = 15
    for col in range(2, 6):
        ws_daily.column_dimensions[get_column_letter(col)].width = 15

    return wb


def main():
    """Main function"""
    if len(sys.argv) < 2:
        print(json.dumps({
            "success": False,
            "error": "Missing filters argument"
        }))
        sys.exit(1)

    try:
        # Parse filters from command line argument
        filters = json.loads(sys.argv[1])

        date_from = filters.get('dateFrom')
        date_to = filters.get('dateTo')
        airlines = filters.get('airlines', [])
        routes = filters.get('routes', [])
        operation_type_id = filters.get('operationTypeId')
        direction = filters.get('direction', 'all')

        if not date_from or not date_to:
            print(json.dumps({
                "success": False,
                "error": "dateFrom and dateTo are required"
            }))
            sys.exit(1)

        # Direction label
        direction_label = 'Dolazni' if direction == 'arrival' else 'Odlazni' if direction == 'departure' else 'Svi'
        filters['direction_label'] = direction_label

        # Fetch data
        data = get_load_factor_data(date_from, date_to, airlines, routes, operation_type_id, direction)

        # Generate Excel
        wb = create_excel_report(data, filters, date_from, date_to)

        # Save file
        OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"Load_Factor_{timestamp}.xlsx"
        output_path = OUTPUT_DIR / filename

        wb.iso_dates = True
        wb.save(output_path)
        wb.close()

        print(json.dumps({
            "success": True,
            "fileName": filename,
            "message": f"Load Factor izvještaj uspješno generisan"
        }))

    except Exception as e:
        print(json.dumps({
            "success": False,
            "error": str(e)
        }))
        sys.exit(1)


if __name__ == "__main__":
    main()
