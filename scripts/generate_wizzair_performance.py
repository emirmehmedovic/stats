#!/usr/bin/env python3
"""
Wizz Air Daily Performance Table Generator

Generiše mjesečni izvještaj performansi za Wizz Air sa jednim sheet-om po danu.
Times are exported as stored in flight records (no timezone conversion).
"""

import sys
import os
import re
import pytz
import calendar
from pathlib import Path
from datetime import datetime, timedelta
import psycopg2
from psycopg2.extras import RealDictCursor
import openpyxl
from openpyxl.utils.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

# Putanja
SCRIPT_DIR = Path(__file__).parent
PROJECT_ROOT = SCRIPT_DIR.parent
OUTPUT_DIR = PROJECT_ROOT / "izvještaji" / "generated"

# Nazivi mjeseci
MONTH_NAMES = {
    1: "Januar", 2: "Februar", 3: "Mart", 4: "April",
    5: "Maj", 6: "Juni", 7: "Juli", 8: "Avgust",
    9: "Septembar", 10: "Oktobar", 11: "Novembar", 12: "Decembar"
}

def get_db_connection():
    """Konekcija na PostgreSQL bazu"""
    DATABASE_URL = os.getenv("DATABASE_URL")
    if not DATABASE_URL:
        raise ValueError("DATABASE_URL environment variable not set")

    return psycopg2.connect(DATABASE_URL, cursor_factory=RealDictCursor)


def create_merged_cell(ws, row, start_col, end_col, value, font=None, fill=None, alignment=None, border=None):
    """
    Helper function to properly create merged cells with consistent formatting.
    Applies formatting to ALL cells in the range before merging to avoid Excel corruption.
    """
    # Convert column letters to numbers if needed
    if isinstance(start_col, str):
        start_col = openpyxl.utils.column_index_from_string(start_col)
    if isinstance(end_col, str):
        end_col = openpyxl.utils.column_index_from_string(end_col)
    
    for col in range(start_col, end_col + 1):
        cell = ws.cell(row=row, column=col)
        if col == start_col and value is not None:
            cell.value = value
        if font:
            cell.font = font
        if fill:
            cell.fill = fill
        if alignment:
            cell.alignment = alignment
        if border:
            cell.border = border
    
    if start_col != end_col:
        ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=end_col)

def sanitize_text(value):
    """Remove control characters that can corrupt Excel XML."""
    if value is None:
        return None
    if not isinstance(value, str):
        return value
    value = ILLEGAL_CHARACTERS_RE.sub('', value)
    return re.sub(r'[\x00-\x08\x0B\x0C\x0E-\x1F]', '', value)

SARAJEVO_TZ = pytz.timezone('Europe/Sarajevo')
UTC_TZ = pytz.utc

def to_local_datetime(dt):
    """Convert aware datetimes to Europe/Sarajevo; keep naive datetimes unchanged."""
    if dt is None:
        return None
    if isinstance(dt, datetime):
        if dt.tzinfo is None:
            return UTC_TZ.localize(dt).astimezone(SARAJEVO_TZ)
        return dt.astimezone(SARAJEVO_TZ)
    return dt


def format_time_local(dt):
    """Format datetime as HH:MM for display"""
    if dt is None:
        return ""
    if hasattr(dt, 'hour') and not hasattr(dt, 'year'):
        return dt.strftime('%H:%M')
    return dt.strftime('%H:%M')


def calculate_delay_minutes(scheduled, actual):
    """
    Calculate delay in minutes between scheduled and actual time.
    Both times should be datetime objects.

    Returns: minutes (int) or None if cannot calculate
    """
    if scheduled is None or actual is None:
        return None

    delay = actual - scheduled
    minutes = int(delay.total_seconds() / 60)

    # Only return positive delays (negative means early)
    return minutes if minutes > 0 else 0


def get_flight_data(year: int, month: int, day: int = None):
    """
    Povuči sve Wizz Air letove za zadati mjesec ili dan
    """
    conn = get_db_connection()
    cursor = conn.cursor()

    # Datum početka i kraja perioda
    if day:
        first_day = f"{year}-{month:02d}-{day:02d}"
        last_date = first_day
    else:
        first_day = f"{year}-{month:02d}-01"
        last_day = calendar.monthrange(year, month)[1]
        last_date = f"{year}-{month:02d}-{last_day:02d}"

    query = """
        SELECT
            f.id,
            f.date,
            f.route,
            f.registration,

            -- Airline
            a.name as airline_name,
            a."icaoCode" as airline_icao,
            a."iataCode" as airline_iata,

            -- Departure info
            f."departureFlightNumber",
            f."departureScheduledTime",
            f."departureActualTime",
            f."departureDoorClosingTime",
            f."departurePassengers",
            f."departureStatus",

            -- Arrival info
            f."arrivalFlightNumber",
            f."arrivalScheduledTime",
            f."arrivalActualTime",
            f."arrivalPassengers",
            f."arrivalStatus"

        FROM "Flight" f
        INNER JOIN "Airline" a ON f."airlineId" = a.id
        WHERE f.date >= %s AND f.date <= %s
          AND (a."iataCode" IN ('W6', 'W4') OR a."icaoCode" IN ('WZZ', 'WMT'))
        ORDER BY f.date, f."departureScheduledTime", f."arrivalScheduledTime"
    """

    cursor.execute(query, (first_day, last_date))
    flights = cursor.fetchall()

    cursor.close()
    conn.close()

    return flights


def get_flight_delays(flight_id):
    """
    Povuči delay kodove za zadati let iz FlightDelay tabele

    Returns:
        List of delays: [{'phase': 'DEP', 'code': 'G8', 'minutes': 30, 'comment': '...'}, ...]
    """
    conn = get_db_connection()
    cursor = conn.cursor()

    query = """
        SELECT
            fd.phase,
            fd.minutes,
            fd."isPrimary",
            fd.comment,
            fd."unofficialReason",
            dc.code as delay_code,
            dc.description as delay_description
        FROM "FlightDelay" fd
        INNER JOIN "DelayCode" dc ON fd."delayCodeId" = dc.id
        WHERE fd."flightId" = %s
        ORDER BY fd.phase, fd."isPrimary" DESC, fd.minutes DESC
    """

    cursor.execute(query, (flight_id,))
    delays = cursor.fetchall()

    cursor.close()
    conn.close()

    return delays


def parse_route(route_str):
    """Extract IATA code from route string"""
    if not route_str or '-' not in route_str:
        return None

    parts = [p.strip().upper() for p in route_str.strip().split('-') if p.strip()]
    if not parts:
        return None
    for part in parts:
        if part != 'TZL':
            return part
    return parts[0]


def generate_wizzair_performance(year: int, month: int, day: int = None, output_path: Path = None):
    """
    Glavni metod za generisanje Wizz Air Performance izvještaja
    """
    if day:
        print(f"Generišem Wizz Air Performance izvještaj za {day:02d}.{month:02d}.{year}...")
    else:
        print(f"Generišem Wizz Air Performance izvještaj za {MONTH_NAMES[month]} {year}...")

    # 1. Povući podatke iz baze
    print("Povlačim podatke iz baze...")
    flights = get_flight_data(year, month, day)
    print(f"Pronađeno {len(flights)} Wizz Air letova.")

    if len(flights) == 0:
        print("UPOZORENJE: Nema Wizz Air letova za zadati period!")
        return None

    # 2. Grupiši letove po danima
    print("Grupisanje letova po danima...")
    flights_by_day = {}
    for flight in flights:
        day = flight['date'].day
        if day not in flights_by_day:
            flights_by_day[day] = []
        flights_by_day[day].append(flight)

    print(f"Letovi pronađeni za {len(flights_by_day)} dana.")

    # 3. Kreiranje Excel workbook-a
    print("Kreiram Excel workbook...")
    wb = openpyxl.Workbook()

    # Ukloni default sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])

    # 4. Stilovi
    header_font = Font(name='Arial', size=10, bold=True)
    header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    title_font = Font(name='Arial', size=12, bold=True)
    regular_font = Font(name='Arial', size=9)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')

    # 5. Kreiranje sheet-a za svaki dan u mjesecu (ili jedan dan)
    num_days = calendar.monthrange(year, month)[1]
    days_to_generate = [day] if day else list(range(1, num_days + 1))

    for day in days_to_generate:
        # Kreiraj sheet za ovaj dan
        sheet_name = f"{day:02d}"
        ws = wb.create_sheet(title=sheet_name)

        # Header - Row 1: Title
        create_merged_cell(
            ws, 1, 'A', 'U',
            "WIZZ Air Daily Performance Table",
            font=title_font,
            alignment=center_align
        )

        # Header - Row 2: Column headers
        headers = [
            'Nr', 'Date', 'Flight Nr', 'Desk', 'A/C Reg', 'STA', 'ATA', 'STD', 'DCT', 'ATD', 'Total delay',
            # Delay 1
            'Min', 'DL code', 'Reason',
            # Delay 2
            'Min', 'DL code', 'Reason',
            # Delay 3
            'Min', 'DL code', 'Reason',
            'PAX'
        ]

        for col_idx, header_text in enumerate(headers, start=1):
            cell = ws.cell(row=2, column=col_idx)
            cell.value = header_text
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = center_align

        # Popuni podatke za ovaj dan
        day_flights = flights_by_day.get(day, [])
        row_idx = 3
        flight_nr = 1

        for flight in day_flights:
            # Parse route
            desk = sanitize_text(parse_route(flight['route']) or '')

            # Use stored times converted to local time (if timezone-aware)
            sta_local = to_local_datetime(flight.get('arrivalScheduledTime'))
            ata_local = to_local_datetime(flight.get('arrivalActualTime'))
            std_local = to_local_datetime(flight.get('departureScheduledTime'))
            dct_local = to_local_datetime(flight.get('departureDoorClosingTime'))
            atd_local = to_local_datetime(flight.get('departureActualTime'))

            # Calculate total delay based on door close time
            total_delay_minutes = calculate_delay_minutes(std_local, dct_local)

            # Get delays from database
            delays = get_flight_delays(flight['id'])

            # Filter only departure delays for this report
            dep_delays = [d for d in delays if d['phase'] == 'DEP']

            # Basic flight info
            ws.cell(row=row_idx, column=1).value = flight_nr  # Nr
            ws.cell(row=row_idx, column=2).value = flight['date'].strftime('%d/%m/%Y')  # Date
            ws.cell(row=row_idx, column=3).value = sanitize_text(
                flight.get('departureFlightNumber') or flight.get('arrivalFlightNumber') or ''
            )  # Flight Nr
            ws.cell(row=row_idx, column=4).value = desk  # Desk
            ws.cell(row=row_idx, column=5).value = sanitize_text(flight.get('registration') or '')  # A/C Reg
            ws.cell(row=row_idx, column=6).value = sanitize_text(format_time_local(sta_local))  # STA
            ws.cell(row=row_idx, column=7).value = sanitize_text(format_time_local(ata_local))  # ATA
            ws.cell(row=row_idx, column=8).value = sanitize_text(format_time_local(std_local))  # STD
            ws.cell(row=row_idx, column=9).value = sanitize_text(format_time_local(dct_local))  # DCT
            ws.cell(row=row_idx, column=10).value = sanitize_text(format_time_local(atd_local))  # ATD

            # Total delay
            if total_delay_minutes and total_delay_minutes > 0:
                # Format as H:MM
                hours = total_delay_minutes // 60
                mins = total_delay_minutes % 60
                ws.cell(row=row_idx, column=11).value = sanitize_text(f"{hours}:{mins:02d}")
            else:
                ws.cell(row=row_idx, column=11).value = sanitize_text("No Delay")

            # Delay codes (up to 3)
            for i, delay in enumerate(dep_delays[:3]):
                base_col = 12 + (i * 3)  # Delay 1 starts at col 12, Delay 2 at 15, Delay 3 at 18
                ws.cell(row=row_idx, column=base_col).value = delay['minutes']  # Min
                ws.cell(row=row_idx, column=base_col + 1).value = sanitize_text(delay.get('delay_code') or '')  # DL code
                ws.cell(row=row_idx, column=base_col + 2).value = sanitize_text(
                    delay.get('comment') or delay.get('delay_description') or ''
                )  # Reason

            # PAX
            pax = flight.get('departurePassengers') or flight.get('arrivalPassengers') or ''
            ws.cell(row=row_idx, column=21).value = sanitize_text(pax)  # PAX

            # Apply styles
            for col in range(1, 22):
                cell = ws.cell(row=row_idx, column=col)
                cell.font = regular_font
                cell.border = border
                if col in [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 12, 13, 15, 16, 18, 19, 21]:
                    cell.alignment = center_align
                else:
                    cell.alignment = left_align

            row_idx += 1
            flight_nr += 1

        # Column widths
        ws.column_dimensions['A'].width = 5   # Nr
        ws.column_dimensions['B'].width = 12  # Date
        ws.column_dimensions['C'].width = 10  # Flight Nr
        ws.column_dimensions['D'].width = 8   # Desk
        ws.column_dimensions['E'].width = 10  # A/C Reg
        ws.column_dimensions['F'].width = 8   # STA
        ws.column_dimensions['G'].width = 8   # ATA
        ws.column_dimensions['H'].width = 8   # STD
        ws.column_dimensions['I'].width = 8   # DCT
        ws.column_dimensions['J'].width = 8   # ATD
        ws.column_dimensions['K'].width = 12  # Total delay
        ws.column_dimensions['L'].width = 6   # Min
        ws.column_dimensions['M'].width = 8   # DL code
        ws.column_dimensions['N'].width = 30  # Reason
        ws.column_dimensions['O'].width = 6   # Min
        ws.column_dimensions['P'].width = 8   # DL code
        ws.column_dimensions['Q'].width = 30  # Reason
        ws.column_dimensions['R'].width = 6   # Min
        ws.column_dimensions['S'].width = 8   # DL code
        ws.column_dimensions['T'].width = 30  # Reason
        ws.column_dimensions['U'].width = 8   # PAX

    # 6. Sačuvaj fajl
    if output_path is None:
        OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
        if day:
            output_path = OUTPUT_DIR / f"Wizz_Air_Performance_{year}-{month:02d}-{day:02d}.xlsx"
        else:
            output_path = OUTPUT_DIR / f"Wizz_Air_Performance_{MONTH_NAMES[month]}_{year}.xlsx"

    print(f"Čuvam izvještaj u: {output_path}")
    wb.save(output_path)
    print(f"✅ Wizz Air Performance izvještaj uspješno generisan!")

    return str(output_path)


if __name__ == "__main__":
    if len(sys.argv) not in (3, 4):
        print("Usage: python generate_wizzair_performance.py <year> <month> [day]")
        print("Example: python generate_wizzair_performance.py 2025 10")
        print("Example (day): python generate_wizzair_performance.py 2025 10 12")
        sys.exit(1)

    year = int(sys.argv[1])
    month = int(sys.argv[2])
    day = int(sys.argv[3]) if len(sys.argv) == 4 else None

    if month < 1 or month > 12:
        print("Mjesec mora biti između 1 i 12")
        sys.exit(1)

    if day:
        max_day = calendar.monthrange(year, month)[1]
        if day < 1 or day > max_day:
            print("Dan nije validan za odabrani mjesec")
            sys.exit(1)

    try:
        output_file = generate_wizzair_performance(year, month, day)
        print(f"\nIzvještaj sačuvan: {output_file}")
    except Exception as e:
        print(f"GREŠKA: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
