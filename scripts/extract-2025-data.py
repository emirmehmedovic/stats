#!/usr/bin/env python3

import openpyxl
import os
import json
import re
from collections import defaultdict

STATS_DIR = "/Users/emir_mw/stats/STATS/2025/Dnevni izvještaji"

airlines = defaultdict(int)
aircraft_types = defaultdict(int)
operation_types = defaultdict(int)

# Valid aircraft type patterns
VALID_AIRCRAFT = re.compile(r'^[A-Z0-9]{2,6}$')
INVALID_AIRCRAFT = re.compile(r'-TZL-|TZL-|^\d{1,3}$')

def normalize_airline(name):
    """Normalize airline name"""
    return name.strip().upper().replace('  ', ' ')

def is_valid_aircraft(aircraft_type):
    """Check if aircraft type is valid"""
    if not aircraft_type:
        return False

    clean = str(aircraft_type).strip()

    # Skip if looks like route
    if INVALID_AIRCRAFT.search(clean):
        return False

    # Skip if too long (probably route)
    if len(clean) > 10:
        return False

    # Valid aircraft pattern
    return VALID_AIRCRAFT.match(clean) is not None

# Parse all months
month_folders = sorted([
    f for f in os.listdir(STATS_DIR)
    if os.path.isdir(os.path.join(STATS_DIR, f))
])

print(f"📂 Found {len(month_folders)} months\n")

total_flights_all = 0

for month_folder in month_folders:
    month_path = os.path.join(STATS_DIR, month_folder)

    # Find Excel file
    files = [
        f for f in os.listdir(month_path)
        if 'Dnevni izvještaj o saobraćaju' in f and f.endswith('.xlsx')
    ]

    if not files:
        continue

    file_path = os.path.join(month_path, files[0])

    try:
        print(f"📊 {month_folder[:10]}: ", end="")

        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        total_flights = 0

        # Iterate through all sheets (days)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            for i, row in enumerate(ws.iter_rows(values_only=True), 1):
                # Skip header
                if i == 1:
                    continue

                # Skip empty rows
                if not row[0]:
                    continue

                total_flights += 1

                # Extract data
                # [0] datum, [1] kompanija, [2] ruta, [3] tip a/c, [4] reg, [5] tip OPER
                airline = row[1]
                aircraft = row[3]
                op_type = row[5]

                # Airline
                if airline and isinstance(airline, str):
                    normalized = normalize_airline(airline)
                    airlines[normalized] += 1

                # Aircraft type
                if aircraft:
                    clean = str(aircraft).strip()
                    if is_valid_aircraft(clean):
                        aircraft_types[clean] += 1

                # Operation type
                if op_type:
                    clean = str(op_type).strip().upper()
                    # Skip if it's a number (capacity)
                    if not re.match(r'^\d+$', clean):
                        operation_types[clean] += 1

        total_flights_all += total_flights
        print(f"✅ {total_flights} flights")
        wb.close()

    except Exception as e:
        print(f"❌ Error: {e}")
        continue

# Sort by count
sorted_airlines = sorted(airlines.items(), key=lambda x: x[1], reverse=True)
sorted_aircraft = sorted(aircraft_types.items(), key=lambda x: x[1], reverse=True)
sorted_op_types = sorted(operation_types.items(), key=lambda x: x[1], reverse=True)

# Print results
print('\n' + '='*80)
print('📊 EKSTRAKTOVANI PODACI IZ 2025 GODINE (CLEAN)')
print('='*80)
print(f'Total flights processed: {total_flights_all}')

print(f'\n✈️  AVIOKOMPANIJE ({len(sorted_airlines)}):')
for name, count in sorted_airlines:
    print(f'  {count:4d} x {name}')

print(f'\n🛩️  TIPOVI AVIONA ({len(sorted_aircraft)}):')
for model, count in sorted_aircraft:
    print(f'  {count:4d} x {model}')

print(f'\n🔧 TIPOVI OPERACIJA ({len(sorted_op_types)}):')
for op_type, count in sorted_op_types:
    print(f'  {count:4d} x {op_type}')

# Save to JSON
output = {
    'airlines': [{'name': name, 'count': count} for name, count in sorted_airlines],
    'aircraftTypes': [{'model': model, 'count': count} for model, count in sorted_aircraft],
    'operationTypes': [{'type': op_type, 'count': count} for op_type, count in sorted_op_types],
    'totalFlights': total_flights_all,
}

os.makedirs('output', exist_ok=True)
with open('output/2025-extracted-data.json', 'w', encoding='utf-8') as f:
    json.dump(output, f, indent=2, ensure_ascii=False)

print('\n✅ Data saved to: output/2025-extracted-data.json')
