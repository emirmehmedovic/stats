#!/usr/bin/env python3

import openpyxl
import os

STATS_DIR = "/Users/emir_mw/stats/STATS/2025/Dnevni izvještaji"

months = [
    "01. JANUAR",
    "02. FEBRUAR",
    "03. MART",
    "04. APRIL",
    "05. MAJ",
    "06. JUNI",
]

for month_folder in months:
    month_path = os.path.join(STATS_DIR, month_folder)

    if not os.path.exists(month_path):
        print(f"❌ {month_folder}: Folder not found")
        continue

    files = [
        f for f in os.listdir(month_path)
        if 'Dnevni izvještaj o saobraćaju' in f and f.endswith('.xlsx')
    ]

    if not files:
        print(f"❌ {month_folder}: No Excel file")
        continue

    file_path = os.path.join(month_path, files[0])

    try:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]

        # Get header row
        header = next(ws.iter_rows(values_only=True))

        print(f"\n{'='*80}")
        print(f"📊 {month_folder}")
        print('='*80)

        # Find key columns
        for i, cell in enumerate(header):
            if cell and ('ruta' in str(cell).lower() or
                        'putnik' in str(cell).lower() or
                        'icao' in str(cell).lower() or
                        'beb' in str(cell).lower()):
                print(f"  [{i:2d}] {cell}")

        wb.close()
    except Exception as e:
        print(f"❌ {month_folder}: Error - {e}")
