#!/usr/bin/env python3

import openpyxl
import os
import re

STATS_DIR = "/Users/emir_mw/stats/STATS/2025/Dnevni izvještaji"
month_folder = "01. JANUAR"
month_path = os.path.join(STATS_DIR, month_folder)

files = [
    f for f in os.listdir(month_path)
    if 'Dnevni izvještaj o saobraćaju' in f and f.endswith('.xlsx')
]

file_path = os.path.join(month_path, files[0])

def parse_passengers(value):
    """Parse passenger string"""
    if not value or value == '-' or value == 'N/A' or str(value).strip() == '':
        return 0, 0

    str_val = str(value).strip()

    # Pattern: "110+6 INF" or "110+6"
    match = re.match(r'^(\d+)\s*\+\s*(\d+)', str_val)
    if match:
        adults = int(match.group(1))
        infants = int(match.group(2))
        return adults, infants

    # Pattern: just a number "110"
    match = re.match(r'^(\d+)$', str_val)
    if match:
        adults = int(match.group(1))
        return adults, 0

    return 0, 0

wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)

total_flights = 0
total_arrival_adults = 0
total_arrival_infants = 0
total_departure_adults = 0
total_departure_infants = 0

flights_with_zero_pax = []

print(f"📊 Counting passengers in Excel file: {files[0]}\n")
print(f"Total sheets (days): {len(wb.sheetnames)}\n")

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]

    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if i == 1:  # Skip header
            continue

        if not row[0]:  # Skip empty rows
            continue

        total_flights += 1

        # Get passenger data
        arrival_pax_str = row[13] if len(row) > 13 else None
        departure_pax_str = row[14] if len(row) > 14 else None

        # Parse
        arr_adults, arr_infants = parse_passengers(arrival_pax_str)
        dep_adults, dep_infants = parse_passengers(departure_pax_str)

        total_arrival_adults += arr_adults
        total_arrival_infants += arr_infants
        total_departure_adults += dep_adults
        total_departure_infants += dep_infants

        # Track flights with zero passengers
        if arr_adults == 0 and arr_infants == 0 and dep_adults == 0 and dep_infants == 0:
            flights_with_zero_pax.append({
                'date': row[0],
                'airline': row[1],
                'route': row[2],
                'arrival_str': arrival_pax_str,
                'departure_str': departure_pax_str,
            })

wb.close()

arrival_total = total_arrival_adults + total_arrival_infants
departure_total = total_departure_adults + total_departure_infants
grand_total = arrival_total + departure_total

print("="*80)
print("📈 EXCEL FILE TOTALS")
print("="*80)
print(f"\nTotal flights: {total_flights}")
print(f"\nArrival:")
print(f"  Adults: {total_arrival_adults}")
print(f"  Infants: {total_arrival_infants}")
print(f"  Total: {arrival_total}")
print(f"\nDeparture:")
print(f"  Adults: {total_departure_adults}")
print(f"  Infants: {total_departure_infants}")
print(f"  Total: {departure_total}")
print(f"\n🎯 GRAND TOTAL: {grand_total}")

print(f"\n⚠️  Flights with ZERO passengers: {len(flights_with_zero_pax)}")
if flights_with_zero_pax:
    print("\nFlights with zero passengers:")
    for f in flights_with_zero_pax:
        print(f"  {f['date']} - {f['airline']} - {f['route']}")
        print(f"    Arrival: '{f['arrival_str']}', Departure: '{f['departure_str']}'")
