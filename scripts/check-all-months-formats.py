#!/usr/bin/env python3

import openpyxl
import os

STATS_DIR = "/Users/emir_mw/stats/STATS/2025/Dnevni izvještaji"

months = [
    "01. JANUAR",
    "02. FEBRUAR",
    "03. MART",
    "04. APRIL",
    "05. MAJ",
    "06. JUNI",
    "07. JULI",
    "08. AUGUST",
    "09. SEPTEMBAR",
    "10. OKTOBAR",
    "11. NOVEMBAR",
]

for month_folder in months:
    month_path = os.path.join(STATS_DIR, month_folder)

    files = [
        f for f in os.listdir(month_path)
        if 'Dnevni izvještaj o saobraćaju' in f and f.endswith('.xlsx')
    ]

    file_path = os.path.join(month_path, files[0])
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)

    # Get first sheet
    ws = wb[wb.sheetnames[0]]
    header = next(ws.iter_rows(values_only=True))

    print(f"\n{'='*80}")
    print(f"📊 {month_folder}")
    print('='*80)

    # Show first 30 columns
    for i in range(min(30, len(header))):
        if header[i]:
            print(f"  [{i:2d}] {header[i]}")

    wb.close()
