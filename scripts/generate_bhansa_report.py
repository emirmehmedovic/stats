#!/usr/bin/env python3
"""
BHANSA Monthly Airport Traffic Report Generator

Generiše mjesečni izvještaj o prometu na aerodromu Tuzla za BHANSA
Format: Lista letova sa detaljima o avionima, putnicima i rutama
"""

import sys
import os
import re
import calendar
from pathlib import Path
from datetime import datetime
import psycopg2
from psycopg2.extras import RealDictCursor
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

# Putanja do template-a (koristićemo novi template za BHANSA)
SCRIPT_DIR = Path(__file__).parent
PROJECT_ROOT = SCRIPT_DIR.parent
TEMPLATE_PATH = PROJECT_ROOT / "izvještaji" / "templates" / "BHANSA_template.xlsx"
OUTPUT_DIR = PROJECT_ROOT / "izvještaji" / "generated"

# Nazivi mjeseci
MONTH_NAMES = {
    1: "Januar", 2: "Februar", 3: "Mart", 4: "April",
    5: "Maj", 6: "Juni", 7: "Juli", 8: "Avgust",
    9: "Septembar", 10: "Oktobar", 11: "Novembar", 12: "Decembar"
}

MONTH_NAMES_UPPER = {
    1: "JANUAR", 2: "FEBRUAR", 3: "MART", 4: "APRIL",
    5: "MAJ", 6: "JUNI", 7: "JULI", 8: "AVGUST",
    9: "SEPTEMBAR", 10: "OKTOBAR", 11: "NOVEMBAR", 12: "DECEMBAR"
}


def get_db_connection():
    """Konekcija na PostgreSQL bazu"""
    DATABASE_URL = os.getenv("DATABASE_URL")
    if not DATABASE_URL:
        raise ValueError("DATABASE_URL environment variable not set")

    return psycopg2.connect(DATABASE_URL, cursor_factory=RealDictCursor)


def sanitize_text(value):
    """
    ULTRA-AGGRESSIVE sanitization to prevent ANY Excel corruption.
    Removes control chars, invalid unicode, and normalizes to ASCII-safe characters.
    """
    if value is None:
        return None
    if not isinstance(value, str):
        return value

    # Step 1: Normalize unicode (removes accents, diacritics to ASCII equivalents)
    import unicodedata
    try:
        # Try ASCII transliteration
        value = unicodedata.normalize('NFKD', value)
        value = value.encode('ascii', 'ignore').decode('ascii')
    except Exception:
        pass

    # Step 2: Remove ALL control characters
    value = re.sub(r'[\x00-\x1F\x7F-\x9F]', '', value)

    # Step 3: Use openpyxl's built-in cleaner
    try:
        from openpyxl.utils.cell import ILLEGAL_CHARACTERS_RE
        value = ILLEGAL_CHARACTERS_RE.sub('', value)
    except Exception:
        pass

    # Step 4: Remove any remaining problematic chars
    value = value.replace('\ufeff', '')  # BOM
    value = value.replace('\u200b', '')  # Zero-width space
    value = value.replace('\u200c', '')  # Zero-width non-joiner
    value = value.replace('\u200d', '')  # Zero-width joiner
    value = value.replace('\xa0', ' ')   # Non-breaking space -> normal space

    # Step 5: Strip whitespace
    value = value.strip()

    return value


def create_merged_cell(ws, row, start_col, end_col, value, font=None, fill=None, alignment=None, border=None):
    """
    Helper function to properly create merged cells with consistent formatting.
    Applies formatting to ALL cells in the range before merging to avoid Excel corruption.
    """
    # Convert column letters to numbers if needed
    if isinstance(start_col, str):
        start_col = openpyxl.utils.column_index_from_string(start_col)
    if isinstance(end_col, str):
        end_col = openpyxl.utils.column_index_from_string(end_col)

    for col in range(start_col, end_col + 1):
        cell = ws.cell(row=row, column=col)
        if col == start_col and value is not None:
            cell.value = sanitize_text(value) if isinstance(value, str) else value
        if font:
            cell.font = font
        if fill:
            cell.fill = fill
        if alignment:
            cell.alignment = alignment
        if border:
            cell.border = border

    if start_col != end_col:
        ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=end_col)


def get_operation_type_code(operation_type_code):
    """
    Mapira operation type code u Trip fare oznaku
    R = Redovan (Scheduled)
    H = Charter
    """
    if not operation_type_code:
        return ""

    code = operation_type_code.upper()
    if code == "SCHEDULED":
        return "R"
    elif code == "CHARTER":
        return "H"
    else:
        # Za ostale tipove (MEDEVAC, CARGO, MILITARY, GENERAL_AVIATION)
        return code[0] if code else ""


def get_flight_data(year: int, month: int):
    """
    Povuči sve letove za zadati mjesec sortiran po datumu
    """
    conn = get_db_connection()
    cursor = conn.cursor()

    # Datum početka i kraja mjeseca
    first_day = f"{year}-{month:02d}-01"
    last_day = calendar.monthrange(year, month)[1]
    last_date = f"{year}-{month:02d}-{last_day:02d}"

    query = """
        SELECT
            f.id,
            f.date,
            f.route,
            f.registration,

            -- Airline
            a.name as airline_name,
            a."icaoCode" as airline_icao,
            a."iataCode" as airline_iata,
            a.address as airline_address,

            -- Aircraft Type
            at.model as aircraft_model,
            at.mtow as aircraft_mtow,

            -- Operation Type
            ot.code as operation_type_code,

            -- Departure info
            f."departureFlightNumber",
            f."departurePassengers",
            f."departureInfants",
            f."departureScheduledTime",
            f."departureActualTime",
            f."departureFerryOut",

            -- Arrival info
            f."arrivalFlightNumber",
            f."arrivalPassengers",
            f."arrivalInfants",
            f."arrivalScheduledTime",
            f."arrivalActualTime",
            f."arrivalFerryIn",

            -- Airports
            dep_ap."iataCode" as departure_airport_iata,
            arr_ap."iataCode" as arrival_airport_iata

        FROM "Flight" f
        INNER JOIN "Airline" a ON f."airlineId" = a.id
        INNER JOIN "AircraftType" at ON f."aircraftTypeId" = at.id
        INNER JOIN "OperationType" ot ON f."operationTypeId" = ot.id
        LEFT JOIN "Airport" arr_ap ON f."arrivalAirportId" = arr_ap.id
        LEFT JOIN "Airport" dep_ap ON f."departureAirportId" = dep_ap.id
        WHERE f.date >= %s AND f.date <= %s
        ORDER BY f.date, f."departureScheduledTime", f."arrivalScheduledTime"
    """

    cursor.execute(query, (first_day, last_date))
    flights = cursor.fetchall()

    cursor.close()
    conn.close()

    return flights


def parse_route(route_str):
    """
    Parse route string to extract airport IATA code.
    Route format in DB: "DTM-Dortmund", "FMM-Memmingen", "MMX-Malmö-Sturup"
    Returns: IATA code (first part before dash) or None

    Since this is Tuzla airport database, the route represents the OTHER airport.
    TZL is always one end of the journey.
    """
    if not route_str or '-' not in route_str:
        return None

    # Take only the first part (IATA code)
    parts = route_str.strip().split('-')
    iata_code = parts[0].strip()

    return iata_code if iata_code else None


def format_pax_with_infants(pax, infants):
    """
    Format passengers with infants: "131+1 INF" or just "131" if no infants
    """
    if pax is None:
        pax = 0

    if infants and infants > 0:
        return f"{pax}+{infants} INF"
    else:
        return str(pax)


def generate_bhansa_report(year: int, month: int, output_path: Path = None):
    """
    Glavni metod za generisanje BHANSA izvještaja
    """
    print(f"Generišem BHANSA izvještaj za {MONTH_NAMES[month]} {year}...")

    # 1. Povući podatke iz baze
    print("Povlačim podatke iz baze...")
    flights = get_flight_data(year, month)
    print(f"Pronađeno {len(flights)} letova.")

    if len(flights) == 0:
        print("UPOZORENJE: Nema letova za zadati period!")
        return None

    # 2. Kreiranje Excel workbook-a
    print("Kreiram Excel workbook...")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # 3. Stilovi
    header_font = Font(name='Arial', size=11, bold=True)
    header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    title_font = Font(name='Arial', size=14, bold=True)
    regular_font = Font(name='Arial', size=10)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')

    # 4. Header - Row 1
    create_merged_cell(
        ws, 1, 'A', 'N',
        f"AERODROM TUZLA                    RIJD {MONTH_NAMES_UPPER[month]} {year}",
        font=title_font,
        alignment=center_align
    )

    # 5. Column Headers - Row 2
    headers = ['Nr.', 'DATE', 'COMPANY', 'A/C', 'REG.', 'MTOW/T', 'FLT.NMB',
               'FROM', 'TO', 'ADDRESS', 'DEP PAX', 'ARR PAX', 'Trip fare', 'Day']

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col_idx)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = center_align

    # 6. Popunjavanje podataka
    print("Popunjavam podatke...")
    row_idx = 3
    flight_counter = 1

    TZL_IATA = 'TZL'

    for flight in flights:
        # Parse route to get OTHER airport (TZL is always one end)
        other_airport = None
        route = flight.get('route')
        if route:
            other_airport = parse_route(route)

        # Debug first 3 flights
        if flight_counter <= 3:
            print(f"\n[DEBUG] Flight {flight_counter}:")
            print(f"  route: {route}")
            print(f"  other_airport: {other_airport}")
            print(f"  departurePassengers: {flight.get('departurePassengers')}")
            print(f"  arrivalPassengers: {flight.get('arrivalPassengers')}")
            print(f"  departureFlightNumber: {flight.get('departureFlightNumber')}")
            print(f"  arrivalFlightNumber: {flight.get('arrivalFlightNumber')}")

        # Skip if we can't determine the route
        if not other_airport:
            continue

        # Determine if this is departure or arrival (or both)
        has_departure = flight.get('departurePassengers') is not None and flight.get('departurePassengers') > 0
        has_arrival = flight.get('arrivalPassengers') is not None and flight.get('arrivalPassengers') > 0

        # If no passenger data, check flight numbers
        if not has_departure and not has_arrival:
            has_departure = flight.get('departureFlightNumber') is not None
            has_arrival = flight.get('arrivalFlightNumber') is not None

        # If still no data, assume it's a departure (default)
        if not has_departure and not has_arrival:
            has_departure = True

        # Departure leg (TZL -> Other)
        if has_departure:
            ws.cell(row=row_idx, column=1).value = flight_counter  # Nr
            ws.cell(row=row_idx, column=2).value = flight['date'].strftime('%d/%m/%Y')  # DATE
            ws.cell(row=row_idx, column=3).value = sanitize_text(flight['airline_name'])  # COMPANY
            ws.cell(row=row_idx, column=4).value = sanitize_text(flight['aircraft_model'])  # A/C
            ws.cell(row=row_idx, column=5).value = sanitize_text(flight['registration'])  # REG
            ws.cell(row=row_idx, column=6).value = flight['aircraft_mtow'] / 1000 if flight['aircraft_mtow'] else None  # MTOW in tonnes
            ws.cell(row=row_idx, column=7).value = sanitize_text(flight['departureFlightNumber'] or '')  # FLT.NMB
            # Za DEPARTURE: FROM=TZL, TO=other_airport
            ws.cell(row=row_idx, column=8).value = sanitize_text(TZL_IATA)  # FROM
            ws.cell(row=row_idx, column=9).value = sanitize_text(other_airport)  # TO
            ws.cell(row=row_idx, column=10).value = sanitize_text(flight['airline_address'] or '')  # ADDRESS
            ws.cell(row=row_idx, column=11).value = sanitize_text(format_pax_with_infants(
                flight.get('departurePassengers'),
                flight.get('departureInfants')
            ))  # DEP PAX
            ws.cell(row=row_idx, column=12).value = ''  # ARR PAX (empty for departure)
            ws.cell(row=row_idx, column=13).value = sanitize_text(get_operation_type_code(flight['operation_type_code']))  # Trip fare
            ws.cell(row=row_idx, column=14).value = sanitize_text(flight['date'].strftime('%A'))  # Day of week

            # Apply styles
            for col in range(1, 15):
                cell = ws.cell(row=row_idx, column=col)
                cell.font = regular_font
                cell.border = border
                if col in [1, 2, 4, 5, 6, 7, 8, 9, 13, 14]:
                    cell.alignment = center_align
                else:
                    cell.alignment = left_align

            row_idx += 1
            flight_counter += 1

        # Arrival leg (Other -> TZL)
        if has_arrival:
            # Samo dodaj arrival ako je različit od departure (ne dupliramo isti let)
            if has_arrival and not has_departure:  # Samo arrival, bez departure
                ws.cell(row=row_idx, column=1).value = flight_counter
                ws.cell(row=row_idx, column=2).value = flight['date'].strftime('%d/%m/%Y')
                ws.cell(row=row_idx, column=3).value = sanitize_text(flight['airline_name'])
                ws.cell(row=row_idx, column=4).value = sanitize_text(flight['aircraft_model'])
                ws.cell(row=row_idx, column=5).value = sanitize_text(flight['registration'])
                ws.cell(row=row_idx, column=6).value = flight['aircraft_mtow'] / 1000 if flight['aircraft_mtow'] else None
                ws.cell(row=row_idx, column=7).value = sanitize_text(flight['arrivalFlightNumber'] or '')
                # Za ARRIVAL: FROM=other_airport, TO=TZL
                ws.cell(row=row_idx, column=8).value = sanitize_text(other_airport)
                ws.cell(row=row_idx, column=9).value = sanitize_text(TZL_IATA)
                ws.cell(row=row_idx, column=10).value = sanitize_text(flight['airline_address'] or '')
                ws.cell(row=row_idx, column=11).value = ''  # DEP PAX (empty for arrival)
                ws.cell(row=row_idx, column=12).value = sanitize_text(format_pax_with_infants(
                    flight.get('arrivalPassengers'),
                    flight.get('arrivalInfants')
                ))  # ARR PAX
                ws.cell(row=row_idx, column=13).value = sanitize_text(get_operation_type_code(flight['operation_type_code']))
                ws.cell(row=row_idx, column=14).value = sanitize_text(flight['date'].strftime('%A'))

                # Apply styles
                for col in range(1, 15):
                    cell = ws.cell(row=row_idx, column=col)
                    cell.font = regular_font
                    cell.border = border
                    if col in [1, 2, 4, 5, 6, 7, 8, 9, 13, 14]:
                        cell.alignment = center_align
                    else:
                        cell.alignment = left_align

                row_idx += 1
                flight_counter += 1

    # 7. Podešavanje širine kolona
    ws.column_dimensions['A'].width = 5   # Nr
    ws.column_dimensions['B'].width = 12  # DATE
    ws.column_dimensions['C'].width = 25  # COMPANY
    ws.column_dimensions['D'].width = 8   # A/C
    ws.column_dimensions['E'].width = 10  # REG
    ws.column_dimensions['F'].width = 10  # MTOW/T
    ws.column_dimensions['G'].width = 12  # FLT.NMB
    ws.column_dimensions['H'].width = 8   # FROM
    ws.column_dimensions['I'].width = 8   # TO
    ws.column_dimensions['J'].width = 50  # ADDRESS
    ws.column_dimensions['K'].width = 12  # DEP PAX
    ws.column_dimensions['L'].width = 12  # ARR PAX
    ws.column_dimensions['M'].width = 10  # Trip fare
    ws.column_dimensions['N'].width = 12  # Day

    # 8. Sačuvaj fajl
    if output_path is None:
        OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
        output_path = OUTPUT_DIR / f"BHANSA_{MONTH_NAMES[month]}_{year}.xlsx"

    print(f"Čuvam izvještaj u: {output_path}")
    wb.iso_dates = True
    wb.save(output_path)
    wb.close()
    print(f"✅ BHANSA izvještaj uspješno generisan!")

    return str(output_path)


if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("Usage: python generate_bhansa_report.py <year> <month>")
        print("Example: python generate_bhansa_report.py 2025 10")
        sys.exit(1)

    year = int(sys.argv[1])
    month = int(sys.argv[2])

    if month < 1 or month > 12:
        print("Mjesec mora biti između 1 i 12")
        sys.exit(1)

    try:
        output_file = generate_bhansa_report(year, month)
        print(f"\nIzvještaj sačuvan: {output_file}")
    except Exception as e:
        print(f"GREŠKA: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
