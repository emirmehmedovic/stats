#!/usr/bin/env python3
"""
Statistika za carinu - Customs Statistics Report Generator

Generiše mjesečni statistički izvještaj za carinu sa:
- Putnici (Redovni promet po kompanijama, Vanredni promet, Ostala slijetanja)
- Teret (Utovareno, Istovareno)
"""

import sys
import os
import re
import calendar
from pathlib import Path
from datetime import datetime
import psycopg2
from psycopg2.extras import RealDictCursor
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

# Putanja
SCRIPT_DIR = Path(__file__).parent
PROJECT_ROOT = SCRIPT_DIR.parent
OUTPUT_DIR = PROJECT_ROOT / "izvještaji" / "generated"

# Nazivi mjeseci
MONTH_NAMES = {
    1: "Januar", 2: "Februar", 3: "Mart", 4: "April",
    5: "Maj", 6: "Juni", 7: "Juli", 8: "Avgust",
    9: "Septembar", 10: "Oktobar", 11: "Novembar", 12: "Decembar"
}


def get_db_connection():
    """Konekcija na PostgreSQL bazu"""
    DATABASE_URL = os.getenv("DATABASE_URL")
    if not DATABASE_URL:
        raise ValueError("DATABASE_URL environment variable not set")

    return psycopg2.connect(DATABASE_URL, cursor_factory=RealDictCursor)


def sanitize_text(value):
    """
    ULTRA-AGGRESSIVE sanitization to prevent ANY Excel corruption.
    Removes control chars, invalid unicode, and normalizes to ASCII-safe characters.
    """
    if value is None:
        return None
    if not isinstance(value, str):
        return value

    # Step 1: Normalize unicode
    import unicodedata
    try:
        value = unicodedata.normalize('NFKD', value)
        value = value.encode('ascii', 'ignore').decode('ascii')
    except Exception:
        pass

    # Step 2: Remove ALL control characters
    value = re.sub(r'[\x00-\x1F\x7F-\x9F]', '', value)

    # Step 3: Use openpyxl's built-in cleaner
    try:
        from openpyxl.utils.cell import ILLEGAL_CHARACTERS_RE
        value = ILLEGAL_CHARACTERS_RE.sub('', value)
    except Exception:
        pass

    # Step 4: Remove problematic chars
    value = value.replace('\ufeff', '').replace('\u200b', '').replace('\u200c', '').replace('\u200d', '').replace('\xa0', ' ')

    # Step 5: Strip whitespace
    return value.strip()


def create_merged_cell(ws, row, start_col, end_col, value, font=None, fill=None, alignment=None, border=None):
    """
    Helper function to properly create merged cells with consistent formatting.
    Applies formatting to ALL cells in the range before merging to avoid Excel corruption.
    """
    # Convert column letters to numbers if needed
    if isinstance(start_col, str):
        start_col = openpyxl.utils.column_index_from_string(start_col)
    if isinstance(end_col, str):
        end_col = openpyxl.utils.column_index_from_string(end_col)

    for col in range(start_col, end_col + 1):
        cell = ws.cell(row=row, column=col)
        if col == start_col and value is not None:
            cell.value = sanitize_text(value) if isinstance(value, str) else value
        if font:
            cell.font = font
        if fill:
            cell.fill = fill
        if alignment:
            cell.alignment = alignment
        if border:
            cell.border = border

    if start_col != end_col:
        ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=end_col)


def get_flight_data(year: int, month: int):
    """
    Povuči sve letove za zadati mjesec
    """
    conn = get_db_connection()
    cursor = conn.cursor()

    # Datum početka i kraja mjeseca
    first_day = f"{year}-{month:02d}-01"
    last_day = calendar.monthrange(year, month)[1]
    last_date = f"{year}-{month:02d}-{last_day:02d}"

    query = """
        SELECT
            f.id,
            f.date,

            -- Airline
            a.name as airline_name,
            a."icaoCode" as airline_icao,
            a."iataCode" as airline_iata,

            -- Operation Type
            ot.code as operation_type_code,

            -- Departure info
            f."departurePassengers",
            f."departureCargo",
            f."departureMail",
            f."departureStatus",
            f."departureFerryOut",

            -- Arrival info
            f."arrivalPassengers",
            f."arrivalCargo",
            f."arrivalMail",
            f."arrivalStatus",
            f."arrivalFerryIn"

        FROM "Flight" f
        INNER JOIN "Airline" a ON f."airlineId" = a.id
        INNER JOIN "OperationType" ot ON f."operationTypeId" = ot.id
        WHERE f.date >= %s AND f.date <= %s
        ORDER BY a.name
    """

    cursor.execute(query, (first_day, last_date))
    flights = cursor.fetchall()

    cursor.close()
    conn.close()

    return flights


def aggregate_customs_data(flights):
    """
    Agregira podatke za statistiku carinskog izvještaja

    Returns:
    {
        'scheduled': {
            'WIZZ AIR HUNGARY LTD': {'flights': 35, 'embarked': 6468, 'disembarked': 5350},
            'WIZZAIR MALTA LTD': {...},
            ...
        },
        'non_scheduled': {'flights': 34, 'embarked': 1077, 'disembarked': 1604},
        'other_landings': {'flights': 2, 'embarked': 0, 'disembarked': 0},
        'freight': {'loaded': 0, 'unloaded': 0}
    }
    """
    data = {
        'scheduled': {},  # Po aviokompanijama
        'non_scheduled': {'flights': 0, 'embarked': 0, 'disembarked': 0},
        'other_landings': {'flights': 0, 'embarked': 0, 'disembarked': 0},
        'freight': {'loaded': 0, 'unloaded': 0}
    }

    for flight in flights:
        operation_code = flight.get('operation_type_code', '').upper()
        is_scheduled = operation_code == 'SCHEDULED'
        airline_name = flight['airline_name']

        # Check if this is a ferry flight (other landings)
        is_ferry = flight.get('arrivalFerryIn') or flight.get('departureFerryOut')

        # Count movements (flights)
        # Each flight record can have both arrival and departure
        has_arrival = flight.get('arrivalPassengers') is not None or flight.get('arrivalStatus') == 'OPERATED'
        has_departure = flight.get('departurePassengers') is not None or flight.get('departureStatus') == 'OPERATED'

        # Determine category
        if is_ferry:
            category = 'other_landings'
        elif is_scheduled:
            category = 'scheduled'
        else:
            category = 'non_scheduled'

        # For scheduled flights, group by airline
        if category == 'scheduled':
            if airline_name not in data['scheduled']:
                data['scheduled'][airline_name] = {'flights': 0, 'embarked': 0, 'disembarked': 0}

            # Count flights (movements)
            if has_arrival:
                data['scheduled'][airline_name]['flights'] += 1
            if has_departure and not has_arrival:  # Don't double count if it's the same flight
                data['scheduled'][airline_name]['flights'] += 1

            # Passengers
            if flight.get('departurePassengers'):
                data['scheduled'][airline_name]['embarked'] += flight['departurePassengers']
            if flight.get('arrivalPassengers'):
                data['scheduled'][airline_name]['disembarked'] += flight['arrivalPassengers']

        else:
            # Non-scheduled or other landings
            target = data[category]

            # Count flights
            if has_arrival:
                target['flights'] += 1
            if has_departure and not has_arrival:
                target['flights'] += 1

            # Passengers
            if flight.get('departurePassengers'):
                target['embarked'] += flight['departurePassengers']
            if flight.get('arrivalPassengers'):
                target['disembarked'] += flight['arrivalPassengers']

        # Freight (for all flights)
        if flight.get('departureCargo'):
            data['freight']['loaded'] += flight['departureCargo'] / 1000.0  # Convert to tonnes
        if flight.get('arrivalCargo'):
            data['freight']['unloaded'] += flight['arrivalCargo'] / 1000.0

        if flight.get('departureMail'):
            data['freight']['loaded'] += flight['departureMail'] / 1000.0
        if flight.get('arrivalMail'):
            data['freight']['unloaded'] += flight['arrivalMail'] / 1000.0

    # Round freight to 2 decimals
    data['freight']['loaded'] = round(data['freight']['loaded'], 2)
    data['freight']['unloaded'] = round(data['freight']['unloaded'], 2)

    return data


def generate_customs_stats(year: int, month: int, output_path: Path = None):
    """
    Glavni metod za generisanje Customs statistike
    """
    print(f"Generišem statistiku za carinu za {MONTH_NAMES[month]} {year}...")

    # 1. Povući podatke iz baze
    print("Povlačim podatke iz baze...")
    flights = get_flight_data(year, month)
    print(f"Pronađeno {len(flights)} letova.")

    if len(flights) == 0:
        print("UPOZORENJE: Nema letova za zadati period!")
        return None

    # 2. Agregirati podatke
    print("Agregatiranje podataka...")
    data = aggregate_customs_data(flights)

    # 3. Kreiranje Excel workbook-a
    print("Kreiram Excel workbook...")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Statistika"

    # 4. Stilovi
    title_font = Font(name='Arial', size=14, bold=True)
    section_font = Font(name='Arial', size=11, bold=True)
    header_font = Font(name='Arial', size=10, bold=True)
    regular_font = Font(name='Arial', size=10)
    indent_font = Font(name='Arial', size=10, italic=True)
    bold_font = Font(name='Arial', size=10, bold=True)

    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')
    right_align = Alignment(horizontal='right', vertical='center')

    # 5. Popunjavanje podataka
    row = 1

    # Title
    create_merged_cell(
        ws, row, 'A', 'E',
        f"{MONTH_NAMES[month]} {year}",
        font=title_font,
        alignment=center_align
    )
    row += 2

    # PUTNICI Section
    create_merged_cell(
        ws, row, 'B', 'E',
        "PUTNICI",
        font=section_font,
        alignment=center_align
    )
    row += 1

    # Headers
    ws[f'B{row}'] = "Broj letova"
    ws[f'C{row}'] = "Ukrcano"
    ws[f'D{row}'] = "Iskrcano"
    ws[f'E{row}'] = "Ukupno"
    for col in ['B', 'C', 'D', 'E']:
        ws[f'{col}{row}'].font = header_font
        ws[f'{col}{row}'].alignment = center_align
    row += 1

    # Calculate scheduled totals
    scheduled_total_flights = sum(airline['flights'] for airline in data['scheduled'].values())
    scheduled_total_embarked = sum(airline['embarked'] for airline in data['scheduled'].values())
    scheduled_total_disembarked = sum(airline['disembarked'] for airline in data['scheduled'].values())
    scheduled_total = scheduled_total_embarked + scheduled_total_disembarked

    # Redovni promet
    ws[f'A{row}'] = "Redovni promet"
    ws[f'A{row}'].font = bold_font
    ws[f'B{row}'] = scheduled_total_flights
    ws[f'C{row}'] = scheduled_total_embarked
    ws[f'D{row}'] = scheduled_total_disembarked
    ws[f'E{row}'] = scheduled_total
    for col in ['B', 'C', 'D', 'E']:
        ws[f'{col}{row}'].font = regular_font
        ws[f'{col}{row}'].alignment = right_align
    row += 1

    # Airlines breakdown (sorted alphabetically)
    for airline_name in sorted(data['scheduled'].keys()):
        airline_data = data['scheduled'][airline_name]
        ws[f'A{row}'] = f"  {airline_name}"
        ws[f'A{row}'].font = indent_font
        ws[f'B{row}'] = airline_data['flights']
        ws[f'C{row}'] = airline_data['embarked']
        ws[f'D{row}'] = airline_data['disembarked']
        ws[f'E{row}'] = airline_data['embarked'] + airline_data['disembarked']
        for col in ['B', 'C', 'D', 'E']:
            ws[f'{col}{row}'].font = regular_font
            ws[f'{col}{row}'].alignment = right_align
        row += 1

    # Vanredni promet
    ws[f'A{row}'] = "Vanredni promet"
    ws[f'A{row}'].font = bold_font
    ws[f'B{row}'] = data['non_scheduled']['flights']
    ws[f'C{row}'] = data['non_scheduled']['embarked']
    ws[f'D{row}'] = data['non_scheduled']['disembarked']
    ws[f'E{row}'] = data['non_scheduled']['embarked'] + data['non_scheduled']['disembarked']
    for col in ['B', 'C', 'D', 'E']:
        ws[f'{col}{row}'].font = regular_font
        ws[f'{col}{row}'].alignment = right_align
    row += 1

    # Ostala slijetanja
    ws[f'A{row}'] = "Ostala slijetanja"
    ws[f'A{row}'].font = bold_font
    ws[f'B{row}'] = data['other_landings']['flights']
    ws[f'C{row}'] = data['other_landings']['embarked']
    ws[f'D{row}'] = data['other_landings']['disembarked']
    ws[f'E{row}'] = data['other_landings']['embarked'] + data['other_landings']['disembarked']
    for col in ['B', 'C', 'D', 'E']:
        ws[f'{col}{row}'].font = regular_font
        ws[f'{col}{row}'].alignment = right_align
    row += 2

    # UKUPNO
    total_flights = scheduled_total_flights + data['non_scheduled']['flights'] + data['other_landings']['flights']
    total_embarked = scheduled_total_embarked + data['non_scheduled']['embarked'] + data['other_landings']['embarked']
    total_disembarked = scheduled_total_disembarked + data['non_scheduled']['disembarked'] + data['other_landings']['disembarked']
    total_passengers = total_embarked + total_disembarked

    ws[f'A{row}'] = "UKUPNO"
    ws[f'A{row}'].font = bold_font
    ws[f'B{row}'] = total_flights
    ws[f'C{row}'] = total_embarked
    ws[f'D{row}'] = total_disembarked
    ws[f'E{row}'] = total_passengers
    for col in ['B', 'C', 'D', 'E']:
        ws[f'{col}{row}'].font = bold_font
        ws[f'{col}{row}'].alignment = right_align
    row += 3

    # TERET Section
    create_merged_cell(
        ws, row, 'B', 'D',
        "TERET",
        font=section_font,
        alignment=center_align
    )
    row += 1

    # Headers
    ws[f'B{row}'] = "Utovareno"
    ws[f'C{row}'] = "Istovareno"
    ws[f'D{row}'] = "Ukupno"
    for col in ['B', 'C', 'D']:
        ws[f'{col}{row}'].font = header_font
        ws[f'{col}{row}'].alignment = center_align
    row += 1

    # Freight data
    ws[f'B{row}'] = data['freight']['loaded']
    ws[f'C{row}'] = data['freight']['unloaded']
    ws[f'D{row}'] = data['freight']['loaded'] + data['freight']['unloaded']
    for col in ['B', 'C', 'D']:
        ws[f'{col}{row}'].font = regular_font
        ws[f'{col}{row}'].alignment = right_align

    # 6. Column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15

    # 7. Sačuvaj fajl
    if output_path is None:
        OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
        output_path = OUTPUT_DIR / f"Statistika_za_carinu_{MONTH_NAMES[month]}_{year}.xlsx"

    print(f"Čuvam izvještaj u: {output_path}")
    wb.iso_dates = True
    wb.save(output_path)
    wb.close()
    print(f"✅ Statistika za carinu uspješno generisana!")

    return str(output_path)


if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("Usage: python generate_customs_stats.py <year> <month>")
        print("Example: python generate_customs_stats.py 2025 9")
        sys.exit(1)

    year = int(sys.argv[1])
    month = int(sys.argv[2])

    if month < 1 or month > 12:
        print("Mjesec mora biti između 1 i 12")
        sys.exit(1)

    try:
        output_file = generate_customs_stats(year, month)
        print(f"\nIzvještaj sačuvan: {output_file}")
    except Exception as e:
        print(f"GREŠKA: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
