#!/usr/bin/env python3
"""
BHDCA (Bosnia and Herzegovina Department of Civil Aviation) Report Generator

Generiše mjesečni ICAO izvještaj o saobraćaju aerodroma prema BHDCA standardima.
Izvještaj sadrži 3 sheet-a:
1. AIRPORT TRAFFIC - Osnovni saobraćaj (aircraft movements, passengers, freight, mail)
2. O-D TRAFFIC (Scheduled) - City-pair podaci za scheduled letove
3. O-D TRAFFIC (Non-Scheduled) - City-pair podaci za non-scheduled letove
"""

import sys
import os
import re
from datetime import datetime
from pathlib import Path
import openpyxl
from openpyxl.styles import Alignment, Font, Border, Side
import psycopg2
from psycopg2.extras import RealDictCursor
import calendar

# Putanje
PROJECT_ROOT = Path(__file__).parent.parent
TEMPLATE_PATH = PROJECT_ROOT / "izvještaji" / "09. BHDCA Septembar 2025.xlsx"
OUTPUT_DIR = PROJECT_ROOT / "izvještaji" / "generated"

# Kreirati output direktorijum ako ne postoji
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

# Konstante za header informacije
CONTACT_PERSON = "Haris Tufekčić"
ORGANIZATION = "Tuzla International Airport"
PHONE = "387 35 814 605"
FAX = "387 35 745 750"
EMAIL = "opcentar@bih.net.ba"
STATE = "Bosnia and Herzegovina"
AIRPORT_NAME = "Tuzla International Airport"
CITY_NAME = "Živinice"

# Mapiranje mjeseca (1-12) na nazive
MONTH_NAMES = {
    1: "January", 2: "February", 3: "March", 4: "April",
    5: "May", 6: "June", 7: "July", 8: "August",
    9: "September", 10: "October", 11: "November", 12: "December"
}


def get_db_connection():
    """Konekcija na PostgreSQL bazu"""
    DATABASE_URL = os.getenv("DATABASE_URL")
    if not DATABASE_URL:
        raise ValueError("DATABASE_URL environment variable not set")

    return psycopg2.connect(DATABASE_URL, cursor_factory=RealDictCursor)


def sanitize_text(value):
    """
    Remove control characters that can corrupt Excel XML.
    Uklanja kontrolne karaktere koji mogu oštetiti Excel XML.
    """
    if value is None:
        return None
    if not isinstance(value, str):
        return value
    try:
        from openpyxl.utils.cell import ILLEGAL_CHARACTERS_RE
        value = ILLEGAL_CHARACTERS_RE.sub('', value)
    except Exception:
        value = re.sub(r'[\x00-\x08\x0B\x0C\x0E-\x1F]', '', value)
    return value


def create_merged_cell(ws, row, start_col, end_col, value, font=None, fill=None, alignment=None, border=None):
    """
    Helper function to properly create merged cells with consistent formatting.
    Applies formatting to ALL cells in the range before merging to avoid Excel corruption.
    """
    # Convert column letters to numbers if needed
    if isinstance(start_col, str):
        start_col = openpyxl.utils.column_index_from_string(start_col)
    if isinstance(end_col, str):
        end_col = openpyxl.utils.column_index_from_string(end_col)

    for col in range(start_col, end_col + 1):
        cell = ws.cell(row=row, column=col)
        if col == start_col and value is not None:
            cell.value = sanitize_text(value) if isinstance(value, str) else value
        if font:
            cell.font = font
        if fill:
            cell.fill = fill
        if alignment:
            cell.alignment = alignment
        if border:
            cell.border = border

    if start_col != end_col:
        ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=end_col)


def clear_merged_cell_range(ws, start_row, end_row, columns):
    """
    Safely clear values in a range that may contain merged cells.
    Only clears the TOP-LEFT cell of each merged range, which is the only writable cell.

    Args:
        ws: Worksheet object
        start_row: First row to clear
        end_row: Last row to clear
        columns: List of column numbers to clear
    """
    # Identify all merged cell top-left positions in our target range
    merged_top_left_cells = set()
    for merged_range in list(ws.merged_cells.ranges):
        # Check if this merged range intersects with our target range
        if (merged_range.min_row >= start_row and merged_range.max_row <= end_row):
            # Check if any of the columns intersect
            for col in columns:
                if col >= merged_range.min_col and col <= merged_range.max_col:
                    # Store top-left cell position
                    merged_top_left_cells.add((merged_range.min_row, merged_range.min_col))
                    break

    # Clear values - only clear top-left cells of merged ranges
    for row in range(start_row, end_row + 1):
        for col in columns:
            cell = ws.cell(row=row, column=col)
            # Check if this is a merged cell
            if hasattr(cell, '__class__') and 'MergedCell' in cell.__class__.__name__:
                # Skip merged cells (not top-left)
                continue
            # Clear the cell value
            try:
                cell.value = None
            except AttributeError:
                # This is a merged cell that's read-only, skip it
                continue


def get_flight_data(year: int, month: int):
    """Povuči podatke o letovima za zadati mjesec"""
    conn = get_db_connection()
    cursor = conn.cursor()

    # Datum početka i kraja mjeseca
    first_day = f"{year}-{month:02d}-01"
    last_day = calendar.monthrange(year, month)[1]
    last_date = f"{year}-{month:02d}-{last_day:02d}"

    query = """
        SELECT
            f.id,
            f.date,
            f.route,
            f.registration,

            -- Airline
            a.name as airline_name,
            a."icaoCode" as airline_icao,
            a."iataCode" as airline_iata,

            -- Aircraft Type
            at.model as aircraft_model,
            at.seats as aircraft_seats,

            -- Operation Type
            ot.code as operation_type_code,
            ot.name as operation_type_name,

            -- Flight Type
            ft.code as flight_type_code,
            ft.name as flight_type_name,

            -- Arrival
            f."arrivalFlightNumber",
            f."arrivalScheduledTime",
            f."arrivalActualTime",
            f."arrivalPassengers",
            f."arrivalMalePassengers",
            f."arrivalFemalePassengers",
            f."arrivalChildren",
            f."arrivalInfants",
            f."arrivalBaggage",
            f."arrivalCargo",
            f."arrivalMail",
            f."arrivalStatus",
            f."arrivalFerryIn",

            -- Arrival Airport
            arr_ap."iataCode" as arrival_airport_iata,
            arr_ap."icaoCode" as arrival_airport_icao,
            arr_ap.name as arrival_airport_name,
            arr_ap.city as arrival_airport_city,
            arr_ap.country as arrival_airport_country,

            -- Departure
            f."departureFlightNumber",
            f."departureScheduledTime",
            f."departureActualTime",
            f."departureDoorClosingTime",
            f."departurePassengers",
            f."departureMalePassengers",
            f."departureFemalePassengers",
            f."departureChildren",
            f."departureInfants",
            f."departureBaggage",
            f."departureCargo",
            f."departureMail",
            f."departureStatus",
            f."departureFerryOut",

            -- Departure Airport
            dep_ap."iataCode" as departure_airport_iata,
            dep_ap."icaoCode" as departure_airport_icao,
            dep_ap.name as departure_airport_name,
            dep_ap.city as departure_airport_city,
            dep_ap.country as departure_airport_country

        FROM "Flight" f
        INNER JOIN "Airline" a ON f."airlineId" = a.id
        INNER JOIN "AircraftType" at ON f."aircraftTypeId" = at.id
        INNER JOIN "OperationType" ot ON f."operationTypeId" = ot.id
        LEFT JOIN "FlightType" ft ON f."flightTypeId" = ft.id
        LEFT JOIN "Airport" arr_ap ON f."arrivalAirportId" = arr_ap.id
        LEFT JOIN "Airport" dep_ap ON f."departureAirportId" = dep_ap.id
        WHERE f.date >= %s AND f.date <= %s
        ORDER BY f.date, f."arrivalScheduledTime", f."departureScheduledTime"
    """

    cursor.execute(query, (first_day, last_date))
    flights = cursor.fetchall()

    cursor.close()
    conn.close()

    return flights


def aggregate_airport_traffic(flights):
    """
    Agregirati podatke za Sheet 1 - AIRPORT TRAFFIC

    Vraća:
    {
        'international_scheduled': {'movements': X, 'embarked': Y, 'disembarked': Z, ...},
        'international_non_scheduled': {...},
        'domestic': {...},
        ...
    }
    """
    data = {
        'international_scheduled': {
            'movements': 0,
            'embarked': 0,
            'disembarked': 0,
            'freight_loaded': 0,
            'freight_unloaded': 0,
            'mail_loaded': 0,
            'mail_unloaded': 0,
        },
        'international_non_scheduled': {
            'movements': 0,
            'embarked': 0,
            'disembarked': 0,
            'freight_loaded': 0,
            'freight_unloaded': 0,
            'mail_loaded': 0,
            'mail_unloaded': 0,
        },
        'domestic': {
            'movements': 0,
            'embarked': 0,
            'disembarked': 0,
            'freight_loaded': 0,
            'freight_unloaded': 0,
            'mail_loaded': 0,
            'mail_unloaded': 0,
        },
    }

    # Debug: Ispis prvog flight-a da vidimo strukturu
    if len(flights) > 0:
        print(f"\n[DEBUG] Prvi flight podaci:")
        print(f"  arrivalFlightNumber: {flights[0].get('arrivalFlightNumber')}")
        print(f"  departureFlightNumber: {flights[0].get('departureFlightNumber')}")
        print(f"  arrivalFerryIn: {flights[0].get('arrivalFerryIn')}")
        print(f"  departureFerryOut: {flights[0].get('departureFerryOut')}")
        print(f"  operation_type_code: {flights[0].get('operation_type_code')}")
        print(f"  flight_type_code: {flights[0].get('flight_type_code')}")
        print()

    def is_scheduled_flight(flight):
        flight_type_code = (flight.get('flight_type_code') or '').upper()
        if flight_type_code:
            return flight_type_code == 'SCHEDULED'
        return (flight.get('operation_type_code') or '').upper() == 'SCHEDULED'

    for flight in flights:
        # Određivanje da li je international ili domestic
        # TZL je u BiH, pa ako drugi aerodrom nije u BiH onda je international
        is_domestic = False
        arr_country = flight.get('arrival_airport_country')
        dep_country = flight.get('departure_airport_country')

        # Ako oba aerodroma su u BiH ili jedan od njih je BiH (TZL)
        if (arr_country == 'Bosnia and Herzegovina' and dep_country == 'Bosnia and Herzegovina'):
            is_domestic = True

        # Određivanje scheduled vs non-scheduled
        is_scheduled = is_scheduled_flight(flight)

        # Odabir kategorije
        if is_domestic:
            category = 'domestic'
        elif is_scheduled:
            category = 'international_scheduled'
        else:
            category = 'international_non_scheduled'

        # Brojanje movementa - svaki arrival i svaki departure se broje odvojeno
        # ICAO standard: Aircraft Movements = broj arrival + broj departure
        # Pošto flight numbers često nisu postavljeni, brojimo po tome da li postoje passengers

        # Arrival movement - ako ima arrival putnika i nije ferry
        if flight.get('arrivalPassengers') is not None and flight.get('arrivalPassengers') > 0 and not flight.get('arrivalFerryIn'):
            data[category]['movements'] += 1
        # Ako nema putnika ali ima arrivalStatus = OPERATED, i dalje brojimo
        elif flight.get('arrivalStatus') == 'OPERATED' and not flight.get('arrivalFerryIn'):
            data[category]['movements'] += 1

        # Departure movement - ako ima departure putnika i nije ferry
        if flight.get('departurePassengers') is not None and flight.get('departurePassengers') > 0 and not flight.get('departureFerryOut'):
            data[category]['movements'] += 1
        # Ako nema putnika ali ima departureStatus = OPERATED, i dalje brojimo
        elif flight.get('departureStatus') == 'OPERATED' and not flight.get('departureFerryOut'):
            data[category]['movements'] += 1

        # Passengers
        if flight.get('departurePassengers') is not None:
            data[category]['embarked'] += flight['departurePassengers']
        if flight.get('arrivalPassengers') is not None:
            data[category]['disembarked'] += flight['arrivalPassengers']

        # Freight (konvertovati kg u tonne)
        if flight.get('departureCargo') is not None:
            data[category]['freight_loaded'] += flight['departureCargo'] / 1000.0
        if flight.get('arrivalCargo') is not None:
            data[category]['freight_unloaded'] += flight['arrivalCargo'] / 1000.0

        # Mail (konvertovati kg u tonne)
        if flight.get('departureMail') is not None:
            data[category]['mail_loaded'] += flight['departureMail'] / 1000.0
        if flight.get('arrivalMail') is not None:
            data[category]['mail_unloaded'] += flight['arrivalMail'] / 1000.0

    # Zaokružiti freight i mail na 2 decimale
    for category in data:
        data[category]['freight_loaded'] = round(data[category]['freight_loaded'], 2)
        data[category]['freight_unloaded'] = round(data[category]['freight_unloaded'], 2)
        data[category]['mail_loaded'] = round(data[category]['mail_loaded'], 2)
        data[category]['mail_unloaded'] = round(data[category]['mail_unloaded'], 2)

    return data


def parse_route(route_str):
    """
    Parse route string to extract departure and arrival airport codes.
    Route formats:
    - Simple: "TZL-MLH" or "MLH-TZL"
    - Round-trip: "TZL-DTM-TZL" or "DTM-TZL-DTM"

    For round-trip format, returns the first leg (outbound).

    Returns: (departure_iata, arrival_iata) or (None, None) if cannot parse
    """
    if not route_str or '-' not in route_str:
        return None, None

    parts = [p.strip() for p in route_str.strip().split('-') if p.strip()]

    if len(parts) == 2:
        # Simple one-way: "TZL-MLH"
        return parts[0], parts[1]
    elif len(parts) == 3:
        # Round-trip: "TZL-DTM-TZL" or "DTM-TZL-DTM"
        # Return first leg (outbound): parts[0] -> parts[1]
        return parts[0], parts[1]
    elif len(parts) > 3:
        # Multi-leg route - use first and last
        return parts[0], parts[-1]

    return None, None


def split_city_pairs(city_pairs, base_iata='TZL'):
    outbound = []
    inbound = []

    for pair_key, data in city_pairs.items():
        from_airport, to_airport = pair_key.split('-')
        if from_airport == base_iata and to_airport != base_iata:
            outbound.append((from_airport, to_airport, data))
        elif to_airport == base_iata and from_airport != base_iata:
            inbound.append((from_airport, to_airport, data))

    return outbound, inbound


def find_row_by_label(ws, label, column=4):
    for row in range(1, ws.max_row + 1):
        value = ws.cell(row=row, column=column).value
        if value and label in str(value):
            return row
    return None


def write_city_pairs_sheet(ws, city_pairs, start_row=23):
    """
    Write city-pair data to worksheet, safely handling merged cells.
    """
    outbound, inbound = split_city_pairs(city_pairs)
    remarks_row = find_row_by_label(ws, 'Remarks', column=4)
    if remarks_row is None:
        remarks_row = ws.max_row + 1

    rows_needed = len(outbound) + len(inbound)
    available_rows = max(0, remarks_row - start_row)

    if rows_needed > available_rows:
        ws.insert_rows(remarks_row, rows_needed - available_rows)
        remarks_row += rows_needed - available_rows

    end_row = remarks_row - 1

    # Safely clear merged cells in the data range
    # Columns: 4 (From), 5 (To), 9 (Passengers), 14 (Freight), 15 (Mail)
    clear_merged_cell_range(ws, start_row, end_row, [4, 5, 9, 14, 15])

    # Write new data
    row_num = start_row
    for from_airport, to_airport, data in outbound + inbound:
        ws.cell(row=row_num, column=4).value = sanitize_text(from_airport)
        ws.cell(row=row_num, column=5).value = sanitize_text(to_airport)
        ws.cell(row=row_num, column=9).value = data['passengers']
        ws.cell(row=row_num, column=14).value = data['freight']
        ws.cell(row=row_num, column=15).value = data['mail']
        row_num += 1


def aggregate_city_pair_data(flights, scheduled_only=True):
    """
    Agregirati city-pair podatke za Sheet 2/3

    VAŽNO: Za round-trip letove (npr. "TZL-DTM-TZL"), departure i arrival se tretiraju odvojeno:
    - Departure leg: TZL -> DTM (koristi departurePassengers)
    - Arrival leg: DTM -> TZL (koristi arrivalPassengers)

    Returns:
    {
        'TZL-MLH': {'passengers': 3529, 'freight': 0, 'mail': 0},
        'MLH-TZL': {'passengers': 3084, 'freight': 0, 'mail': 0},
        ...
    }
    """
    city_pairs = {}

    # Debug: Prvi flight podaci
    if len(flights) > 0:
        print(f"\n[DEBUG City-Pair] Prvi flight:")
        print(f"  route: {flights[0].get('route')}")
        print(f"  departure_airport_iata (from DB): {flights[0].get('departure_airport_iata')}")
        print(f"  arrival_airport_iata (from DB): {flights[0].get('arrival_airport_iata')}")

        # Test route parsing
        if flights[0].get('route'):
            parsed_dep, parsed_arr = parse_route(flights[0].get('route'))
            print(f"  Parsed from route: {parsed_dep} -> {parsed_arr}")

        print(f"  departurePassengers: {flights[0].get('departurePassengers')}")
        print(f"  arrivalPassengers: {flights[0].get('arrivalPassengers')}")
        print(f"  operation_type_code: {flights[0].get('operation_type_code')}")
        print(f"  flight_type_code: {flights[0].get('flight_type_code')}")
        print()

    def is_scheduled_flight(flight):
        flight_type_code = (flight.get('flight_type_code') or '').upper()
        if flight_type_code:
            return flight_type_code == 'SCHEDULED'
        return (flight.get('operation_type_code') or '').upper() == 'SCHEDULED'

    TZL_IATA = 'TZL'

    for flight in flights:
        is_scheduled = is_scheduled_flight(flight)

        # Filtrirati po scheduled/non-scheduled
        if scheduled_only and not is_scheduled:
            continue
        if not scheduled_only and is_scheduled:
            continue

        # Parse route to get airport codes
        route = flight.get('route')
        if not route:
            continue

        dep_iata, arr_iata = parse_route(route)
        if not dep_iata or not arr_iata:
            continue

        # Departure leg: TZL -> Other
        if flight.get('departurePassengers') is not None and flight.get('departurePassengers') > 0 and not flight.get('departureFerryOut'):
            # Determine other airport based on route
            # For round-trip "TZL-DTM-TZL", departure is TZL -> DTM
            other_airport = arr_iata if dep_iata == TZL_IATA else dep_iata

            if other_airport and other_airport != TZL_IATA:
                pair_key = f"{TZL_IATA}-{other_airport}"
                if pair_key not in city_pairs:
                    city_pairs[pair_key] = {'passengers': 0, 'freight': 0, 'mail': 0}

                city_pairs[pair_key]['passengers'] += flight.get('departurePassengers', 0) or 0
                city_pairs[pair_key]['freight'] += (flight.get('departureCargo', 0) or 0) / 1000.0
                city_pairs[pair_key]['mail'] += (flight.get('departureMail', 0) or 0) / 1000.0

        # Arrival leg: Other -> TZL
        if flight.get('arrivalPassengers') is not None and flight.get('arrivalPassengers') > 0 and not flight.get('arrivalFerryIn'):
            # Determine other airport based on route
            # For round-trip "TZL-DTM-TZL", arrival is DTM -> TZL
            # For simple route "DTM-TZL", arrival is DTM -> TZL
            other_airport = dep_iata if arr_iata == TZL_IATA else arr_iata

            if other_airport and other_airport != TZL_IATA:
                pair_key = f"{other_airport}-{TZL_IATA}"
                if pair_key not in city_pairs:
                    city_pairs[pair_key] = {'passengers': 0, 'freight': 0, 'mail': 0}

                city_pairs[pair_key]['passengers'] += flight.get('arrivalPassengers', 0) or 0
                city_pairs[pair_key]['freight'] += (flight.get('arrivalCargo', 0) or 0) / 1000.0
                city_pairs[pair_key]['mail'] += (flight.get('arrivalMail', 0) or 0) / 1000.0

    # Zaokružiti freight i mail na 2 decimale
    for pair in city_pairs:
        city_pairs[pair]['freight'] = round(city_pairs[pair]['freight'], 2)
        city_pairs[pair]['mail'] = round(city_pairs[pair]['mail'], 2)

    # Sortirati po broju putnika (descending)
    city_pairs = dict(sorted(city_pairs.items(), key=lambda x: x[1]['passengers'], reverse=True))

    # Debug output
    print(f"\n[DEBUG] Pronađeno {len(city_pairs)} city-pairs:")
    for pair_key, data in list(city_pairs.items())[:5]:  # Prikaži prvih 5
        print(f"  {pair_key}: {data['passengers']} putnika")

    return city_pairs


def generate_bhdca_report(year: int, month: int, output_path: Path = None):
    """
    Glavni metod za generisanje BHDCA izvještaja
    """
    print(f"Generišem BHDCA izvještaj za {MONTH_NAMES[month]} {year}...")

    # 1. Povući podatke iz baze
    print("Povlačim podatke iz baze...")
    flights = get_flight_data(year, month)
    print(f"Pronađeno {len(flights)} letova.")

    if len(flights) == 0:
        print("UPOZORENJE: Nema letova za zadati period!")
        return None

    # 2. Agregirati podatke
    print("Agregatiranje podataka...")
    airport_traffic = aggregate_airport_traffic(flights)
    city_pairs_scheduled = aggregate_city_pair_data(flights, scheduled_only=True)
    city_pairs_non_scheduled = aggregate_city_pair_data(flights, scheduled_only=False)

    # Debug output
    print("\n=== AGREGIRANI PODACI ===")
    print(f"International Scheduled: {airport_traffic['international_scheduled']}")
    print(f"International Non-Scheduled: {airport_traffic['international_non_scheduled']}")
    print(f"Domestic: {airport_traffic['domestic']}")
    print(f"\nCity-pairs Scheduled: {len(city_pairs_scheduled)} parova")
    print(f"City-pairs Non-Scheduled: {len(city_pairs_non_scheduled)} parova")
    print("========================\n")

    # 3. Učitati template
    print(f"Učitavanje template-a: {TEMPLATE_PATH}")
    if not TEMPLATE_PATH.exists():
        raise FileNotFoundError(f"Template ne postoji: {TEMPLATE_PATH}")

    wb = openpyxl.load_workbook(TEMPLATE_PATH)

    # 4. Popuniti Sheet 1 - AIRPORT TRAFFIC
    print("Popunjavam Sheet 1 - AIRPORT TRAFFIC...")
    ws1 = wb['Sheet1']

    # Header informacije
    ws1['D10'] = year
    ws1['N11'] = MONTH_NAMES[month]

    # Podaci
    # Red 20: International scheduled
    print(f"  Int. Scheduled - Movements: {airport_traffic['international_scheduled']['movements']}, Embarked: {airport_traffic['international_scheduled']['embarked']}, Disembarked: {airport_traffic['international_scheduled']['disembarked']}")
    ws1['F20'] = airport_traffic['international_scheduled']['movements']
    ws1['G20'] = airport_traffic['international_scheduled']['embarked']
    ws1['H20'] = airport_traffic['international_scheduled']['disembarked']
    ws1['K20'] = airport_traffic['international_scheduled']['freight_loaded']
    ws1['L20'] = airport_traffic['international_scheduled']['freight_unloaded']
    ws1['N20'] = airport_traffic['international_scheduled']['mail_loaded']
    ws1['O20'] = airport_traffic['international_scheduled']['mail_unloaded']

    # Red 21: International non-scheduled
    ws1['F21'] = airport_traffic['international_non_scheduled']['movements']
    ws1['G21'] = airport_traffic['international_non_scheduled']['embarked']
    ws1['H21'] = airport_traffic['international_non_scheduled']['disembarked']
    ws1['K21'] = airport_traffic['international_non_scheduled']['freight_loaded']
    ws1['L21'] = airport_traffic['international_non_scheduled']['freight_unloaded']
    ws1['N21'] = airport_traffic['international_non_scheduled']['mail_loaded']
    ws1['O21'] = airport_traffic['international_non_scheduled']['mail_unloaded']

    # Red 22: Total international (umjesto formule, eksplicitna vrijednost)
    total_int_movements = airport_traffic['international_scheduled']['movements'] + airport_traffic['international_non_scheduled']['movements']
    total_int_embarked = airport_traffic['international_scheduled']['embarked'] + airport_traffic['international_non_scheduled']['embarked']
    total_int_disembarked = airport_traffic['international_scheduled']['disembarked'] + airport_traffic['international_non_scheduled']['disembarked']
    total_int_freight_loaded = airport_traffic['international_scheduled']['freight_loaded'] + airport_traffic['international_non_scheduled']['freight_loaded']
    total_int_freight_unloaded = airport_traffic['international_scheduled']['freight_unloaded'] + airport_traffic['international_non_scheduled']['freight_unloaded']
    total_int_mail_loaded = airport_traffic['international_scheduled']['mail_loaded'] + airport_traffic['international_non_scheduled']['mail_loaded']
    total_int_mail_unloaded = airport_traffic['international_scheduled']['mail_unloaded'] + airport_traffic['international_non_scheduled']['mail_unloaded']

    ws1['F22'] = total_int_movements
    ws1['G22'] = total_int_embarked
    ws1['H22'] = total_int_disembarked
    ws1['K22'] = total_int_freight_loaded
    ws1['L22'] = total_int_freight_unloaded
    ws1['N22'] = total_int_mail_loaded
    ws1['O22'] = total_int_mail_unloaded

    # Red 23: Domestic
    ws1['F23'] = airport_traffic['domestic']['movements']
    ws1['G23'] = airport_traffic['domestic']['embarked']
    ws1['H23'] = airport_traffic['domestic']['disembarked']
    ws1['K23'] = airport_traffic['domestic']['freight_loaded']
    ws1['L23'] = airport_traffic['domestic']['freight_unloaded']
    ws1['N23'] = airport_traffic['domestic']['mail_loaded']
    ws1['O23'] = airport_traffic['domestic']['mail_unloaded']

    # Red 24: Total commercial air transport (umjesto formule, eksplicitna vrijednost)
    total_comm_movements = total_int_movements + airport_traffic['domestic']['movements']
    total_comm_embarked = total_int_embarked + airport_traffic['domestic']['embarked']
    total_comm_disembarked = total_int_disembarked + airport_traffic['domestic']['disembarked']
    total_comm_freight_loaded = total_int_freight_loaded + airport_traffic['domestic']['freight_loaded']
    total_comm_freight_unloaded = total_int_freight_unloaded + airport_traffic['domestic']['freight_unloaded']
    total_comm_mail_loaded = total_int_mail_loaded + airport_traffic['domestic']['mail_loaded']
    total_comm_mail_unloaded = total_int_mail_unloaded + airport_traffic['domestic']['mail_unloaded']

    ws1['F24'] = total_comm_movements
    ws1['G24'] = total_comm_embarked
    ws1['H24'] = total_comm_disembarked
    ws1['K24'] = total_comm_freight_loaded
    ws1['L24'] = total_comm_freight_unloaded
    ws1['N24'] = total_comm_mail_loaded
    ws1['O24'] = total_comm_mail_unloaded

    # 5. Popuniti Sheet 2 - O-D TRAFFIC (Scheduled)
    print("Popunjavam Sheet 2 - O-D TRAFFIC (Scheduled)...")
    ws2 = wb['Sheet2']

    # Header
    ws2['N9'] = MONTH_NAMES[month]
    ws2['N10'] = year

    # Očistiti stare podatke (redovi 23-35) - skip merged cells
    write_city_pairs_sheet(ws2, city_pairs_scheduled, start_row=23)

    # 6. Popuniti Sheet 3 - O-D TRAFFIC (Non-Scheduled)
    print("Popunjavam Sheet 3 - O-D TRAFFIC (Non-Scheduled)...")
    ws3 = wb['Sheet3']

    # Header
    ws3['N9'] = MONTH_NAMES[month]
    ws3['N10'] = year

    # Očistiti stare podatke (redovi 23-35) - skip merged cells
    write_city_pairs_sheet(ws3, city_pairs_non_scheduled, start_row=23)

    # 7. Sačuvati Excel fajl
    if output_path is None:
        month_name = MONTH_NAMES[month]
        output_filename = f"BHDCA_{month_name}_{year}.xlsx"
        output_path = OUTPUT_DIR / output_filename

    print(f"Čuvam izvještaj u: {output_path}")
    wb.iso_dates = True
    wb.save(output_path)
    wb.close()

    print(f"✅ BHDCA izvještaj uspješno generisan: {output_path}")

    return output_path


def main():
    """CLI interfejs za skriptu"""
    if len(sys.argv) < 3:
        print("Usage: python generate_bhdca_report.py <year> <month>")
        print("Example: python generate_bhdca_report.py 2025 9")
        sys.exit(1)

    try:
        year = int(sys.argv[1])
        month = int(sys.argv[2])

        if month < 1 or month > 12:
            print("Mjesec mora biti između 1 i 12!")
            sys.exit(1)

        output_path = generate_bhdca_report(year, month)

        if output_path:
            print(f"\n✅ Izvještaj generisan: {output_path}")
        else:
            print("\n❌ Izvještaj nije generisan zbog greške.")
            sys.exit(1)

    except Exception as e:
        print(f"\n❌ Greška: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == "__main__":
    main()
