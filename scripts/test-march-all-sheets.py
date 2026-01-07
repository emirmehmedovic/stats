#!/usr/bin/env python3

import openpyxl
import os
import re

STATS_DIR = "/Users/emir_mw/stats/STATS/2025/Dnevni izvještaji"
month_folder = "03. MART"
month_path = os.path.join(STATS_DIR, month_folder)

files = [
    f for f in os.listdir(month_path)
    if 'Dnevni izvještaj o saobraćaju' in f and f.endswith('.xlsx')
]

file_path = os.path.join(month_path, files[0])
wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)

print(f"📊 Processing all March sheets\n")

total_flights = 0

for sheet_name in wb.sheetnames:
    # Extract day from sheet name
    day_str = re.match(r'^(\d+)', sheet_name.strip())
    if not day_str:
        print(f"  ⏭️  Skipping '{sheet_name}' (not a day)")
        continue

    day = day_str.group(1).zfill(2)
    ws = wb[sheet_name]

    # Count data rows
    row_count = 0
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if i == 1:  # Skip header
            continue
        if row[0]:  # Has date
            row_count += 1

    total_flights += row_count
    print(f"  Day {day} (sheet '{sheet_name}'): {row_count} flights")

print(f"\n✅ Total flights in March: {total_flights}")

wb.close()
