#!/usr/bin/env python3
"""
Custom Advanced Multi-Sheet Report Generator

Generiše napredni custom izvještaj sa više sheet-ova:
- Sheet 1: Super detaljni SUMMARY sa monthly comparison
- Sheet 2+: Mjesečni sheet-ovi sa monthly summary + daily breakdown
"""

import sys
import os
import json
import calendar
from datetime import datetime, timedelta
from pathlib import Path
from collections import defaultdict
import psycopg2
from psycopg2.extras import RealDictCursor
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

# Putanja
SCRIPT_DIR = Path(__file__).parent
PROJECT_ROOT = SCRIPT_DIR.parent
OUTPUT_DIR = PROJECT_ROOT / "izvještaji" / "generated"

# Nazivi mjeseci
MONTH_NAMES = {
    1: "Januar", 2: "Februar", 3: "Mart", 4: "April",
    5: "Maj", 6: "Juni", 7: "Juli", 8: "Avgust",
    9: "Septembar", 10: "Oktobar", 11: "Novembar", 12: "Decembar"
}

MONTH_NAMES_SHORT = {
    1: "Jan", 2: "Feb", 3: "Mar", 4: "Apr",
    5: "Maj", 6: "Jun", 7: "Jul", 8: "Avg",
    9: "Sep", 10: "Okt", 11: "Nov", 12: "Dec"
}


def get_db_connection():
    """Konekcija na PostgreSQL bazu"""
    DATABASE_URL = os.getenv("DATABASE_URL")
    if not DATABASE_URL:
        raise ValueError("DATABASE_URL environment variable not set")
    return psycopg2.connect(DATABASE_URL, cursor_factory=RealDictCursor)


def create_merged_cell(ws, row, start_col, end_col, value, font=None, fill=None, alignment=None, border=None):
    """
    Helper function to properly create merged cells with consistent formatting.
    Applies formatting to ALL cells in the range before merging to avoid Excel corruption.
    """
    for col in range(start_col, end_col + 1):
        cell = ws.cell(row=row, column=col)
        if col == start_col and value is not None:
            cell.value = value
        if font:
            cell.font = font
        if fill:
            cell.fill = fill
        if alignment:
            cell.alignment = alignment
        if border:
            cell.border = border
    
    if start_col != end_col:
        ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=end_col)


def format_date(dt):
    """Format date za prikaz"""
    if not dt:
        return "-"
    if isinstance(dt, str):
        dt = datetime.fromisoformat(dt.split('T')[0])
    return dt.strftime("%d.%m.%Y")


def format_datetime(dt):
    """Format datetime za prikaz"""
    if not dt:
        return "-"
    if isinstance(dt, str):
        dt = datetime.fromisoformat(dt.replace('Z', '+00:00'))
    return dt.strftime("%d.%m.%Y %H:%M")


def get_flight_data(date_from, date_to, operation_types=None, airlines=None, routes=None):
    """Povuči letove iz baze sa filterima"""
    conn = get_db_connection()
    cursor = conn.cursor()

    query = """
        SELECT
            f.id,
            f.date,
            f.route,
            f.registration,
            
            a.name as airline_name,
            at.model as aircraft_model,
            ot.name as operation_type_name,
            
            f."departureFlightNumber",
            f."departureScheduledTime",
            f."departureActualTime",
            f."departurePassengers",
            f."departureMalePassengers",
            f."departureFemalePassengers",
            f."departureChildren",
            f."departureInfants",
            
            f."arrivalFlightNumber",
            f."arrivalScheduledTime",
            f."arrivalActualTime",
            f."arrivalPassengers",
            f."arrivalMalePassengers",
            f."arrivalFemalePassengers",
            f."arrivalChildren",
            f."arrivalInfants"
            
        FROM "Flight" f
        LEFT JOIN "Airline" a ON f."airlineId" = a.id
        LEFT JOIN "AircraftType" at ON f."aircraftTypeId" = at.id
        LEFT JOIN "OperationType" ot ON f."operationTypeId" = ot.id
        WHERE f.date >= %s AND f.date <= %s
    """
    
    params = [date_from, date_to]
    
    if operation_types and len(operation_types) > 0:
        query += " AND f.\"operationTypeId\" = ANY(%s)"
        params.append(operation_types)
    
    if airlines and len(airlines) > 0:
        query += " AND f.\"airlineId\" = ANY(%s)"
        params.append(airlines)
    
    if routes and len(routes) > 0:
        query += " AND f.route = ANY(%s)"
        params.append(routes)
    
    query += " ORDER BY f.date, f.\"departureScheduledTime\", f.\"arrivalScheduledTime\""
    
    cursor.execute(query, params)
    flights = cursor.fetchall()
    
    cursor.close()
    conn.close()
    
    return flights


def aggregate_by_month(flights):
    """Grupiši letove po mjesecima"""
    monthly_data = defaultdict(list)
    
    for flight in flights:
        date = flight['date']
        if isinstance(date, str):
            date = datetime.fromisoformat(date.split('T')[0])
        
        month_key = (date.year, date.month)
        monthly_data[month_key].append(flight)
    
    return dict(monthly_data)


def calculate_monthly_stats(flights, passenger_type):
    """Izračunaj statistiku za mjesec"""
    total_flights = len(flights)
    total_operations = 0
    total_passengers = 0
    total_departure_passengers = 0
    total_arrival_passengers = 0
    
    for flight in flights:
        if flight['arrivalFlightNumber']:
            total_operations += 1
        if flight['departureFlightNumber']:
            total_operations += 1
        
        dep_pax = flight['departurePassengers'] or 0
        arr_pax = flight['arrivalPassengers'] or 0
        
        total_departure_passengers += dep_pax
        total_arrival_passengers += arr_pax
        
        if passenger_type == "departure":
            total_passengers += dep_pax
        elif passenger_type == "arrival":
            total_passengers += arr_pax
        elif passenger_type == "all":
            total_passengers += dep_pax + arr_pax
    
    return {
        'flights': total_flights,
        'operations': total_operations,
        'passengers': total_passengers,
        'departure_passengers': total_departure_passengers,
        'arrival_passengers': total_arrival_passengers,
        'avg_per_flight': round(total_passengers / total_flights, 1) if total_flights > 0 else 0,
        'avg_per_operation': round(total_passengers / total_operations, 1) if total_operations > 0 else 0
    }


def create_summary_sheet(wb, all_flights, monthly_data, date_from, date_to, passenger_type):
    """Kreira super detaljni summary sheet"""
    ws = wb.active
    ws.title = "SUMMARY"
    
    # Styles
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    row_idx = 1
    
    # Main header
    summary_header_fill = PatternFill(start_color="F59E0B", end_color="F59E0B", fill_type="solid")
    create_merged_cell(
        ws, row_idx, 1, 8,
        "SAŽETAK IZVJEŠTAJA",
        font=Font(bold=True, size=16, color="FFFFFF"),
        fill=summary_header_fill,
        alignment=Alignment(horizontal="center", vertical="center"),
        border=thin_border
    )
    row_idx += 1
    
    # Period info
    period_cell = ws.cell(row=row_idx, column=1, value="Period:")
    period_cell.font = Font(bold=True, size=12)
    period_cell.border = thin_border
    
    create_merged_cell(
        ws, row_idx, 2, 4,
        f"{date_from} do {date_to}",
        font=Font(size=12),
        border=thin_border
    )
    row_idx += 2
    
    # Overall statistics
    stats_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
    
    create_merged_cell(
        ws, row_idx, 1, 8,
        "OSNOVNE STATISTIKE",
        font=Font(bold=True, size=12, color="FFFFFF"),
        fill=PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid"),
        alignment=Alignment(horizontal="center"),
        border=thin_border
    )
    row_idx += 1
    
    # Calculate overall stats
    overall_stats = calculate_monthly_stats(all_flights, passenger_type)
    
    stats_data = [
        ("Ukupno letova:", overall_stats['flights']),
        ("Ukupno operacija:", overall_stats['operations']),
        ("Ukupno putnika:", overall_stats['passengers']),
        ("  - Odlazeći putnici:", overall_stats['departure_passengers']),
        ("  - Dolazeći putnici:", overall_stats['arrival_passengers']),
        ("Prosječno putnika po letu:", overall_stats['avg_per_flight']),
        ("Prosječno putnika po operaciji:", overall_stats['avg_per_operation']),
    ]
    
    for label, value in stats_data:
        label_cell = ws.cell(row=row_idx, column=1, value=label)
        label_cell.font = Font(bold=True if not label.startswith("  -") else False, italic=label.startswith("  -"))
        label_cell.fill = stats_fill
        label_cell.border = thin_border
        
        value_cell = ws.cell(row=row_idx, column=2, value=value)
        value_cell.font = Font(bold=True, size=11, color="D97706" if "putnika" in label.lower() else "000000")
        value_cell.fill = stats_fill
        value_cell.alignment = Alignment(horizontal="center")
        value_cell.border = thin_border
        row_idx += 1
    
    row_idx += 2
    
    # Monthly comparison table
    create_merged_cell(
        ws, row_idx, 1, 8,
        "MJESEČNA KOMPARACIJA",
        font=Font(bold=True, size=12, color="FFFFFF"),
        fill=PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid"),
        alignment=Alignment(horizontal="center"),
        border=thin_border
    )
    row_idx += 1
    
    # Headers for monthly comparison
    headers = ["Mjesec", "Letova", "Operacija", "Putnici", "Odlazeći", "Dolazeći", "Avg/Let", "Growth%"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="E5E7EB", end_color="E5E7EB", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border
    row_idx += 1
    
    # Monthly data rows
    sorted_months = sorted(monthly_data.keys())
    prev_passengers = None
    total_row_data = [0, 0, 0, 0, 0, 0]
    
    for year, month in sorted_months:
        month_flights = monthly_data[(year, month)]
        stats = calculate_monthly_stats(month_flights, passenger_type)
        
        # Calculate growth
        growth = None
        if prev_passengers is not None and prev_passengers > 0:
            growth = ((stats['passengers'] - prev_passengers) / prev_passengers) * 100
        prev_passengers = stats['passengers']
        
        # Month name
        month_name = f"{MONTH_NAMES_SHORT[month]} {year}"
        
        row_data = [
            month_name,
            stats['flights'],
            stats['operations'],
            stats['passengers'],
            stats['departure_passengers'],
            stats['arrival_passengers'],
            stats['avg_per_flight'],
            f"{growth:+.1f}%" if growth is not None else "-"
        ]
        
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            if col_idx > 1:
                cell.alignment = Alignment(horizontal="center")
        
        # Accumulate totals
        total_row_data[0] += stats['flights']
        total_row_data[1] += stats['operations']
        total_row_data[2] += stats['passengers']
        total_row_data[3] += stats['departure_passengers']
        total_row_data[4] += stats['arrival_passengers']
        
        row_idx += 1
    
    # Total row
    total_avg = round(total_row_data[2] / total_row_data[0], 1) if total_row_data[0] > 0 else 0
    total_row = ["UKUPNO", total_row_data[0], total_row_data[1], total_row_data[2], 
                 total_row_data[3], total_row_data[4], total_avg, "-"]
    
    for col_idx, value in enumerate(total_row, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
        cell.border = thin_border
        if col_idx > 1:
            cell.alignment = Alignment(horizontal="center")
    
    row_idx += 2
    
    # Airline breakdown for entire period
    row_idx = create_airline_breakdown(ws, row_idx, all_flights, passenger_type, thin_border, is_summary=True)
    row_idx += 2
    
    # Route breakdown for entire period
    row_idx = create_route_breakdown(ws, row_idx, all_flights, passenger_type, thin_border, is_summary=True)
    
    # Set column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 12


def create_airline_breakdown(ws, start_row, flights, passenger_type, thin_border, is_summary=False):
    """Kreira airline breakdown sekciju"""
    row_idx = start_row
    
    # Aggregate airline data
    airline_data = defaultdict(lambda: {'flights': 0, 'operations': 0, 'passengers': 0, 
                                        'departure': 0, 'arrival': 0})
    
    for flight in flights:
        airline = flight['airline_name'] or "Nepoznato"
        airline_data[airline]['flights'] += 1
        
        if flight['arrivalFlightNumber']:
            airline_data[airline]['operations'] += 1
        if flight['departureFlightNumber']:
            airline_data[airline]['operations'] += 1
        
        dep_pax = flight['departurePassengers'] or 0
        arr_pax = flight['arrivalPassengers'] or 0
        airline_data[airline]['departure'] += dep_pax
        airline_data[airline]['arrival'] += arr_pax
        
        if passenger_type == "departure":
            airline_data[airline]['passengers'] += dep_pax
        elif passenger_type == "arrival":
            airline_data[airline]['passengers'] += arr_pax
        elif passenger_type == "all":
            airline_data[airline]['passengers'] += dep_pax + arr_pax
    
    if airline_data:
        # Header
        breakdown_fill = PatternFill(start_color="64748B", end_color="64748B", fill_type="solid")
        header_text = "PREGLED PO AVIOKOMPANIJAMA (cijeli period)" if is_summary else "PREGLED PO AVIOKOMPANIJAMA"
        header_end_col = 8 if is_summary else 7
        
        create_merged_cell(
            ws, row_idx, 1, header_end_col,
            header_text,
            font=Font(bold=True, size=11, color="FFFFFF"),
            fill=breakdown_fill,
            alignment=Alignment(horizontal="center"),
            border=thin_border
        )
        row_idx += 1
        
        # Column headers
        headers = ["Aviokompanija", "Letova", "Operacija", "Odlazeći", "Dolazeći", "Ukupno", "Prosječno"]
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")
        row_idx += 1
        
        # Data rows - sorted by passenger count (highest first)
        sorted_airlines = sorted(airline_data.keys(), key=lambda x: airline_data[x]['passengers'], reverse=True)
        for airline in sorted_airlines:
            data = airline_data[airline]
            avg = round(data['passengers'] / data['operations'], 1) if data['operations'] > 0 else 0
            
            row_data = [airline, data['flights'], data['operations'], 
                       data['departure'], data['arrival'], data['passengers'], avg]
            
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                if col_idx > 1:
                    cell.alignment = Alignment(horizontal="center")
            row_idx += 1
    
    return row_idx


def create_route_breakdown(ws, start_row, flights, passenger_type, thin_border, is_summary=False):
    """Kreira route breakdown sekciju"""
    row_idx = start_row
    
    # Aggregate route data
    route_data = defaultdict(lambda: {'flights': 0, 'operations': 0, 'passengers': 0, 
                                      'departure': 0, 'arrival': 0})
    
    for flight in flights:
        route = flight['route'] or "Nepoznato"
        route_data[route]['flights'] += 1
        
        if flight['arrivalFlightNumber']:
            route_data[route]['operations'] += 1
        if flight['departureFlightNumber']:
            route_data[route]['operations'] += 1
        
        dep_pax = flight['departurePassengers'] or 0
        arr_pax = flight['arrivalPassengers'] or 0
        route_data[route]['departure'] += dep_pax
        route_data[route]['arrival'] += arr_pax
        
        if passenger_type == "departure":
            route_data[route]['passengers'] += dep_pax
        elif passenger_type == "arrival":
            route_data[route]['passengers'] += arr_pax
        elif passenger_type == "all":
            route_data[route]['passengers'] += dep_pax + arr_pax
    
    if route_data:
        # Header
        breakdown_fill = PatternFill(start_color="64748B", end_color="64748B", fill_type="solid")
        header_text = "PREGLED PO RUTAMA (cijeli period)" if is_summary else "PREGLED PO RUTAMA"
        header_end_col = 8 if is_summary else 7
        
        create_merged_cell(
            ws, row_idx, 1, header_end_col,
            header_text,
            font=Font(bold=True, size=11, color="FFFFFF"),
            fill=breakdown_fill,
            alignment=Alignment(horizontal="center"),
            border=thin_border
        )
        row_idx += 1
        
        # Column headers
        headers = ["Ruta", "Letova", "Operacija", "Odlazeći", "Dolazeći", "Ukupno", "Prosječno"]
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")
        row_idx += 1
        
        # Data rows - sorted by flight count
        for route in sorted(route_data.keys(), key=lambda x: route_data[x]['flights'], reverse=True):
            data = route_data[route]
            avg = round(data['passengers'] / data['operations'], 1) if data['operations'] > 0 else 0
            
            row_data = [route, data['flights'], data['operations'], 
                       data['departure'], data['arrival'], data['passengers'], avg]
            
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                if col_idx > 1:
                    cell.alignment = Alignment(horizontal="center")
            row_idx += 1
    
    return row_idx


def create_monthly_sheet(wb, month_flights, year, month, passenger_type):
    """Kreira sheet za pojedinačni mjesec"""
    month_name = f"{MONTH_NAMES[month].upper()}_{year}"
    ws = wb.create_sheet(title=month_name)
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    row_idx = 1
    
    # Month header
    create_merged_cell(
        ws, row_idx, 1, 8,
        f"{MONTH_NAMES[month]} {year}",
        font=Font(bold=True, size=14, color="FFFFFF"),
        fill=PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid"),
        alignment=Alignment(horizontal="center"),
        border=thin_border
    )
    row_idx += 2
    
    # Monthly summary
    stats = calculate_monthly_stats(month_flights, passenger_type)
    stats_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
    
    create_merged_cell(
        ws, row_idx, 1, 4,
        "MJESEČNI SAŽETAK",
        font=Font(bold=True, size=11),
        fill=PatternFill(start_color="E5E7EB", end_color="E5E7EB", fill_type="solid"),
        alignment=Alignment(horizontal="center"),
        border=thin_border
    )
    row_idx += 1
    
    summary_data = [
        ("Ukupno letova:", stats['flights']),
        ("Ukupno operacija:", stats['operations']),
        ("Ukupno putnika:", stats['passengers']),
        ("Prosječno po letu:", stats['avg_per_flight']),
    ]
    
    for label, value in summary_data:
        label_cell = ws.cell(row=row_idx, column=1, value=label)
        label_cell.font = Font(bold=True)
        label_cell.fill = stats_fill
        label_cell.border = thin_border
        
        value_cell = ws.cell(row=row_idx, column=2, value=value)
        value_cell.font = Font(bold=True)
        value_cell.fill = stats_fill
        value_cell.alignment = Alignment(horizontal="center")
        value_cell.border = thin_border
        row_idx += 1
    
    row_idx += 1
    
    # Airline breakdown for this month
    row_idx = create_airline_breakdown(ws, row_idx, month_flights, passenger_type, thin_border, is_summary=False)
    row_idx += 2
    
    # Route breakdown for this month
    row_idx = create_route_breakdown(ws, row_idx, month_flights, passenger_type, thin_border, is_summary=False)
    row_idx += 2
    
    # Daily breakdown header
    create_merged_cell(
        ws, row_idx, 1, 11,
        "DNEVNI PREGLED",
        font=Font(bold=True, size=11, color="FFFFFF"),
        fill=PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid"),
        alignment=Alignment(horizontal="center"),
        border=thin_border
    )
    row_idx += 1
    
    # Daily data headers
    base_headers = ["Datum", "Aviokompanija", "Tip aviona", "Registracija", "Ruta", "Tip operacije"]
    
    if passenger_type == "all":
        headers = base_headers + [
            "Broj leta (dolazak)", "Putnici (dolazak)",
            "Broj leta (odlazak)", "Putnici (odlazak)",
            "Ukupno putnika"
        ]
    elif passenger_type == "departure":
        headers = base_headers + [
            "Broj leta (odlazak)", "Planirano vrijeme", "Stvarno vrijeme",
            "Putnici", "Muškarci", "Žene", "Djeca", "Bebe"
        ]
    elif passenger_type == "arrival":
        headers = base_headers + [
            "Broj leta (dolazak)", "Planirano vrijeme", "Stvarno vrijeme",
            "Putnici", "Muškarci", "Žene", "Djeca", "Bebe"
        ]
    else:  # infants
        headers = base_headers + [
            "Broj leta (dolazak)", "Broj leta (odlazak)",
            "Bebe (dolazak)", "Bebe (odlazak)", "Ukupno beba"
        ]
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=header)
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
    row_idx += 1
    
    # Daily flight data
    for flight in month_flights:
        base_data = [
            format_date(flight['date']),
            flight['airline_name'] or "-",
            flight['aircraft_model'] or "-",
            flight['registration'] or "-",
            flight['route'] or "-",
            flight['operation_type_name'] or "-"
        ]
        
        if passenger_type == "departure":
            passenger_data = [
                flight['departureFlightNumber'] or "-",
                format_datetime(flight['departureScheduledTime']),
                format_datetime(flight['departureActualTime']),
                flight['departurePassengers'] or 0,
                flight['departureMalePassengers'] or 0,
                flight['departureFemalePassengers'] or 0,
                flight['departureChildren'] or 0,
                flight['departureInfants'] or 0
            ]
        elif passenger_type == "arrival":
            passenger_data = [
                flight['arrivalFlightNumber'] or "-",
                format_datetime(flight['arrivalScheduledTime']),
                format_datetime(flight['arrivalActualTime']),
                flight['arrivalPassengers'] or 0,
                flight['arrivalMalePassengers'] or 0,
                flight['arrivalFemalePassengers'] or 0,
                flight['arrivalChildren'] or 0,
                flight['arrivalInfants'] or 0
            ]
        elif passenger_type == "infants":
            arr_infants = flight['arrivalInfants'] or 0
            dep_infants = flight['departureInfants'] or 0
            passenger_data = [
                flight['arrivalFlightNumber'] or "-",
                flight['departureFlightNumber'] or "-",
                arr_infants,
                dep_infants,
                arr_infants + dep_infants
            ]
        else:  # all
            arr_pax = flight['arrivalPassengers'] or 0
            dep_pax = flight['departurePassengers'] or 0
            passenger_data = [
                flight['arrivalFlightNumber'] or "-",
                arr_pax,
                flight['departureFlightNumber'] or "-",
                dep_pax,
                arr_pax + dep_pax
            ]
        
        row_data = base_data + passenger_data
        
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            if col_idx > 6:
                cell.alignment = Alignment(horizontal="center")
        
        row_idx += 1
    
    # Set column widths
    column_widths = [12, 25, 15, 12, 15, 15] + [18] * (len(headers) - 6)
    for col_idx, width in enumerate(column_widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width


def create_advanced_report(date_from, date_to, passenger_type, operation_types=None, airlines=None, routes=None):
    """Glavna funkcija za kreiranje advanced multi-sheet izvještaja"""
    
    # Fetch all flights (use stderr for debug messages to not break JSON parsing)
    sys.stderr.write("Povlačim podatke iz baze...\n")
    all_flights = get_flight_data(date_from, date_to, operation_types, airlines, routes)
    sys.stderr.write(f"Pronađeno {len(all_flights)} letova.\n")
    
    if len(all_flights) == 0:
        return None
    
    # Aggregate by month
    sys.stderr.write("Agregatiranje po mjesecima...\n")
    monthly_data = aggregate_by_month(all_flights)
    
    # Create workbook
    sys.stderr.write("Kreiram Excel workbook...\n")
    wb = openpyxl.Workbook()
    
    # Create summary sheet
    sys.stderr.write("Kreiram summary sheet...\n")
    create_summary_sheet(wb, all_flights, monthly_data, date_from, date_to, passenger_type)
    
    # Create monthly sheets
    sys.stderr.write("Kreiram mjesečne sheet-ove...\n")
    sorted_months = sorted(monthly_data.keys())
    for year, month in sorted_months:
        month_flights = monthly_data[(year, month)]
        create_monthly_sheet(wb, month_flights, year, month, passenger_type)
        sys.stderr.write(f"  - {MONTH_NAMES[month]} {year}: {len(month_flights)} letova\n")
    
    return wb


def main():
    """Main function"""
    if len(sys.argv) < 2:
        print(json.dumps({"error": "Missing filters argument"}))
        sys.exit(1)
    
    try:
        filters = json.loads(sys.argv[1])
        
        date_from = filters.get('dateFrom')
        date_to = filters.get('dateTo')
        passenger_type = filters.get('passengerType', 'all')
        operation_types = filters.get('operationTypes', [])
        airlines = filters.get('airlines', [])
        routes = filters.get('routes', [])
        
        if not date_from or not date_to:
            print(json.dumps({"error": "Missing dateFrom or dateTo"}))
            sys.exit(1)
        
        # Generate report
        wb = create_advanced_report(
            date_from, date_to, passenger_type,
            operation_types if operation_types else None,
            airlines if airlines else None,
            routes if routes else None
        )
        
        if wb is None:
            print(json.dumps({"error": "No flights found for the specified period"}))
            sys.exit(1)
        
        # Save file
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"Custom_Advanced_{date_from.replace('-', '')}_{date_to.replace('-', '')}_{timestamp}.xlsx"
        filepath = OUTPUT_DIR / filename
        
        OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
        wb.save(filepath)
        
        print(json.dumps({
            "success": True,
            "fileName": filename,
            "message": f"Izvještaj uspješno generisan: {filename}"
        }))
        
    except Exception as e:
        print(json.dumps({"error": str(e)}))
        sys.exit(1)


if __name__ == "__main__":
    main()
