#!/usr/bin/env python3

import openpyxl
import os
import sys
import json
import re
from datetime import datetime

STATS_DIR = "/Users/emir_mw/stats/STATS/2025/Dnevni izvještaji"

def parse_passengers(value):
    """Parse passenger string like '165+6 INF' or '110' or '-'"""
    if not value or value == '-' or value == 'N/A' or str(value).strip() == '':
        return None

    str_val = str(value).strip()

    # Pattern: "110+6 INF" or "110+6" (with infants)
    match = re.match(r'^(\d+)\s*\+\s*(\d+)', str_val)
    if match:
        adults = int(match.group(1))
        infants = int(match.group(2))
        return {
            'adults': adults,
            'infants': infants,
            'total': adults + infants
        }

    # Pattern: just a number "110" (no infants)
    match = re.match(r'^(\d+)$', str_val)
    if match:
        adults = int(match.group(1))
        return {
            'adults': adults,
            'infants': 0,
            'total': adults
        }

    return None

def parse_route(route):
    """Parse route like 'TZL-AYT' and return [departure, arrival]"""
    if not route or route == '-' or route == 'N/A':
        return None

    parts = [p.strip() for p in route.split('-') if p.strip()]

    if len(parts) >= 2:
        return {
            'departure': parts[0],
            'arrival': parts[-1],
        }

    return None

def extract_month(month_folder):
    """Extract flight data from one month"""
    month_path = os.path.join(STATS_DIR, month_folder)

    # Find Excel file
    files = [
        f for f in os.listdir(month_path)
        if 'Dnevni izvještaj o saobraćaju' in f and f.endswith('.xlsx')
    ]

    if not files:
        print(f"   ⚠️  No Excel file found in {month_folder}")
        return []

    file_path = os.path.join(month_path, files[0])
    flights = []

    try:
        print(f"📊 {month_folder}: ", end="")

        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)

        # Extract year and month from folder name (e.g., "01. JANUAR" -> 2025-01)
        month_number = month_folder.split('.')[0].strip().zfill(2)  # "01", "02", etc.
        year = "2025"  # Hardcoded for 2025 folder

        # Iterate through all sheets (days)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            # Extract day from sheet name (e.g., "01", "1", "09" -> "01")
            day_str = re.match(r'^(\d+)', sheet_name.strip())
            if not day_str:
                continue  # Skip sheets that don't start with a number

            day = day_str.group(1).zfill(2)  # "01", "02", etc.

            # Construct proper date string
            proper_date_str = f"{year}-{month_number}-{day}"

            for i, row in enumerate(ws.iter_rows(values_only=True), 1):
                # Skip header
                if i == 1:
                    continue

                # Skip empty rows
                if not row[0]:
                    continue

                try:
                    # Extract data from row
                    # [0] datum, [1] kompanija, [2] ruta, [3] tip a/c, [4] reg, [5] tip OPER,
                    # [6] MTOW(kg), [7] br leta u dol, [8] br leta u odl,
                    # [9] pl vrijeme dol, [10] st vrijeme dol, [11] pl vrijeme odl, [12] st vrijeme odl,
                    # [13] br putnika DOLAZAK, [14] br putnika ODLAZAK,
                    # [15] prtljag D(kg), [16] prtljag O(kg),
                    # [17] cargo D(kg), [18] cargo O(kg),
                    # [19] pošta D(kg), [20] pošta O(kg)

                    date_value = row[0]
                    airline_name = row[1]
                    route_str = row[2]
                    aircraft_model = row[3]
                    registration = row[4]
                    operation_type_str = row[5]
                    mtow = row[6] if len(row) > 6 else None
                    arrival_flight_number = row[7] if len(row) > 7 else None
                    departure_flight_number = row[8] if len(row) > 8 else None
                    arrival_scheduled_time = row[9] if len(row) > 9 else None
                    arrival_actual_time = row[10] if len(row) > 10 else None
                    departure_scheduled_time = row[11] if len(row) > 11 else None
                    departure_actual_time = row[12] if len(row) > 12 else None
                    arrival_passengers = row[13] if len(row) > 13 else None
                    departure_passengers = row[14] if len(row) > 14 else None
                    arrival_baggage = row[15] if len(row) > 15 else None
                    departure_baggage = row[16] if len(row) > 16 else None
                    arrival_cargo = row[17] if len(row) > 17 else None
                    departure_cargo = row[18] if len(row) > 18 else None
                    arrival_mail = row[19] if len(row) > 19 else None
                    departure_mail = row[20] if len(row) > 20 else None

                    # Use proper date from sheet name instead of Excel cell
                    # (Excel cells have wrong years in some cases)
                    try:
                        flight_date = datetime.strptime(proper_date_str, '%Y-%m-%d').isoformat()
                    except:
                        # If parsing fails, skip this row
                        continue

                    # Parse route
                    route = parse_route(str(route_str))
                    if not route:
                        continue

                    # Parse passengers (with infant support)
                    arrival_pax = parse_passengers(arrival_passengers)
                    departure_pax = parse_passengers(departure_passengers)

                    # Create flight object
                    flight = {
                        'date': flight_date,
                        'airline': str(airline_name).strip().upper(),
                        'route': str(route_str),
                        'departureAirport': route['departure'],
                        'arrivalAirport': route['arrival'],
                        'aircraftModel': str(aircraft_model).strip(),
                        'registration': str(registration or ''),
                        'operationType': str(operation_type_str).strip().upper() if operation_type_str else 'N/A',

                        # Additional data
                        'mtow': int(mtow) if mtow and str(mtow).strip() and str(mtow).strip() != '-' else None,
                        'arrivalFlightNumber': str(arrival_flight_number) if arrival_flight_number else None,
                        'departureFlightNumber': str(departure_flight_number) if departure_flight_number else None,

                        # Passengers (parsed)
                        'arrivalPassengers': arrival_pax['adults'] if arrival_pax else None,
                        'arrivalInfants': arrival_pax['infants'] if arrival_pax else None,
                        'departurePassengers': departure_pax['adults'] if departure_pax else None,
                        'departureInfants': departure_pax['infants'] if departure_pax else None,

                        # Baggage, cargo, mail
                        'arrivalBaggage': int(arrival_baggage) if arrival_baggage and str(arrival_baggage).strip() != '-' else None,
                        'departureBaggage': int(departure_baggage) if departure_baggage and str(departure_baggage).strip() != '-' else None,
                        'arrivalCargo': int(arrival_cargo) if arrival_cargo and str(arrival_cargo).strip() != '-' else None,
                        'departureCargo': int(departure_cargo) if departure_cargo and str(departure_cargo).strip() != '-' else None,
                        'arrivalMail': int(arrival_mail) if arrival_mail and str(arrival_mail).strip() != '-' else None,
                        'departureMail': int(departure_mail) if departure_mail and str(departure_mail).strip() != '-' else None,

                        'sourceFile': files[0],
                        'sheet': sheet_name,
                    }

                    flights.append(flight)

                except Exception as e:
                    # Skip problematic rows silently
                    pass

        print(f"✅ {len(flights)} flights")
        wb.close()
        return flights

    except Exception as e:
        print(f"❌ Error: {str(e)}")
        return []

def main():
    """Main function"""
    args = sys.argv[1:]
    month_arg = None

    for arg in args:
        if not arg.startswith('--'):
            month_arg = arg
            break

    print('🚀 Extracting 2025 Flight Data to JSON...\n')

    all_flights = []

    if month_arg:
        # Extract specific month
        print(f"📅 Extracting single month: {month_arg}\n")
        flights = extract_month(month_arg)
        all_flights.extend(flights)
    else:
        # Extract all months
        month_folders = sorted([
            f for f in os.listdir(STATS_DIR)
            if os.path.isdir(os.path.join(STATS_DIR, f))
        ])

        print(f"📅 Found {len(month_folders)} months\n")

        for month_folder in month_folders:
            flights = extract_month(month_folder)
            all_flights.extend(flights)

    # Save to JSON
    output_dir = 'output'
    os.makedirs(output_dir, exist_ok=True)
    output_file = os.path.join(output_dir, '2025-flights-data.json')

    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump({
            'flights': all_flights,
            'totalCount': len(all_flights),
            'extractedAt': datetime.now().isoformat(),
        }, f, indent=2, ensure_ascii=False)

    print('\n' + '='*80)
    print('✅ EXTRACTION COMPLETED')
    print('='*80)
    print(f'\n📋 Summary:')
    print(f'   - Total flights: {len(all_flights)}')
    print(f'   - Output file: {output_file}\n')

if __name__ == '__main__':
    main()
