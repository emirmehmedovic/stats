#!/usr/bin/env python3

import openpyxl
import os

STATS_DIR = "/Users/emir_mw/stats/STATS/2025/Dnevni izvještaji"
month_folder = "07. JULI"
month_path = os.path.join(STATS_DIR, month_folder)

files = [
    f for f in os.listdir(month_path)
    if 'Dnevni izvještaj o saobraćaju' in f and f.endswith('.xlsx')
]

file_path = os.path.join(month_path, files[0])
wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)

print(f"📊 Analyzing July 2025 Excel file\n")
print(f"Total sheets: {len(wb.sheetnames)}")
print(f"Sheet names (first 10): {wb.sheetnames[:10]}\n")

# Check first sheet
sheet_name = wb.sheetnames[0]
ws = wb[sheet_name]

print(f"{'='*80}")
print(f"Sheet: {sheet_name}")
print('='*80)

# Header
header = next(ws.iter_rows(values_only=True))
print(f"\nHeader row:")
for j, cell in enumerate(header):
    if cell:
        print(f"  [{j:2d}] {cell}")

# First few data rows
print(f"\nFirst 5 data rows:")
row_count = 0
for i, row in enumerate(ws.iter_rows(values_only=True), 1):
    if i == 1:
        continue

    if row[0] and row_count < 5:
        row_count += 1
        print(f"\nRow {i}:")
        for j in range(min(15, len(row))):
            if row[j]:
                print(f"  [{j:2d}] {row[j]}")

wb.close()
