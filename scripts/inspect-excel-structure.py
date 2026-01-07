#!/usr/bin/env python3

import openpyxl
import os

STATS_DIR = "/Users/emir_mw/stats/STATS/2025/Dnevni izvještaji"

# Check January
month_folder = "01. JANUAR"
month_path = os.path.join(STATS_DIR, month_folder)

files = [
    f for f in os.listdir(month_path)
    if 'Dnevni izvještaj o saobraćaju' in f and f.endswith('.xlsx')
]

file_path = os.path.join(month_path, files[0])

wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)

# Get first sheet
first_sheet = wb.sheetnames[0]
ws = wb[first_sheet]

print(f"📊 Excel Structure - Sheet: {first_sheet}\n")

# Print header row
print("Header Row (Row 1):")
for i, row in enumerate(ws.iter_rows(values_only=True), 1):
    if i == 1:
        for col_idx, value in enumerate(row):
            if value:
                print(f"  Column {col_idx}: {value}")
        break

print("\n" + "="*80)
print("First 5 Data Rows:")
print("="*80 + "\n")

# Print first 5 data rows
for i, row in enumerate(ws.iter_rows(values_only=True), 1):
    if i == 1:  # Skip header
        continue
    if i > 6:  # Show only first 5 data rows
        break

    print(f"Row {i}:")
    for col_idx, value in enumerate(row):
        if value is not None and str(value).strip() != '':
            print(f"  Column {col_idx}: {value}")
    print()

wb.close()
