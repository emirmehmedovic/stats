#!/usr/bin/env python3

import openpyxl
import os

STATS_DIR = "/Users/emir_mw/stats/STATS/2025/Dnevni izvještaji"
month_path = os.path.join(STATS_DIR, "03. MART")

files = [
    f for f in os.listdir(month_path)
    if 'Dnevni izvještaj o saobraćaju' in f and f.endswith('.xlsx')
]

file_path = os.path.join(month_path, files[0])
wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)

print("Checking headers across March sheets:\n")

for sheet_name in ['01', '02', '03', 'Sheet1']:
    if sheet_name not in wb.sheetnames:
        continue

    ws = wb[sheet_name]
    header = next(ws.iter_rows(values_only=True))

    print(f"{'='*80}")
    print(f"Sheet: {sheet_name}")
    print('='*80)

    # Show columns 0-10
    for i in range(min(11, len(header))):
        print(f"  [{i:2d}] {header[i] if header[i] else '(empty)'}")

    # Show first data row
    for j, row in enumerate(ws.iter_rows(values_only=True)):
        if j == 0:
            continue
        if row[0]:
            print(f"\n  First data row:")
            for i in range(min(11, len(row))):
                print(f"    [{i:2d}] {row[i]}")
            break

    print()

wb.close()
