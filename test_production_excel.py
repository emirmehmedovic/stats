#!/usr/bin/env python3
"""
Test Excel creation on production to compare with dev
"""
import sys
import os
from pathlib import Path

print("=" * 60)
print("PRODUCTION EXCEL CREATION TEST")
print("=" * 60)

print(f"\nPython version: {sys.version}")
print(f"Python executable: {sys.executable}")

try:
    import openpyxl
    print(f"openpyxl version: {openpyxl.__version__}")
    print(f"openpyxl location: {openpyxl.__file__}")
except ImportError as e:
    print(f"ERROR: Cannot import openpyxl: {e}")
    sys.exit(1)

from openpyxl.styles import Font, Alignment

# Get project root
PROJECT_ROOT = Path(__file__).parent
OUTPUT_DIR = PROJECT_ROOT / "izvještaji" / "generated"
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

print(f"\nOutput directory: {OUTPUT_DIR}")

# Test 1: Absolute minimal
print("\n--- Test 1: Creating absolute minimal Excel ---")
try:
    wb1 = openpyxl.Workbook()
    ws1 = wb1.active
    ws1['A1'] = "Production Test"
    ws1['B1'] = 123
    wb1.iso_dates = True

    output1 = OUTPUT_DIR / "prod_test_minimal.xlsx"
    wb1.save(output1)
    wb1.close()

    # Check file size
    file_size = os.path.getsize(output1)
    print(f"✅ Created: {output1}")
    print(f"   File size: {file_size} bytes")

    # Check file permissions
    import stat
    file_stat = os.stat(output1)
    print(f"   Permissions: {oct(file_stat.st_mode)}")
    print(f"   Owner UID: {file_stat.st_uid}")
    print(f"   Group GID: {file_stat.st_gid}")

except Exception as e:
    print(f"❌ ERROR: {e}")
    import traceback
    traceback.print_exc()

# Test 2: With styling and merged cells
print("\n--- Test 2: Creating with styling and merged cells ---")
try:
    wb2 = openpyxl.Workbook()
    ws2 = wb2.active
    ws2.title = "Production Test"

    # Title (merged)
    title_font = Font(name='Arial', size=14, bold=True)
    center_align = Alignment(horizontal='center', vertical='center')

    for col in range(1, 6):
        cell = ws2.cell(row=1, column=col)
        if col == 1:
            cell.value = "PRODUCTION TEST REPORT"
        cell.font = title_font
        cell.alignment = center_align

    ws2.merge_cells('A1:E1')

    # Add some data
    headers = ['ID', 'Name', 'Value', 'Status']
    for col_idx, header in enumerate(headers, start=1):
        cell = ws2.cell(row=2, column=col_idx)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = center_align

    # Data rows
    for row_idx in range(3, 8):
        ws2.cell(row=row_idx, column=1).value = row_idx - 2
        ws2.cell(row=row_idx, column=2).value = f"Item {row_idx - 2}"
        ws2.cell(row=row_idx, column=3).value = 100 + row_idx
        ws2.cell(row=row_idx, column=4).value = "OK"

    wb2.iso_dates = True

    output2 = OUTPUT_DIR / "prod_test_styled.xlsx"
    wb2.save(output2)
    wb2.close()

    # Check file size
    file_size2 = os.path.getsize(output2)
    print(f"✅ Created: {output2}")
    print(f"   File size: {file_size2} bytes")

except Exception as e:
    print(f"❌ ERROR: {e}")
    import traceback
    traceback.print_exc()

print("\n" + "=" * 60)
print("TEST COMPLETE")
print("=" * 60)
print("\nNEXT STEPS:")
print("1. Download both files from production")
print("2. Open them in Excel on your Mac")
print("3. Compare with dev-generated files")
print("4. Report if production files have corruption warning")
print("=" * 60)
