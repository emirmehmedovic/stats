#!/usr/bin/env python3
"""
Test Excel creation with verbose warnings
"""
import warnings
import sys
warnings.simplefilter('always')  # Show all warnings

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from pathlib import Path

OUTPUT_DIR = Path("izvještaji/generated")
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

print("openpyxl version:", openpyxl.__version__)
print("\n" + "="*60)

# Test creating a typical report structure
print("Creating typical report structure...")
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Test Report"

# Title row (merged)
title_font = Font(name='Arial', size=14, bold=True)
center_align = Alignment(horizontal='center', vertical='center')

# Method 1: Old way (what we're doing)
print("Setting up merged cell the OLD way...")
for col in range(1, 6):
    cell = ws.cell(row=1, column=col)
    if col == 1:
        cell.value = "TEST REPORT TITLE"
    cell.font = title_font
    cell.alignment = center_align
ws.merge_cells('A1:E1')

# Headers
header_font = Font(name='Arial', size=10, bold=True)
header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

headers = ['ID', 'Name', 'Value', 'Date', 'Status']
for col_idx, header in enumerate(headers, start=1):
    cell = ws.cell(row=2, column=col_idx)
    cell.value = header
    cell.font = header_font
    cell.fill = header_fill
    cell.border = border
    cell.alignment = center_align

# Data rows
for row_idx in range(3, 10):
    ws.cell(row=row_idx, column=1).value = row_idx - 2
    ws.cell(row=row_idx, column=2).value = f"Item {row_idx - 2}"
    ws.cell(row=row_idx, column=3).value = 100 + row_idx
    ws.cell(row=row_idx, column=4).value = "2026-01-01"
    ws.cell(row=row_idx, column=5).value = "OK"

# Save with all best practices
print("Saving with best practices...")
wb.iso_dates = True

output_path = OUTPUT_DIR / "test_typical_report.xlsx"
wb.save(output_path)
wb.close()

print(f"✅ Saved: {output_path}")
print("\n" + "="*60)
print("MOLIM VAS:")
print("1. Otvorite test_typical_report.xlsx")
print("2. Javite da li ima Excel corruption warning")
print("3. Provjerite i absolute_minimal.xlsx i minimal_with_styling.xlsx")
print("="*60)
